from django.shortcuts import render, redirect, get_object_or_404
from django.utils.html import escape
from django.contrib import messages
from Feedbacksystem.models import Course, Section, SectionSubjectFaculty, Faculty, Department, Student, Subject, LikertEvaluation, EvaluationStatus, Event, SchoolEventModel, FacultyEvaluationQuestions, SchoolEventQuestions, WebinarSeminarModel, WebinarSeminarQuestions, StakeholderFeedbackModel, StakeholderFeedbackQuestions, Message, PeertoPeerEvaluation, PeertoPeerEvaluationQuestions, StakeholderAgency, Attendance 
from .forms import TeacherForm, StudentForm, CourseForm, SectionForm, SectionSubjectFacultyForm, SubjectForm, StudentRegistrationForm, StudentLoginForm, LikertEvaluationForm, FacultyRegistrationForm, FacultyLoginForm, EvaluationStatusForm, DepartmentForm, EventCreationForm, SchoolEventForm, WebinarSeminarForm, StudentProfileForm, EditQuestionForm, EditSchoolEventQuestionForm,  WebinarSeminarForm, EditWebinarSeminarQuestionForm, StakeholderFeedbackForm, MessageForm, PeertoPeerEvaluationForm, FacultyProfileForm, AdminRegistrationForm, EditStakeholdersQuestionForm
from django.contrib.auth.models import User, Group
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import UserCreationForm
from django.db.models import Avg, Value, CharField
from django.http import Http404
from .filters import EvaluationFilter, FacultyFilter, StudentFilter, UserFilter, EventFilter, SectionFilter, SubjectFilter, StakeholderFilter, LikertEvaluationFilter, PeertoPeerEvaluationFilter, SchoolEventFilter, WebinarSeminarEventFilter
from .resources import StudentResource
from tablib import Dataset
from django.contrib.auth.decorators import login_required
from .decorators import allowed_users
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, F, Window, Min
import json
from django.core.exceptions import ObjectDoesNotExist
import csv
from django.utils import timezone
from django.db.models.functions import Concat, Rank
import qrcode
from django.utils.timezone import now
from datetime import timedelta, datetime
from collections import defaultdict
from statistics import mean
from django.db import models  
from django.core.mail import EmailMessage
from notifications.signals import notify
from notifications.models import Notification
from django.utils.text import Truncator
from django.contrib.auth import views as auth_views
from django.utils.http import urlsafe_base64_decode
from django.contrib.auth.tokens import default_token_generator
from django.contrib.auth.models import User
from django.template.loader import render_to_string
from django.contrib.sites.shortcuts import get_current_site
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.core.mail import EmailMessage
from django.contrib.auth import get_user_model
from .tokens import account_activation_token
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.date import DateTrigger
from xhtml2pdf import pisa
from django.templatetags.static import static
import os
from django.conf import settings
from itertools import chain
from django.db.models import Q, FloatField, Count, Case, When, IntegerField
from django.http import FileResponse
from openpyxl import Workbook 
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
import calendar
from xhtml2pdf.default import DEFAULT_FONT
from .jobs import close_evaluations, approve_pending_evaluations, start_event_evaluations, end_event_evaluations
from apscheduler.jobstores.base import JobLookupError
from .scheduler import scheduler
from django.db.models.functions import Coalesce, Round
from .utils import generate_wordcloud, get_sentiment
ITEMS_PER_PAGE = 10
            # ------------------------------------------------------
            #                Login-Page Views
            # ------------------------------------------------------
def signin(request):
    registration_form = StudentRegistrationForm()
    login_form = StudentLoginForm()

    if request.method == 'POST':

        if 'signin' in request.POST:
            # Handle login form submission
            login_form = StudentLoginForm(request.POST)
            if login_form.is_valid():
               student_number = login_form.cleaned_data['student_number']
               password = login_form.cleaned_data['password']

            # Authenticate the user
            user = authenticate(request, username=student_number, password=password)

            if user is not None:
                # Login the user
                login(request, user)
                # Get the value of the 'next' parameter
                next_url = request.GET.get('next')

                # Redirect to 'next' URL if it exists, otherwise redirect to the home page
                if next_url:
                    return redirect(next_url)
                else:
                    return redirect('home')
                
            else:

                try:
                        user = User.objects.get(username=student_number)
                        if user.is_active == False:
                            messages.error(request, 'Your account is not activated. Please check your email to activate your account.')
                        else:
                            messages.error(request, 'Invalid password.')
                except User.DoesNotExist:
                        messages.error(request, 'Student number is not registered.')
                
                
               
              
        elif 'register' in request.POST:
            # Handle registration form submission
            registration_form = StudentRegistrationForm(request.POST)
            if registration_form.is_valid():
                    # Get the student number from the form
                student_number = registration_form.cleaned_data.get('student_number')

                # Fetch the student record from the database
                student = get_object_or_404(Student, student_number=student_number)

                # Create a new user and associate it with the student
                user = registration_form.save(commit=False)
                user.email = student.email  # Automatically set the email
                user.is_active = False  # Deactivate account until email verification
                user.save()

                # Link the user to the student model
                student.user = user
                student.save()

                # Assign the user to the "student" group
                group = Group.objects.get(name='student')
                user.groups.add(group)

                # Send email activation link
                activateEmail(request, user, student.email)
                return redirect('signin')  # Redirect to the home page or any desired URL after successful registration
            else:
                for field, errors in registration_form.errors.items(): 
                     for error in errors:
                          if 'password' in field: 
                               if 'match' in error: messages.error(request, 'Passwords do not match.') 
                               elif 'too short' in error: messages.error(request, 'Password is too short. It must contain at least 8 characters.') 
                          elif field == 'student_number': 
                              if 'exist' in error: 
                                messages.error(request, 'Student with this student number does not exist. Please contact the admin to create an account.') 
                              elif 'registered' in error:
                                messages.error(request, 'This student number is already registered. Please login to continue.')
                return render(request, 'pages/login.html', {'registration_form': registration_form, 'login_form': login_form,})
                                     

    context = {'registration_form': registration_form, 'login_form': login_form,}
    
    return render(request, 'pages/login.html', context)

def activateEmail(request, user, to_email):
    mail_subject = 'Activate your user account.'
    message = render_to_string('pages/template_activate_account.html', {
        'user': user.username,
        'domain': get_current_site(request).domain,
        'uid': urlsafe_base64_encode(force_bytes(user.pk)),
        'token': account_activation_token.make_token(user),
        'protocol': 'https' if request.is_secure() else 'http'
    })
    email = EmailMessage(mail_subject, message, to=[to_email])
    if email.send():
        messages.success(request, f'A confirmation email has been sent to {to_email}. Please check your inbox and click the activation link to complete your registration. Note: If you do not see the email, please check your spam folder.')
    else:
        messages.error(request, f'Problem sending confirmation email to {to_email}, check if you typed it correctly.')

def activate(request, uidb64, token):
    User = get_user_model()
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
        
    except(TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    if user is not None and account_activation_token.check_token(user, token):
        user.is_active = True
        user.save()

        messages.success(request, 'Thank you for confirming your email. You can now log in to your account.')
        return redirect('signin')
    else:
        print(user)
        messages.error(request, 'Activation link is invalid!')
    
    return redirect('signin')
  
def adminlogin(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        try:
            user = User.objects.get(username=username)
        except User.DoesNotExist:
            messages.error(request, "Username does not exist")
            return render(request, 'pages/adminlogin.html', {})

        # Authenticate the user using username and password
        user = authenticate(request, username=username, password=password)
        if user is not None:
            if user.groups.filter(name="HR admin").exists():
                login(request, user)
                return redirect('hr_dashboard')
            elif user.groups.filter(name="admin").exists():
                login(request, user)
                return redirect('admin')
            else:
                messages.error(request, "You are not authorized to log in as an admin")
        else:
            messages.error(request, "Invalid password")

    context = {}
    return render(request, 'pages/adminlogin.html', context)



def facultylogin(request):
    registration_form = FacultyRegistrationForm()
    login_form = FacultyLoginForm()

    if request.method == 'POST':

        if 'signin' in request.POST:
            # Handle login form submission
            login_form = FacultyLoginForm(request.POST)
            if login_form.is_valid():
               email = login_form.cleaned_data['email']
               password = login_form.cleaned_data['password']

            # Authenticate the user
            user = authenticate(request, username=email, password=password)

            if user is not None:
                login(request, user)
                next_url = request.GET.get('next')

                # Redirect to 'next' URL if it exists, otherwise redirect to the home page
                if next_url:
                    return redirect(next_url)
                else:
                    return redirect('facultydashboard')
            else:

                try:
                        user = User.objects.get(username=email)
                        if user.is_active == False:
                            messages.error(request, 'Your account is not activated. Please check your email to activate your account.')
                        else:
                            messages.error(request, 'Invalid password.')
                except User.DoesNotExist:
                        messages.error(request, 'Your email is not registered.')
                      
        elif 'register' in request.POST:
            # Handle registration form submission
            registration_form = FacultyRegistrationForm(request.POST)
            if registration_form.is_valid():
                user = registration_form.save(commit=False)
                user.is_active = False
                user.save()

                email = registration_form.cleaned_data.get('email')
                faculty = Faculty.objects.get(email=email)
                faculty.user = user
                faculty.save()  # Save the faculty with the linked user

                group = Group.objects.get(name = 'faculty')
                user.groups.add(group)
                activateEmail(request, user, email)
                return redirect('facultylogin')  # Redirect to the home page or any desired URL after successful registration
            else:
               for field, errors in registration_form.errors.items(): 
                     for error in errors:
                          if 'password' in field: 
                               if 'match' in error: messages.error(request, 'Passwords do not match.') 
                               elif 'too short' in error: messages.error(request, 'Password is too short. It must contain at least 8 characters.') 
                               elif 'similar' in error: messages.error(request, 'The password is too similar to the email address.') 
                          elif field == 'email': 
                              if 'exist' in error: 
                                messages.error(request, 'Faculty with this email does not exist. Please contact the admin to create an account.') 
                              elif 'registered' in error:
                                messages.error(request, 'This email is already registered. Please login to continue.')
                                return redirect('facultylogin') 

    context = {'registration_form': registration_form, 'login_form': login_form,}
    
    return render(request, 'pages/facultylogin.html', context)

class CustomPasswordResetConfirmView(auth_views.PasswordResetConfirmView):
    template_name = 'pages/password_reset_form.html'
    
    def dispatch(self, *args, **kwargs):
        # Attempt to decode the user ID and validate the token
        try:
            uid = urlsafe_base64_decode(kwargs['uidb64']).decode()
            user = User.objects.get(pk=uid)
            token = kwargs['token']
            if not default_token_generator.check_token(user, token):
                raise ValueError("Invalid token")
        except:
            pass
        # Proceed with default behavior if token is valid
        return super().dispatch(*args, **kwargs)

    
def stakeholder_feedback_form(request):
    questions = StakeholderFeedbackQuestions.objects.all()
    form = StakeholderFeedbackForm()
    if request.method == 'POST':
        form = StakeholderFeedbackForm(request.POST)
        if form.is_valid():

            name = form.cleaned_data['name']
            agency_name = form.cleaned_data['agency']
            email = form.cleaned_data['email']
            purpose = form.cleaned_data['purpose']
            date = form.cleaned_data['date']
            staff = form.cleaned_data['staff']
            courtesy = form.cleaned_data['courtesy']
            quality = form.cleaned_data['quality']
            timeliness = form.cleaned_data['timeliness']
            efficiency = form.cleaned_data['efficiency']
            cleanliness = form.cleaned_data['cleanliness']
            comfort = form.cleaned_data['comfort']
            suggestions_and_comments = form.cleaned_data['suggestions_and_comments']
            predicted_sentiment = get_sentiment(suggestions_and_comments)
            sentiment_label = 'None'
            if predicted_sentiment and isinstance(predicted_sentiment, list):
                try:
                    # Access the first list in the response, then sort by score
                    sentiment_list = predicted_sentiment[0]
                    # Get the prediction with the highest score (sorted descending)
                    top_prediction = max(sentiment_list, key=lambda x: x['score'])
                    
                    # Directly use the label from the API response
                    sentiment_label = top_prediction['label'].capitalize()  # "Negative", "Neutral", or "Positive"
                
                except (IndexError, KeyError, TypeError) as e:
                    print(f"Error processing sentiment: {e}")
                    sentiment_label = "None"


            agency_instance = get_object_or_404(StakeholderAgency, name=agency_name)

            form = StakeholderFeedbackModel(
                name=name,
                agency=agency_instance,
                email=email,
                purpose=purpose,
                date=date,
                staff=staff,
                courtesy=courtesy,
                quality=quality,
                timeliness=timeliness,
                efficiency=efficiency,
                cleanliness=cleanliness,
                comfort=comfort,
                suggestions_and_comments=suggestions_and_comments,
                predicted_sentiment = sentiment_label
            )           
            form.save()
            messages.success(request, 'Form Submitted Successfully')
            return redirect('signin')
    context = {'questions': questions, 'form': form}

    return render(request, 'pages/stakeholder_feedback_form.html', context)

def makeqrcode(destination_url):
    image = qrcode.make(destination_url)
    name = "CvSU_Stakeholder_Feedback_Form.jpg"
    path = 'Feedbacksystem/static/images/'  
    image.save(path + name)

def get_code(request):
    if request.method == 'GET':
        destination_url = "https://feedback-and-evaluation-system-capstone.onrender.com/stakeholder_feedback_form"
        makeqrcode(destination_url)

        return render(request, 'pages/qrcode.html')


            # ------------------------------------------------------
            #                Student-Page Views
            # ------------------------------------------------------


@login_required(login_url='signin')
@allowed_users(allowed_roles=['student', 'society president'])
def home(request):
    user = request.user
    is_president = request.user.groups.filter(name='society president').exists()  # Check if the user is in the "student" group
    student = Student.objects.filter(student_number=request.user.username).first()
    event_notifications = Notification.objects.filter(recipient=user, level='success')[:5] 
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    notifications_unread_count = unread_notifications.count()
    courses = student.Course 
    evaluation_status = EvaluationStatus.objects.first()
    is_irregular = Student.objects.filter( Q(student_number=request.user.username) & (Q(status='IRREGULAR') | Q(status='Irregular')) ).first()
    #pending events to display
    events = Event.objects.filter(course_attendees=courses, admin_status="Approved").distinct()  # Get events related to those courses
    # Get evaluated event IDs by the current user
     # Get evaluated event IDs by the current user for SchoolEventModel
    evaluated_school_event_ids = SchoolEventModel.objects.filter(user=user).values_list('event_id', flat=True)
    
    # Get evaluated event IDs by the current user for WebinarSeminarModel
    evaluated_webinar_event_ids = WebinarSeminarModel.objects.filter(user=user).values_list('event_id', flat=True)

    # Combine evaluated event IDs from both models
    evaluated_event_ids = list(evaluated_school_event_ids) + list(evaluated_webinar_event_ids)
    
    # Exclude events that have been evaluated
    current_time = timezone.now()
    upcoming_events = events.filter(date__gt=current_time, evaluation_status='Closed')  
    past_events = events.filter(date__lt=current_time, evaluation_status='Closed')  
    unevaluated_events = events.exclude(id__in=evaluated_event_ids).exclude(id__in=past_events).exclude(id__in=upcoming_events) 
    for event in unevaluated_events:
        event.attended = Attendance.objects.filter(user=user, event=event, attended=True).exists()

    #faculty evaluated count
       # Fetch section subjects faculty based on the student's section
    section_subjects_faculty = SectionSubjectFaculty.objects.filter(section=student.Section)

    # Count the total number of unique faculty members
    total_faculty = section_subjects_faculty.values('faculty').distinct().count()
   
    pending_count = 0
    completed_count = 0

    # Iterate over each section_subject_faculty and check if evaluation exists
    for section_subject_faculty in section_subjects_faculty:
        try:
            # Check if the user has submitted an evaluation for this faculty
            evaluation_status = EvaluationStatus.objects.first()
            LikertEvaluation.objects.get(user=request.user, section_subject_faculty=section_subject_faculty, academic_year=evaluation_status.academic_year,
            semester=evaluation_status.semester)
            completed_count += 1  # Increment completed count if evaluation exists
        except:
            pass 
      
    #pagination
    paginator = Paginator(unevaluated_events, 3) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {'evaluation_status': evaluation_status, 'student': student, 'unevaluated_events': unevaluated_events, 'page_obj': page_obj, 'completed_count': completed_count, 'total_faculty': total_faculty, 'is_president': is_president, 'event_notifications': event_notifications, 'notifications_unread_count': notifications_unread_count, 'is_irregular': is_irregular}
    return render(request, 'pages/home.html', context)
    
@login_required(login_url='signin')
@allowed_users(allowed_roles=['student'])
def student_notifications(request):
    user=request.user
    is_president = request.user.groups.filter(name='society president').exists()  # Check if the user is in the "student" group
    student = Student.objects.filter(student_number=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    notifications_unread_count = unread_notifications.count()
    notifications = Notification.objects.filter(recipient=user, level='success')
  
    return render(request, 'pages/student_notifications.html', {'student': student, 'is_president': is_president, 'notifications': notifications, 'event_notifications': event_notifications,
        'notifications_unread_count': notifications_unread_count,  })  # Return the response

@login_required(login_url='signin')
@allowed_users(allowed_roles=['student', 'society president'])
def clear_student_notifications(request):
    user=request.user
    is_president = request.user.groups.filter(name='society president').exists()  # Check if the user is in the "student" group
    student = Student.objects.filter(student_number=request.user.username).first()   
    notifications = Notification.objects.filter(recipient=user, level='success')
    notifications.delete()
    return redirect('home') # Return the response

@login_required(login_url='signin')
@allowed_users(allowed_roles=['student', 'society president'])
def student_profile(request):
    user=request.user
    is_president = request.user.groups.filter(name='society president').exists()  # Check if the user is in the "student" group
    student = Student.objects.filter(student_number=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    notifications_unread_count = unread_notifications.count()
    user_groups = user.groups.all()
    all_student = Student.objects.all()
    # Example of checking if the user is in a specific group, e.g., "society president"
    is_society_president = user_groups.filter(name="society president").exists()
    is_irregular = Student.objects.filter( Q(student_number=request.user.username) & (Q(status='IRREGULAR') | Q(status='Irregular')) ).first()
    is_regular = Student.objects.filter( Q(student_number=request.user.username) & (Q(status='REGULAR') | Q(status='Regular')) ).first()  

    evaluation_status = EvaluationStatus.objects.first()
    section_subjects_faculty = SectionSubjectFaculty.objects.filter(section=student.Section)
    current_academic_year = evaluation_status.academic_year 
    current_semester = evaluation_status.semester
    
        # Create an empty dictionary to store evaluation status for each faculty
    evaluation_status_list = []

    # Iterate over each section_subject_faculty
    for section_subject_faculty in section_subjects_faculty:
        try:
            # Check if the user has submitted an evaluation for this faculty  
            evaluation = LikertEvaluation.objects.get(user=request.user, section_subject_faculty=section_subject_faculty, academic_year=evaluation_status.academic_year,
            semester=evaluation_status.semester)
            evaluation_status_list.append((section_subject_faculty.subjects, section_subject_faculty.faculty, evaluation.status))

        except LikertEvaluation.DoesNotExist:
            # If evaluation doesn't exist, set status to 'Pending'
            evaluation_status_list.append((section_subject_faculty.subjects, section_subject_faculty.faculty, 'Pending'))
    
    student_recent_faculty_evaluations = LikertEvaluation.objects.filter(user=user, academic_year=current_academic_year, semester=current_semester).order_by('-updated')

    school_event_evaluations = SchoolEventModel.objects.filter(
        user=user,
        academic_year=current_academic_year,
        semester=current_semester
    ).order_by('-updated')
    
    # Query evaluations for WebinarSeminarModel
    webinar_seminar_evaluations = WebinarSeminarModel.objects.filter(
        user=user,
        academic_year=current_academic_year,
        semester=current_semester
    ).order_by('-updated')
    
    recent_evaluations = list(chain(
        student_recent_faculty_evaluations, 
        school_event_evaluations, 
        webinar_seminar_evaluations
    ))
            
    context = {'student': student, 'section_subjects_faculty': section_subjects_faculty, 'evaluation_status_list': evaluation_status_list, 'is_society_president': is_society_president, 'is_president': is_president, 'event_notifications': event_notifications, 'notifications_unread_count': notifications_unread_count, 'recent_evaluations': recent_evaluations, 'current_semester': current_semester, 'is_irregular': is_irregular, 'is_regular': is_regular}
    return render(request, 'pages/student_profile.html', context)

def student_edit_evaluation_form(request, pk):
    user=request.user
    is_president = request.user.groups.filter(name='society president').exists()  # Check if the user is in the "student" group
    student = Student.objects.filter(student_number=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    notifications_unread_count = unread_notifications.count()
    # Check which model the evaluation belongs to
    evaluation = None
    form_class = None  # To store the appropriate form class
    questions = None
    form_type = None
    
    if LikertEvaluation.objects.filter(pk=pk).exists():
        evaluation = get_object_or_404(LikertEvaluation, pk=pk)
        form_class = LikertEvaluationForm
        questions = FacultyEvaluationQuestions.objects.all().order_by('order')
        form_type = 'SET'
    elif SchoolEventModel.objects.filter(pk=pk).exists():
        evaluation = get_object_or_404(SchoolEventModel, pk=pk)
        form_class = SchoolEventForm
        questions = SchoolEventQuestions.objects.all().order_by('order')
        form_type = 'School Event'
    elif WebinarSeminarModel.objects.filter(pk=pk).exists():
        evaluation = get_object_or_404(WebinarSeminarModel, pk=pk)
        form_class = WebinarSeminarForm
        questions = WebinarSeminarQuestions.objects.all().order_by('order')
        form_type = 'Webinar/Seminar Event'
    if evaluation is None or form_class is None:
        return render(request, '404.html', {"message": "Evaluation not found."})

    # Handle form submission
    if request.method == 'POST':
        # Initialize the form with POST data
        form = form_class(request.POST)
        if form.is_valid():
            # Update the instance manually with the cleaned data
            for field, value in form.cleaned_data.items():
                setattr(evaluation, field, value)
            evaluation.save()
            return redirect('student_profile')  # Redirect to recent evaluations page
    else:
        # Initialize the form with the instance's data
        initial_data = {field.name: getattr(evaluation, field.name) for field in evaluation._meta.fields}
        form = form_class(initial=initial_data)

    return render(request, 'pages/student_edit_evaluation_form.html', {'form': form, 'evaluation': evaluation, 'questions': questions, 'form_type': form_type, 'student': student, 'is_president': is_president, 'event_notifications': event_notifications, 'notifications_unread_count': notifications_unread_count,})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['student', 'society president'])
def edit_student_profile(request):
    user=request.user.student
    user_student = request.user
    is_president = request.user.groups.filter(name='society president').exists()  # Check if the user is in the "student" group
    student = Student.objects.filter(student_number=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user_student, level='success')
    unread_notifications = Notification.objects.filter(recipient=user_student, level='success', unread=True)
    notifications_unread_count = unread_notifications.count()
    form = StudentProfileForm(instance = user)
    if request.method == 'POST':
        form = StudentProfileForm(request.POST, request.FILES, instance = user)
        if form.is_valid():
            form.save(commit=True)  # Ensure commit is set to True to save to the database
            messages.success(request, 'Profile Updated Successfully')
            return redirect('student_profile')

    return render(request, 'pages/edit_student_profile.html', {'student': student, 'form': form, 'is_president': is_president, 'event_notifications': event_notifications, 'notifications_unread_count': notifications_unread_count})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['student', 'society president'])
def facultyeval(request):
    evaluation_status = EvaluationStatus.objects.first()
    user=request.user
    is_president = request.user.groups.filter(name='society president').exists()  # Check if the user is in the "student" group
    student = Student.objects.filter(student_number=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    notifications_unread_count = unread_notifications.count()
    is_irregular = Student.objects.filter( Q(student_number=request.user.username) & (Q(status='IRREGULAR') | Q(status='Irregular')) ).first()
    
    if is_irregular:
        section_subjects_faculty = (
            SectionSubjectFaculty.objects.order_by('faculty__pk')
            .distinct('faculty__pk')
        )
        section_subjects_faculty = sorted(
            section_subjects_faculty, key=lambda x: x.faculty.last_name
        )

    else:
        section_subjects_faculty = SectionSubjectFaculty.objects.filter(section=student.Section)

    evaluated_faculty_ids = LikertEvaluation.objects.filter(
    user=request.user,
    section_subject_faculty__in=section_subjects_faculty,
    academic_year=evaluation_status.academic_year,
    semester=evaluation_status.semester
    ).values_list('section_subject_faculty__faculty__pk', flat=True).distinct()  # Get evaluated faculty IDs


    return render(request, 'pages/facultyeval.html', {'student': student, 'section_subjects_faculty': section_subjects_faculty, 'evaluation_status': evaluation_status, 'is_president': is_president, 'event_notifications': event_notifications, 'notifications_unread_count': notifications_unread_count, 'evaluated_faculty_ids': evaluated_faculty_ids })

@login_required(login_url='signin')
@allowed_users(allowed_roles=['student', 'society president'])
def evaluate_subject_faculty(request,pk):
    section_subject_faculty = get_object_or_404(SectionSubjectFaculty, pk=pk)
    user=request.user
    is_president = request.user.groups.filter(name='society president').exists()  # Check if the user is in the "student" group
    student = Student.objects.filter(student_number=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    notifications_unread_count = unread_notifications.count()
    section_subjects_faculty = SectionSubjectFaculty.objects.filter(section=student.Section)
    questions = FacultyEvaluationQuestions.objects.all().order_by('order')
    user = request.user
    evaluation_status = EvaluationStatus.objects.first()
    user_evaluations = LikertEvaluation.objects.filter(
        user=request.user,
        section_subject_faculty=section_subject_faculty,
        academic_year=evaluation_status.academic_year,
        semester=evaluation_status.semester
    )

    evaluated_faculty_ids = LikertEvaluation.objects.filter(
    user=request.user,
    section_subject_faculty__in=section_subjects_faculty,
    academic_year=evaluation_status.academic_year,
    semester=evaluation_status.semester
    ).values_list('section_subject_faculty__faculty__pk', flat=True).distinct()  # Get evaluated faculty IDs

   
    if request.method == 'POST':
        form = LikertEvaluationForm(request.POST)
        if form.is_valid():
            # Process the evaluation form
            command_and_knowledge_of_the_subject = form.cleaned_data['command_and_knowledge_of_the_subject']
            depth_of_mastery = form.cleaned_data['depth_of_mastery']
            practice_in_respective_discipline = form.cleaned_data['practice_in_respective_discipline']
            up_to_date_knowledge = form.cleaned_data['up_to_date_knowledge']
            integrates_subject_to_practical_circumstances = form.cleaned_data['integrates_subject_to_practical_circumstances']

            organizes_the_subject_matter = form.cleaned_data['organizes_the_subject_matter']
            provides_orientation_on_course_content = form.cleaned_data['provides_orientation_on_course_content']
            efforts_of_class_preparation = form.cleaned_data['efforts_of_class_preparation']
            summarizes_main_points = form.cleaned_data['summarizes_main_points']
            monitors_online_class = form.cleaned_data['monitors_online_class']

            holds_interest_of_students = form.cleaned_data['holds_interest_of_students']
            provides_relevant_feedback = form.cleaned_data['provides_relevant_feedback']
            encourages_participation = form.cleaned_data['encourages_participation']
            shows_enthusiasm = form.cleaned_data['shows_enthusiasm']
            shows_sense_of_humor = form.cleaned_data['shows_sense_of_humor']

            teaching_methods = form.cleaned_data['teaching_methods']
            flexible_learning_strategies = form.cleaned_data['flexible_learning_strategies']
            student_engagement = form.cleaned_data['student_engagement']
            clear_examples = form.cleaned_data['clear_examples']
            focused_on_objectives = form.cleaned_data['focused_on_objectives']

            starts_with_motivating_activities = form.cleaned_data['starts_with_motivating_activities']
            speaks_in_clear_and_audible_manner = form.cleaned_data['speaks_in_clear_and_audible_manner']
            uses_appropriate_medium_of_instruction = form.cleaned_data['uses_appropriate_medium_of_instruction']
            establishes_online_classroom_environment = form.cleaned_data['establishes_online_classroom_environment']
            observes_proper_classroom_etiquette = form.cleaned_data['observes_proper_classroom_etiquette']

            uses_time_wisely = form.cleaned_data['uses_time_wisely']
            gives_ample_time_for_students_to_prepare = form.cleaned_data['gives_ample_time_for_students_to_prepare']
            updates_the_students_of_their_progress = form.cleaned_data['updates_the_students_of_their_progress']
            demonstrates_leadership_and_professionalism = form.cleaned_data['demonstrates_leadership_and_professionalism']
            understands_possible_distractions = form.cleaned_data['understands_possible_distractions']

            sensitivity_to_student_culture = form.cleaned_data['sensitivity_to_student_culture']
            responds_appropriately = form.cleaned_data['responds_appropriately']
            assists_students_on_concerns = form.cleaned_data['assists_students_on_concerns']
            guides_the_students_in_accomplishing_tasks = form.cleaned_data['guides_the_students_in_accomplishing_tasks']
            extends_consideration_to_students = form.cleaned_data['extends_consideration_to_students']


           # Extract the cleaned data from the form
            requires_less_task_for_credit = form.cleaned_data['requires_less_task_for_credit']
            


            strengths_of_the_faculty = form.cleaned_data['strengths_of_the_faculty']
            other_suggestions_for_improvement =  form.cleaned_data['other_suggestions_for_improvement']
            comments = form.cleaned_data['comments']
            predicted_sentiment = get_sentiment(comments)
            sentiment_label = 'None'
            if predicted_sentiment and isinstance(predicted_sentiment, list):
                try:
                    # Access the first list in the response, then sort by score
                    sentiment_list = predicted_sentiment[0]
                    # Get the prediction with the highest score (sorted descending)
                    top_prediction = max(sentiment_list, key=lambda x: x['score'])
                    
                    # Directly use the label from the API response
                    sentiment_label = top_prediction['label'].capitalize()  # "Negative", "Neutral", or "Positive"
                
                except (IndexError, KeyError, TypeError) as e:
                    print(f"Error processing sentiment: {e}")
                    sentiment_label = "None"

          
            # Save the data to the database
            # Create the LikertEvaluation instance
            evaluation_instance = LikertEvaluation(
                section_subject_faculty=section_subject_faculty,
                user=user,
                command_and_knowledge_of_the_subject=command_and_knowledge_of_the_subject,
                depth_of_mastery=depth_of_mastery,
                practice_in_respective_discipline=practice_in_respective_discipline,
                up_to_date_knowledge=up_to_date_knowledge,
                integrates_subject_to_practical_circumstances=integrates_subject_to_practical_circumstances,
                organizes_the_subject_matter=organizes_the_subject_matter,
                provides_orientation_on_course_content=provides_orientation_on_course_content,
                efforts_of_class_preparation=efforts_of_class_preparation,
                summarizes_main_points=summarizes_main_points,
                monitors_online_class=monitors_online_class,
                holds_interest_of_students=holds_interest_of_students,
                provides_relevant_feedback=provides_relevant_feedback,
                encourages_participation=encourages_participation,
                shows_enthusiasm=shows_enthusiasm,
                shows_sense_of_humor=shows_sense_of_humor,
                teaching_methods=teaching_methods,
                flexible_learning_strategies=flexible_learning_strategies,
                student_engagement=student_engagement,
                clear_examples=clear_examples,
                focused_on_objectives=focused_on_objectives,
                starts_with_motivating_activities=starts_with_motivating_activities,
                speaks_in_clear_and_audible_manner=speaks_in_clear_and_audible_manner,
                uses_appropriate_medium_of_instruction=uses_appropriate_medium_of_instruction,
                establishes_online_classroom_environment=establishes_online_classroom_environment,
                observes_proper_classroom_etiquette=observes_proper_classroom_etiquette,
                uses_time_wisely=uses_time_wisely,
                gives_ample_time_for_students_to_prepare=gives_ample_time_for_students_to_prepare,
                updates_the_students_of_their_progress=updates_the_students_of_their_progress,
                demonstrates_leadership_and_professionalism=demonstrates_leadership_and_professionalism,
                understands_possible_distractions=understands_possible_distractions,
                sensitivity_to_student_culture=sensitivity_to_student_culture,
                responds_appropriately=responds_appropriately,
                assists_students_on_concerns=assists_students_on_concerns,
                guides_the_students_in_accomplishing_tasks=guides_the_students_in_accomplishing_tasks,
                extends_consideration_to_students=extends_consideration_to_students,
                requires_less_task_for_credit=requires_less_task_for_credit,
                strengths_of_the_faculty=strengths_of_the_faculty,
                other_suggestions_for_improvement=other_suggestions_for_improvement,
                comments=comments,
                predicted_sentiment=sentiment_label
            )

            # Attempt to save the instance
            try:
                evaluation_instance.save()
                messages.success(request, 'Evaluation submitted successfully.')
                return redirect('facultyeval')  # Redirect to a success page
            except Exception as e:
                # If the save fails, print the error message and return a failure message
                print(f"Error saving evaluation: {e}")
                messages.error(request, 'Failed to submit evaluation. Please try again.')

        else:
            # If the form is invalid, print form errors
            print(form.errors)
            messages.error(request, 'There was an error with your submission. Please check your input.')

    else:
        form = LikertEvaluationForm()
    context = { 'form': form, 'section_subject_faculty': section_subject_faculty, 'section_subjects_faculty': section_subjects_faculty, 'student': student, 'questions': questions, 'user_evaluations':user_evaluations, 'is_president': is_president, 'event_notifications': event_notifications, 'notifications_unread_count': notifications_unread_count, 'evaluated_faculty_ids': evaluated_faculty_ids}
    return render(request, 'pages/evaluate_subject_faculty.html', context)


@login_required(login_url='signin')
@allowed_users(allowed_roles=['student', 'society president'])
def eventhub(request):
    user=request.user
    is_president = request.user.groups.filter(name='society president').exists()  # Check if the user is in the "student" group
    student = Student.objects.filter(student_number=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    notifications_unread_count = unread_notifications.count()
    courses = student.Course  # Get all courses the student is enrolled in
    events = Event.objects.filter(course_attendees=courses, admin_status='Approved').distinct()  # Get events related to those courses

    # Get evaluated event IDs by the current user
    # Get evaluated event IDs by the current user for SchoolEventModel
    evaluated_school_event_ids = SchoolEventModel.objects.filter(user=user).values_list('event_id', flat=True)
    
    # Get evaluated event IDs by the current user for WebinarSeminarModel
    evaluated_webinar_event_ids = WebinarSeminarModel.objects.filter(user=user).values_list('event_id', flat=True)

    # Combine evaluated event IDs from both models
    evaluated_event_ids = list(evaluated_school_event_ids) + list(evaluated_webinar_event_ids)
    # Get current time
    current_time = timezone.now()
    upcoming_events = events.filter(date__gt=current_time, evaluation_status='Closed')  # Events in the future
    # Past events with closed evaluation
    past_events = events.filter(date__lt=current_time, evaluation_status='Closed')  # Past events with evaluation closed   
    # Exclude events that have been evaluated
    unevaluated_events = events.exclude(id__in=evaluated_event_ids).exclude(id__in=past_events).exclude(id__in=upcoming_events).order_by('-date') 

    # Add attendance information to each event
    for event in unevaluated_events:
        event.attended = Attendance.objects.filter(user=user, event=event, attended=True).exists()
    return render(request, 'pages/eventhub.html', {'student': student, 'unevaluated_events': unevaluated_events, 'past_events': past_events, 'upcoming_events': upcoming_events, 'is_president': is_president, 'event_notifications': event_notifications, 'notifications_unread_count': notifications_unread_count})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['student', 'society president'])
def eventhub_upcoming(request):
    user=request.user
    is_president = request.user.groups.filter(name='society president').exists()  # Check if the user is in the "student" group
    student = Student.objects.filter(student_number=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    notifications_unread_count = unread_notifications.count()
    courses = student.Course  # Get all courses the student is enrolled in
    events = Event.objects.filter(course_attendees=courses, admin_status='Approved').distinct()  # Get events related to those courses
    
    current_time = timezone.now()
    upcoming_events = events.filter(date__gt=current_time, evaluation_status='Closed')  # Events in the future

    return render(request, 'pages/eventhub_upcoming.html', {'student': student,'upcoming_events': upcoming_events, 'is_president': is_president, 'event_notifications': event_notifications, 'notifications_unread_count': notifications_unread_count})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['student', 'society president'])
def eventhub_evaluated(request):
    user=request.user
    is_president = request.user.groups.filter(name='society president').exists()  # Check if the user is in the "student" group
    student = Student.objects.filter(student_number=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    notifications_unread_count = unread_notifications.count()
    courses = student.Course  # Get all courses the student is enrolled in
    events = Event.objects.filter(course_attendees=courses).distinct()  # Get events related to those courses
      # Get evaluated event IDs by the current user
     # Get evaluated event IDs by the current user for SchoolEventModel
    evaluated_school_event_ids = SchoolEventModel.objects.filter(user=user).values_list('event_id', flat=True)
    
    # Get evaluated event IDs by the current user for WebinarSeminarModel
    evaluated_webinar_event_ids = WebinarSeminarModel.objects.filter(user=user).values_list('event_id', flat=True)

    # Combine evaluated event IDs from both models
    evaluated_event_ids = list(evaluated_school_event_ids) + list(evaluated_webinar_event_ids)
    
    # Exclude events that have been evaluated
    evaluated_events = events.filter(id__in=evaluated_event_ids).order_by('-date')[:10]
    return render(request, 'pages/eventhub_evaluated.html', {'student': student, 'evaluated_events': evaluated_events, 'is_president': is_president, 'event_notifications': event_notifications, 'notifications_unread_count': notifications_unread_count})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['student', 'society president'])
def eventhub_closed(request):
    user=request.user
    is_president = request.user.groups.filter(name='society president').exists()  # Check if the user is in the "student" group
    student = Student.objects.filter(student_number=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    notifications_unread_count = unread_notifications.count()
    courses = student.Course  # Get all courses the student is enrolled in
    events = Event.objects.filter(course_attendees=courses).distinct()  # Get events related to those courses

    # Get current time
    current_time = timezone.now()
    # Past events with closed evaluation
    past_events = events.filter(date__lt=current_time, evaluation_status='Closed')  # Past events with evaluation closed   
    return render(request, 'pages/eventhub_closed.html', {'student': student, 'past_events': past_events, 'is_president': is_president, 'event_notifications': event_notifications, 'notifications_unread_count': notifications_unread_count})

@login_required(login_url='signin')
def scan_qr_code(request, pk):
    user = request.user
    event = get_object_or_404(Event, pk=pk)

    # Mark attendance
    attendance, created = Attendance.objects.get_or_create(user=user, event=event)
    attendance.attended = True
    attendance.save()

    # Redirect based on user role

    if user.groups.filter(Q(name='faculty') | Q(name='head of OSAS') | Q(name='department head/program coordinator')).exists():
        return redirect('faculty_event_detail', pk=pk)

    elif user.groups.filter(Q(name='student') | Q(name='society president')).exists():
        return redirect('event_detail', pk=pk)

    else:
        # Default redirection if no specific group is found
        return redirect('signin')
    
@login_required(login_url='signin')
@allowed_users(allowed_roles=['student', 'society president'])
def event_detail(request, pk):
    user=request.user
    is_president = request.user.groups.filter(name='society president').exists()  # Check if the user is in the "student" group
    student = Student.objects.filter(student_number=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    notifications_unread_count = unread_notifications.count()
    event = Event.objects.get(pk=pk)
    questions = SchoolEventQuestions.objects.all().order_by('order')

    # attendance = Attendance.objects.filter(user=user, event=event, attended=True).exists()
    #if not attendance: 
    #    messages.error(request, "You must scan the QR code to access this evaluation form.")
     #   return redirect('eventhub') # Redirect to home if user has not attended the event
    
    if event.event_type.name == 'School Event' or event.event_type.name == 'Training Workshop':
        user_evaluations = SchoolEventModel.objects.filter(
        user=request.user,
        event=event,
        )
        form = SchoolEventForm()
        if request.method == 'POST':
            form = SchoolEventForm(request.POST)
            if form.is_valid():
                 # Process the evaluation form
                meeting_expectation = form.cleaned_data['meeting_expectation']
                attainment_of_the_objectives = form.cleaned_data['attainment_of_the_objectives']
                topics_discussed = form.cleaned_data['topics_discussed']
                input_presentation = form.cleaned_data['input_presentation']
                management_team = form.cleaned_data['management_team']
                venue_and_physical_arrangement = form.cleaned_data['venue_and_physical_arrangement']
                overall_assessment = form.cleaned_data['overall_assessment']
                suggestions_and_comments = form.cleaned_data['suggestions_and_comments']
                predicted_sentiment = get_sentiment(suggestions_and_comments)
                sentiment_label = 'None'
                if predicted_sentiment and isinstance(predicted_sentiment, list):
                    try:
                        # Access the first list in the response, then sort by score
                        sentiment_list = predicted_sentiment[0]
                        # Get the prediction with the highest score (sorted descending)
                        top_prediction = max(sentiment_list, key=lambda x: x['score'])
                        
                        # Directly use the label from the API response
                        sentiment_label = top_prediction['label'].capitalize()  # "Negative", "Neutral", or "Positive"
                    
                    except (IndexError, KeyError, TypeError) as e:
                        print(f"Error processing sentiment: {e}")
                        sentiment_label = "None"
                # Save the data to database
                form = SchoolEventModel(
                    user=user,
                    event=event,
                    meeting_expectation=meeting_expectation,
                    attainment_of_the_objectives=attainment_of_the_objectives,
                    topics_discussed=topics_discussed,
                    input_presentation=input_presentation,
                    management_team=management_team,
                    venue_and_physical_arrangement=venue_and_physical_arrangement,
                    overall_assessment=overall_assessment,
                    suggestions_and_comments=suggestions_and_comments,
                    predicted_sentiment = sentiment_label
                )
                form.save()
            messages.success(request, 'Evaluation submitted successfully.')
            return redirect('eventhub')
        return render(request, 'pages/school_event_form.html', context = {'event': event, 'form': form, 'student': student, 'questions': questions, 'is_president': is_president, 'event_notifications': event_notifications, 'notifications_unread_count': notifications_unread_count, 'user_evaluations': user_evaluations})
    
    elif event.event_type.name == 'Webinar/Seminar':
        user_evaluations = WebinarSeminarModel.objects.filter(
        user=request.user,
        event=event,
        )
        questions = WebinarSeminarQuestions.objects.all().order_by('order')
        form = WebinarSeminarForm()  # Assuming WebinarSeminarForm is similar to SchoolEventForm
        if request.method == 'POST':
            form = WebinarSeminarForm(request.POST)
            if form.is_valid():
                # Process Webinar/Seminar evaluation
                relevance_of_the_activity = form.cleaned_data['relevance_of_the_activity']

                quality_of_the_activity = form.cleaned_data['quality_of_the_activity']

                timeliness = form.cleaned_data['timeliness']

                suggestions_and_comments = form.cleaned_data['suggestions_and_comments']

                attainment_of_the_objective = form.cleaned_data['attainment_of_the_objective']

                appropriateness_of_the_topic_to_attain_the_objective = form.cleaned_data['appropriateness_of_the_topic_to_attain_the_objective']

                appropriateness_of_the_searching_methods_used = form.cleaned_data['appropriateness_of_the_searching_methods_used']
                
                topics_to_be_included = form.cleaned_data['topics_to_be_included']

                appropriateness_of_the_topic_in_the_present_time = form.cleaned_data['appropriateness_of_the_topic_in_the_present_time']

                usefulness_of_the_topic_discusssed_in_the_activity = form.cleaned_data['usefulness_of_the_topic_discusssed_in_the_activity']

                appropriateness_of_the_searching_methods = form.cleaned_data['appropriateness_of_the_searching_methods']

                displayed_a_thorough_knowledge_of_the_topic = form.cleaned_data['displayed_a_thorough_knowledge_of_the_topic']

                explained_activities = form.cleaned_data['explained_activities']

                able_to_create_a_good_learning_environment = form.cleaned_data['able_to_create_a_good_learning_environment']

                able_to_manage_her_time_well = form.cleaned_data['able_to_manage_her_time_well']

                demonstrated_keenness_to_the_participant_needs = form.cleaned_data['demonstrated_keenness_to_the_participant_needs']

                timeliness_or_suitability_of_service = form.cleaned_data['timeliness_or_suitability_of_service']

                overall_satisfaction = form.cleaned_data['overall_satisfaction']

                predicted_sentiment = get_sentiment(suggestions_and_comments)
                sentiment_label = 'None'
                if predicted_sentiment and isinstance(predicted_sentiment, list):
                    try:
                        # Access the first list in the response, then sort by score
                        sentiment_list = predicted_sentiment[0]
                        # Get the prediction with the highest score (sorted descending)
                        top_prediction = max(sentiment_list, key=lambda x: x['score'])
                        
                        # Directly use the label from the API response
                        sentiment_label = top_prediction['label'].capitalize()  # "Negative", "Neutral", or "Positive"
                    
                    except (IndexError, KeyError, TypeError) as e:
                        print(f"Error processing sentiment: {e}")
                        sentiment_label = "None"

                # Save the data to the database
                form = WebinarSeminarModel(
                    user=user,
                    
                    event=event,

                    relevance_of_the_activity=relevance_of_the_activity,
                    
                    quality_of_the_activity=quality_of_the_activity,

                    timeliness=timeliness,

                    suggestions_and_comments=suggestions_and_comments,

                    attainment_of_the_objective=attainment_of_the_objective,

                    appropriateness_of_the_topic_to_attain_the_objective = appropriateness_of_the_topic_to_attain_the_objective,

                    appropriateness_of_the_searching_methods_used=appropriateness_of_the_searching_methods_used,

                    topics_to_be_included=topics_to_be_included,

                    appropriateness_of_the_topic_in_the_present_time=appropriateness_of_the_topic_in_the_present_time,

                    usefulness_of_the_topic_discusssed_in_the_activity=usefulness_of_the_topic_discusssed_in_the_activity,

                    appropriateness_of_the_searching_methods=appropriateness_of_the_searching_methods,

                    displayed_a_thorough_knowledge_of_the_topic=displayed_a_thorough_knowledge_of_the_topic,

                    explained_activities=explained_activities,

                    able_to_create_a_good_learning_environment=able_to_create_a_good_learning_environment,

                    able_to_manage_her_time_well=able_to_manage_her_time_well,

                    demonstrated_keenness_to_the_participant_needs=demonstrated_keenness_to_the_participant_needs,

                    timeliness_or_suitability_of_service=timeliness_or_suitability_of_service,

                    overall_satisfaction=overall_satisfaction,

                    predicted_sentiment = sentiment_label
                )
                form.save()
                messages.success(request, 'Evaluation submitted successfully.')
                return redirect('eventhub')

        return render(request, 'pages/webinar_seminar_form.html', context={'event': event, 'form': form, 'student': student, 'questions': questions, 'is_president': is_president, 'event_notifications': event_notifications, 'notifications_unread_count': notifications_unread_count, 'user_evaluations': user_evaluations})
    else:
        # Handle other event types
        pass

@login_required(login_url='signin')
@allowed_users(allowed_roles=['student', 'society president'])
def about(request):
    user=request.user
    is_president = request.user.groups.filter(name='society president').exists()  # Check if the user is in the "student" group
    student = Student.objects.filter(student_number=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    notifications_unread_count = unread_notifications.count()  
    return render(request, 'pages/about.html', {'student': student, 'is_president': is_president, 'event_notifications': event_notifications, 'notifications_unread_count': notifications_unread_count})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['student', 'society president'])
def contactUs(request):
    user=request.user
    is_president = request.user.groups.filter(name='society president').exists()  # Check if the user is in the "student" group
    student = Student.objects.filter(student_number=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    notifications_unread_count = unread_notifications.count()  
    return render(request, 'pages/contactUs.html', {'student': student, 'is_president': is_president, 'event_notifications': event_notifications, 'notifications_unread_count': notifications_unread_count})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['student', 'society president'])
def suggestionbox(request):
    student = Student.objects.filter(student_number=request.user.username).first()
    return render(request, 'pages/suggestionbox.html', {'student': student})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['society president'])
def society_president_events(request):
     user=request.user
     is_president = request.user.groups.filter(name='society president'). exists()  # Check if the user is in the "student" group
     student = Student.objects.filter(student_number=request.user.username).first()   
     event_notifications = Notification.objects.filter(recipient=user, level='success')
     unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
     notifications_unread_count = unread_notifications.count()    
     event = Event.objects.filter(author=user, admin_status='Approved').order_by('-updated')
     form = EventCreationForm()
     if request.method == 'POST':
        form = EventCreationForm(request.POST, request.FILES)
        
        today = datetime.today().date()


        if form.is_valid():
            event = form.save(commit=False)
            event.author = request.user  # Set the author to the currently logged-in user
            form.save()
            messages.success(request, "The event has been submitted for approval to the head of OSAS.")
            evaluation_start_date_str = request.POST.get("evaluation_start_datetime")  
            evaluation_end_date_str = request.POST.get("evaluation_end_datetime") 

            event_pk = event.pk
            print(event_pk)
            if evaluation_start_date_str:

                start_date = datetime.strptime(evaluation_start_date_str, "%Y-%m-%d").date()

                # Check if the end date exceeds today's date
                if start_date > today:
                    # Schedule the task if the date is valid
                    trigger = DateTrigger(run_date=start_date) 
                    scheduler.add_job(start_event_evaluations, trigger=trigger, args=[event_pk])
                elif start_date == today:
                    # Schedule the task if the date is valid
                    run_date = datetime.now() + timedelta(minutes=1)
                    trigger = DateTrigger(run_date=run_date) 
                    scheduler.add_job(start_event_evaluations, trigger=trigger, args=[event_pk])
                else:
                    # Display an error if the date exceeds today
                    messages.error(request, "The start date cannot be in the past.")
                    return redirect('faculty_event_evaluations')

            if evaluation_end_date_str:
                end_date = datetime.strptime(evaluation_end_date_str, "%Y-%m-%d").date()

                # Check if the release date exceeds today's date
                if end_date > today:
                    # Schedule the task if the date is valid
                    trigger = DateTrigger(run_date=end_date)
                    scheduler.add_job(end_event_evaluations, trigger=trigger, args=[event_pk])
                else:
                    # Display an error if the date exceeds today
                    messages.error(request, "The end date cannot be in the past.")
                    return redirect('faculty_event_evaluations')
       
            return redirect('society_president_events')
        else:
            # Add form errors to messages
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Error in {field}: {error}")

            return redirect('society_president_events')
     for i in event:
        # Get combined courses and departments names
        attendees_text = ", ".join(course.name for course in i.course_attendees.all()) + " | " + ", ".join(dept.name for dept in i.department_attendees.all())
        # Truncate combined text to 50 characters (or your chosen length)
        i.attendees_summary = Truncator(attendees_text).chars(50)   

     context = {'event': event, 'student': student, 'form':form, 'is_president': is_president, 'event_notifications': event_notifications, 'notifications_unread_count': notifications_unread_count}
     return render(request, 'pages/society_president_events.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['society president'])
def president_view_event_evaluations(request, pk):
    user=request.user
    is_president = request.user.groups.filter(name='society president').exists()  # Check if the user is in the "student" group
    student = Student.objects.filter(student_number=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    notifications_unread_count = unread_notifications.count()    
    event = get_object_or_404(Event, pk=pk)
        # Filter evaluations from both SchoolEventModel and WebinarSeminarModel
    school_event_evaluations = list(SchoolEventModel.objects.filter(event=event))
    webinar_seminar_evaluations = list(WebinarSeminarModel.objects.filter(event=event))

    # Combine the results into one list
    evaluations = school_event_evaluations + webinar_seminar_evaluations
    
    paginator = Paginator(evaluations, 10) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'event': event,
        'evaluations': evaluations,
        'student':student,
        'is_president': is_president,
        'page_obj': page_obj, 'event_notifications': event_notifications, 'notifications_unread_count': notifications_unread_count
    }

    return render(request, 'pages/president_view_event_evaluations.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['society president'])
def view_society_president_event_evaluations(request, pk):
    user=request.user
    is_president = request.user.groups.filter(name='society president').exists()  # Check if the user is in the "student" group
    student = Student.objects.filter(student_number=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    notifications_unread_count = unread_notifications.count()
    # Try to get the event from SchoolEventModel or WebinarSeminarModel
    try:
        event_form_details = SchoolEventModel.objects.get(pk=pk)
        event_type = 'school_event'
    except SchoolEventModel.DoesNotExist:
        try:
            event_form_details = WebinarSeminarModel.objects.get(pk=pk)
            event_type = 'webinar_event'
        except WebinarSeminarModel.DoesNotExist:
            # If no event is found in either model, handle the error
            return HttpResponse("Event not found.")

    # Render different templates based on the event type
    if event_type == 'school_event':
        questions = SchoolEventQuestions.objects.all().order_by('order')

        excellent_count = SchoolEventModel.objects.filter(pk=pk).filter(
            meeting_expectation=5
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            attainment_of_the_objectives=5
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(pk=pk).filter(
            topics_discussed=5
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            input_presentation=5
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            management_team=5
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            venue_and_physical_arrangement=5
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            overall_assessment=5
        ).count()

        very_satisfactory_count = SchoolEventModel.objects.filter(pk=pk).filter(
            meeting_expectation=4
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            attainment_of_the_objectives=4
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(pk=pk).filter(
            topics_discussed=4
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            input_presentation=4
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            management_team=4
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            venue_and_physical_arrangement=4
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            overall_assessment=4
        ).count()

        satisfactory_count = SchoolEventModel.objects.filter(pk=pk).filter(
            meeting_expectation=3
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            attainment_of_the_objectives=3
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(pk=pk).filter(
            topics_discussed=3
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            input_presentation=3
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            management_team=3
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            venue_and_physical_arrangement=3
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            overall_assessment=3
        ).count()

        fair_count = SchoolEventModel.objects.filter(pk=pk).filter(
            meeting_expectation=2
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            attainment_of_the_objectives=2
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(pk=pk).filter(
            topics_discussed=2
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            input_presentation=2
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            management_team=2
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            venue_and_physical_arrangement=2
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            overall_assessment=2
        ).count()

        poor_count = SchoolEventModel.objects.filter(pk=pk).filter(
            meeting_expectation=1
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            attainment_of_the_objectives=1
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(pk=pk).filter(
            topics_discussed=1
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            input_presentation=1
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            management_team=1
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            venue_and_physical_arrangement=1
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            overall_assessment=1
        ).count()

        template_name = 'pages/view_society_president_schoolevent_evaluations.html'
        
    elif event_type == 'webinar_event':
        questions = WebinarSeminarQuestions.objects.all().order_by('order')

        excellent_count = WebinarSeminarModel.objects.filter(pk=pk).filter(
            relevance_of_the_activity=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(pk=pk).filter(
            quality_of_the_activity=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness=5
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            attainment_of_the_objective=5
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_to_attain_the_objective=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods_used=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_in_the_present_time=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            usefulness_of_the_topic_discusssed_in_the_activity=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            displayed_a_thorough_knowledge_of_the_topic=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            explained_activities=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_create_a_good_learning_environment=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_manage_her_time_well=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            demonstrated_keenness_to_the_participant_needs=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness_or_suitability_of_service=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            overall_satisfaction=5
        ).count()

        very_satisfactory_count = WebinarSeminarModel.objects.filter(pk=pk).filter(
            relevance_of_the_activity=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(pk=pk).filter(
            quality_of_the_activity=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness=4
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            attainment_of_the_objective=4
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_to_attain_the_objective=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods_used=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_in_the_present_time=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            usefulness_of_the_topic_discusssed_in_the_activity=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            displayed_a_thorough_knowledge_of_the_topic=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            explained_activities=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_create_a_good_learning_environment=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_manage_her_time_well=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            demonstrated_keenness_to_the_participant_needs=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness_or_suitability_of_service=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            overall_satisfaction=4
        ).count()

        satisfactory_count = WebinarSeminarModel.objects.filter(pk=pk).filter(
            relevance_of_the_activity=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(pk=pk).filter(
            quality_of_the_activity=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness=3
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            attainment_of_the_objective=3
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_to_attain_the_objective=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods_used=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_in_the_present_time=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            usefulness_of_the_topic_discusssed_in_the_activity=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            displayed_a_thorough_knowledge_of_the_topic=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            explained_activities=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_create_a_good_learning_environment=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_manage_her_time_well=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            demonstrated_keenness_to_the_participant_needs=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness_or_suitability_of_service=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            overall_satisfaction=3
        ).count()

        fair_count = WebinarSeminarModel.objects.filter(pk=pk).filter(
            relevance_of_the_activity=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(pk=pk).filter(
            quality_of_the_activity=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness=2
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            attainment_of_the_objective=2
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_to_attain_the_objective=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods_used=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_in_the_present_time=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            usefulness_of_the_topic_discusssed_in_the_activity=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            displayed_a_thorough_knowledge_of_the_topic=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            explained_activities=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_create_a_good_learning_environment=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_manage_her_time_well=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            demonstrated_keenness_to_the_participant_needs=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness_or_suitability_of_service=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            overall_satisfaction=2
        ).count()

        poor_count = WebinarSeminarModel.objects.filter(pk=pk).filter(
            relevance_of_the_activity=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(pk=pk).filter(
            quality_of_the_activity=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness=1
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            attainment_of_the_objective=1
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_to_attain_the_objective=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods_used=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_in_the_present_time=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            usefulness_of_the_topic_discusssed_in_the_activity=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            displayed_a_thorough_knowledge_of_the_topic=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            explained_activities=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_create_a_good_learning_environment=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_manage_her_time_well=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            demonstrated_keenness_to_the_participant_needs=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness_or_suitability_of_service=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            overall_satisfaction=1
        ).count()

       
        template_name = 'pages/view_society_president_webinar_evaluations.html'
    else:
        return HttpResponse("Invalid event type.")


    return render(request, template_name, {'event_form_details': event_form_details, 'faculty': faculty, 'questions': questions, 'excellent_count': excellent_count, 'very_satisfactory_count': very_satisfactory_count, 'satisfactory_count': satisfactory_count, 'fair_count': fair_count, 'poor_count': poor_count, 'student': student, 'event_notifications': event_notifications, 'notifications_unread_count': notifications_unread_count})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['society president'])
def president_edit_event_evaluations(request, pk):
    user=request.user
    is_president = request.user.groups.filter(name='society president').exists()  # Check if the user is in the "student" group
    student = Student.objects.filter(student_number=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    notifications_unread_count = unread_notifications.count()
    event = Event.objects.get(pk=pk)
    form = EventCreationForm(instance=event)
    if request.method == 'POST':
        form = EventCreationForm(request.POST, request.FILES, instance=event)
        if form.is_valid():
            form.save(commit=True)
            return redirect('society_president_events')

           
    context = {'event': event, 'form':form, 'student': student, 'event_notifications': event_notifications, 'notifications_unread_count': notifications_unread_count}
    return render(request, 'pages/president_edit_event_evaluations.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['society president'])
def president_delete_event_evaluations(request, pk):
    user=request.user
    is_president = request.user.groups.filter(name='society president').exists()  # Check if the user is in the "student" group
    student = Student.objects.filter(student_number=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    notifications_unread_count = unread_notifications.count()
    event = Event.objects.get(pk=pk)
    if request.method == 'POST':
            event.delete()
            return redirect('society_president_events')

    return render(request, 'pages/society_president_delete_form.html', {'obj':event, 'student': student, 'event_notifications': event_notifications, 'notifications_unread_count': notifications_unread_count})

def studentlogout(request):
    logout(request)
    return redirect('signin')


            # ------------------------------------------------------
            #                Admin-Page Views
            # ------------------------------------------------------

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def admin(request):
    is_admin = request.user.groups.filter(name='admin').exists()
    is_hr_admin = request.user.groups.filter(name='HR admin').exists()
    student = Student.objects.all()
    course = Course.objects.all()
    section = Section.objects.all()
    faculty = Faculty.objects.all()
    subject = Subject.objects.all()
    user = User.objects.all()
    admins = User.objects.filter(groups__name__in=['admin', 'HR admin'])
    departments = Department.objects.all()

    total_students = student.count()
    total_courses = course.count()
    total_admins = admins.count()
    total_faculty = faculty.count()
    total_subject = subject.count()
    total_user = user.count()

    registered_faculty = Faculty.objects.filter(user__isnull=False).count() 


    registered_students = Student.objects.filter(user__isnull=False).count()

    evaluation_status = EvaluationStatus.objects.first()  # Assuming there's only one status entry
    current_academic_year = evaluation_status.academic_year 
    current_semester = evaluation_status.semester
    
    if request.method == 'POST':
        form = EvaluationStatusForm(request.POST, instance=evaluation_status)
        evaluation_release_date_str = request.POST.get("evaluation_release_date")  # Format: YYYY-MM-DD
        evaluation_end_date_str = request.POST.get("evaluation_end_date")  # Format: YYYY-MM-DD

        today = datetime.today().date()

        if evaluation_end_date_str:
            end_date = datetime.strptime(evaluation_end_date_str, "%Y-%m-%d").date()

            # Check if the end date exceeds today's date
            if end_date > today:
                # Schedule the task if the date is valid
                trigger = DateTrigger(run_date=end_date) 
                scheduler.add_job(close_evaluations, trigger=trigger)
            else:
                # Display an error if the date exceeds today
                messages.error(request, "The end date cannot be in the past.")
                return redirect('admin')

        if evaluation_release_date_str:
            release_date = datetime.strptime(evaluation_release_date_str, "%Y-%m-%d").date()

            # Check if the release date exceeds today's date
            if release_date > today:
                # Schedule the task if the date is valid
                trigger = DateTrigger(run_date=release_date)
                scheduler.add_job(approve_pending_evaluations, trigger=trigger)
            else:
                # Display an error if the date exceeds today
                messages.error(request, "The release date cannot be in the past.")
                return redirect('admin')
        if form.is_valid():
            form.save()
             # Set email_sent to False for all departments
            Department.objects.update(email_sent=False)
            # Set email_sent to False for all faculties
            Faculty.objects.update(email_sent=False)
            messages.success(request, 'Status updated successfully')

    else:
        form = EvaluationStatusForm(instance=evaluation_status)
    
    filterset = LikertEvaluationFilter(request.GET, queryset=LikertEvaluation.objects.all())
   

    if not request.GET or (request.GET.get('academic_year') == '' and request.GET.get('semester') == ''):
        data = LikertEvaluation.objects.filter(academic_year=current_academic_year, semester=current_semester)
        
    else:
        data = filterset.qs

    positive_count = data.filter(predicted_sentiment='Positive').count()
    negative_count = data.filter(predicted_sentiment='Negative').count()
    neutral_count = data.filter(predicted_sentiment='Neutral').count()

    context = {'student': student, 
               'course': course,
               'section': section,
               'faculty': faculty,
               'subject': subject,
               'departments': departments,
                'user': user,
                'form': form,
                'evaluation_status': evaluation_status,
               'total_students': total_students,
                'total_courses': total_courses,
                'total_admins': total_admins,
                'total_faculty': total_faculty,
                'total_subject': total_subject,
                'total_user': total_user,
                'data': data,
                'filterset': filterset,
                'is_admin': is_admin,
                'is_hr_admin': is_hr_admin,
                'registered_faculty': registered_faculty,
                'registered_students': registered_students,
                'positive_count': positive_count,
                'negative_count': negative_count,
                'neutral_count': neutral_count
                }
    return render(request, 'pages/admin.html', context)

def get_all_evaluation_data(request):
    evaluation_status = EvaluationStatus.objects.first()
    current_academic_year = evaluation_status.academic_year 
    current_semester = evaluation_status.semester
    
    faculty_evaluations = LikertEvaluation.objects.filter(academic_year=current_academic_year, semester=current_semester)
    positive_evaluations = faculty_evaluations.filter(predicted_sentiment='Positive').count()
    negative_evaluations = faculty_evaluations.filter(predicted_sentiment='Negative').count()
    neutral_evaluations = faculty_evaluations.filter(predicted_sentiment='Neutral').count()

    data = {
        'positive_evaluations': positive_evaluations,
        'negative_evaluations': negative_evaluations,
        'neutral_evaluations': neutral_evaluations,
    }
    return JsonResponse(data)

def evaluation_response_chart_data(request):
    evaluations = LikertEvaluation.objects.all()

    data = {}
    for evaluation in evaluations:
        # Abbreviate semester names
        semester_abbr = "1S" if evaluation.semester == "1st Semester" else "2S"
        key = f"{evaluation.academic_year} {semester_abbr}"
        if key in data:
            data[key] += 1
        else:
            data[key] = 1

    # Sort data chronologically
    sorted_keys = sorted(data.keys(), key=lambda x: (x.split(" ")[0], x.split(" ")[1]))
    sorted_data = {key: data[key] for key in sorted_keys}

    chart_data = {
        'labels': list(sorted_data.keys()),
        'data': list(sorted_data.values()),
    }
    return JsonResponse(chart_data)

def department_response_chart_data(request):
    evaluation_status = EvaluationStatus.objects.first()
    current_academic_year = evaluation_status.academic_year 
    current_semester = evaluation_status.semester

    departments = Department.objects.annotate(
        total_positive=Count(
            Case(
                When(
                    Q(faculty__sectionsubjectfaculty__likertevaluation__predicted_sentiment="Positive") &
                    Q(faculty__sectionsubjectfaculty__likertevaluation__academic_year=current_academic_year) &
                    Q(faculty__sectionsubjectfaculty__likertevaluation__semester=current_semester),
                    then=1
                ),
                output_field=IntegerField()
            )
        ),
        total_negative=Count(
            Case(
                When(
                    Q(faculty__sectionsubjectfaculty__likertevaluation__predicted_sentiment="Negative") &
                    Q(faculty__sectionsubjectfaculty__likertevaluation__academic_year=current_academic_year) &
                    Q(faculty__sectionsubjectfaculty__likertevaluation__semester=current_semester),
                    then=1
                ),
                output_field=IntegerField()
            )
        ),
        total_neutral=Count(
            Case(
                When(
                    Q(faculty__sectionsubjectfaculty__likertevaluation__predicted_sentiment="Neutral") &
                    Q(faculty__sectionsubjectfaculty__likertevaluation__academic_year=current_academic_year) &
                    Q(faculty__sectionsubjectfaculty__likertevaluation__semester=current_semester),
                    then=1
                ),
                output_field=IntegerField()
            )
        )
    )

    chart_data = {
        'labels': [dept.name for dept in departments],
        'data': [
            [dept.total_positive, dept.total_negative, dept.total_neutral]
            for dept in departments
        ],
    }

    return JsonResponse(chart_data)

def faculty_response_chart_data(request, department_id):
    evaluation_status = EvaluationStatus.objects.first()
    current_academic_year = evaluation_status.academic_year
    current_semester = evaluation_status.semester
    # 1. Rename annotations to avoid method conflicts
    # 2. Add explicit ordering
    faculty_queryset = Faculty.objects.filter(
        department_id=department_id
    ).annotate(
        total_positive=Count(
            Case(
                When(
                    Q(sectionsubjectfaculty__likertevaluation__predicted_sentiment="Positive") &
                    Q(sectionsubjectfaculty__likertevaluation__academic_year=current_academic_year) &
                    Q(sectionsubjectfaculty__likertevaluation__semester=current_semester),
                    then=1
                ),
                output_field=IntegerField()
            )
        ),
        total_negative=Count(
            Case(
                When(
                    Q(sectionsubjectfaculty__likertevaluation__predicted_sentiment="Negative") &
                    Q(sectionsubjectfaculty__likertevaluation__academic_year=current_academic_year) &
                    Q(sectionsubjectfaculty__likertevaluation__semester=current_semester),
                    then=1
                ),
                output_field=IntegerField()
            )
        )
    ).order_by('last_name', 'first_name')  # 3. Required for consistent pagination

    paginator = Paginator(faculty_queryset, 10)
    page_number = request.GET.get('page', 1)

    try:
        page_obj = paginator.page(page_number)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)

    # Convert to serializable types
    chart_data = {
        'labels': [
            f"{faculty.first_name} {faculty.last_name}" 
            for faculty in page_obj
        ],
        'data': [
            [faculty.total_positive, faculty.total_negative]
            for faculty in page_obj
        ],
        'pagination': {
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'has_next': bool(page_obj.has_next()),
            'has_previous': bool(page_obj.has_previous()),
        }
    }
    return JsonResponse(chart_data)



@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def adminregister(request):
    is_admin = request.user.groups.filter(name='admin').exists()
    form = AdminRegistrationForm()
    if request.method == 'POST':
        form = AdminRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()

            user.is_staff = True

            role = form.cleaned_data['role']

            group = Group.objects.get(name=role)

            user.groups.add(group)

            user.save()

            messages.success(request, 'Admin registration successful!')
            return redirect('admin')

 
    return render(request, 'pages/adminregister.html', {'form': form, 'is_admin': is_admin})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def admin_faculty_evaluations(request):
    is_admin = request.user.groups.filter(name='admin').exists()
    evaluation_status = EvaluationStatus.objects.first()  # Assuming there's only one status entry
    current_academic_year = evaluation_status.academic_year 
    current_semester = evaluation_status.semester
    current_evaluation_status = evaluation_status.evaluation_status

    evaluation = LikertEvaluation.objects.filter(academic_year=current_academic_year, semester=current_semester)

       # Count evaluations per course
    evaluations_per_course = evaluation.values(
        'user__student__Course__name'
    ).annotate(total=Count('id'), pk=F('user__student__Course__pk'))

    total_evaluations = evaluation.count()
    departments = Department.objects.all()
    #pagination
    page_number = request.GET.get('page', 1)
    evaluation_paginator = Paginator(evaluation, ITEMS_PER_PAGE)
   

    try:
        page = evaluation_paginator.page(page_number)
    except EmptyPage:
        page = evaluation_paginator.page(evaluation_paginator.num_pages)
 
    context = {'evaluation': page.object_list,'total_evaluations': total_evaluations, 'page_obj':page, 'is_paginated': True, 'paginator':evaluation_paginator,'is_admin': is_admin, 'evaluations_per_course': evaluations_per_course, 'current_academic_year': current_academic_year, 'current_semester': current_semester, 'current_evaluation_status': current_evaluation_status, 'departments': departments }
    
    return render(request, 'pages/admin_faculty_evaluations.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def admin_send_report_to_department_heads(request):
    if request.method == 'POST':
        department_ids = request.POST.getlist('departments')  # Get selected departments
        departments = Department.objects.filter(id__in=department_ids)

        # Filter department head/program coordinators: Faculties with the "department head/program coordinator" role and assigned to the selected departments
        department_heads = Faculty.objects.filter(
            user__groups__name="department head/program coordinator",  # Assuming roles are managed via Django Groups
            department__in=departments
        )

        evaluation_status = EvaluationStatus.objects.first()
        current_academic_year = evaluation_status.academic_year
        current_semester = evaluation_status.semester

        # Group department head/program coordinators by their respective departments
        department_heads_by_department = {}
        for head in department_heads:
            department_heads_by_department.setdefault(head.department, []).append(head)

        # Generate and send reports per department
        for department, heads in department_heads_by_department.items():
            # Filter faculties in the department
            faculties_in_department = Faculty.objects.filter(department=department).order_by('last_name')

            # Filter evaluations for faculties in the department
            evaluations_in_department = LikertEvaluation.objects.filter(
                section_subject_faculty__faculty__in=faculties_in_department,
                academic_year=current_academic_year,
                semester=current_semester
            )
            department.email_sent = True
            department.save()

            summary_data = []
            comments_data = []

            for faculty in faculties_in_department:
                evaluations = evaluations_in_department.filter(section_subject_faculty__faculty=faculty)
                num_evaluators = evaluations.count()

                if num_evaluators > 0:
                    category_sums = {
                        'Subject Matter Content': 0,
                        'Organization': 0,
                        'Teacher-Student Rapport': 0,
                        'Teaching Methods': 0,
                        'Presentation': 0,
                        'Classroom Management': 0,
                        'Sensitivity and Support to Students': 0,
                        'Overall': 0
                    }
                    MAX_SCORE = 5
                    NUM_CATEGORIES = len(category_sums) - 1  # Exclude 'Overall' key
                    for evaluation in evaluations:
                        category_averages = evaluation.calculate_category_averages()
                        for category, average in category_averages.items():
                            if average is not None:
                                category_sums[category] += average
                        category_sums['Overall'] += evaluation.average_rating

                    category_averages = {category: round(total / num_evaluators, 2) for category, total in category_sums.items()}
                    avg_rating = category_averages['Overall']

                    # Calculate overall percentage of category averages
                    total_score = sum(category_averages[category] for category in category_sums if category != 'Overall')
                    max_possible_score = MAX_SCORE * NUM_CATEGORIES
                    overall_percentage = round((total_score / max_possible_score) * 100, 2)

                    rating_category = (
                        "Poor" if avg_rating <= 1.99 else
                        "Unsatisfactory" if avg_rating <= 2.99 else
                        "Satisfactory" if avg_rating <= 3.99 else
                        "Very Satisfactory" if avg_rating <= 4.99 else
                        "Outstanding"
                    )

                    summary_data.append({
                        'faculty': f"{faculty.first_name} {faculty.last_name}",
                        'num_evaluators': num_evaluators,
                        'subject_matter_content_avg': category_averages['Subject Matter Content'],
                        'organization_avg': category_averages['Organization'],
                        'teacher_student_rapport_avg': category_averages['Teacher-Student Rapport'],
                        'teaching_methods_avg': category_averages['Teaching Methods'],
                        'presentation_avg': category_averages['Presentation'],
                        'classroom_management_avg': category_averages['Classroom Management'],
                        'sensitivity_support_students_avg': category_averages['Sensitivity and Support to Students'],
                        'overall_avg': category_averages['Overall'],
                        'overall_percentage': overall_percentage,
                        'rating_category': rating_category
                    })

                    for evaluation in evaluations:
                        comments_data.append({
                            'faculty': f"{faculty.first_name} {faculty.last_name}",
                            'requires_less_task_for_credit': evaluation.requires_less_task_for_credit,
                            'strengths_of_the_faculty': evaluation.strengths_of_the_faculty,
                            'other_suggestions_for_improvement': evaluation.other_suggestions_for_improvement,
                            'comments': evaluation.comments
                        })
            department_name = department.name
            # Render the summary report for this department
            image_path = os.path.join(settings.BASE_DIR, 'static/images/cvsulogo.png')
            html = render_to_string('pages/faculty_evaluations_department_summary_report.html', {
                'summary_data': summary_data,
                'image_path': image_path,
                'comments_data': comments_data,
                'current_academic_year': current_academic_year,
                'current_semester': current_semester,
                'department_name': department_name
            })

            pdf_path = os.path.join(settings.MEDIA_ROOT, f'faculty_evaluations_summary_{department.id}.pdf')
            with open(pdf_path, 'wb') as pdf_file:
                pisa_status = pisa.CreatePDF(html, dest=pdf_file, link_callback=link_callback)

            if pisa_status.err:
                return HttpResponse(f'Error generating PDF for department {department.id}')

            # Send the email to all department head/program coordinators of this department
            for head in heads:
                email = EmailMessage(
                    subject=f'Faculty Evaluations Summary Report - {department.name}',
                    body='Please find attached the faculty evaluations summary report for your department.',
                    from_email='admin@example.com',
                    to=[head.email],
                )
                email.attach_file(pdf_path)
                email.send()

                notify.send(request.user, 
                            recipient=head.user, 
                            verb='Summary Report Sent', 
                            description=f'The summary report for the {head.department.name} has been sent to your email.',
                            level='info')
                
        messages.success(request, "The summary reports have been successfully sent to the department head and program coordinators.") 
        return redirect('admin_faculty_evaluations')  # Redirect to a success page


@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def admin_faculty_evaluations_sections(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    evaluation_status = EvaluationStatus.objects.first()  # Assuming there's only one status entry
    current_academic_year = evaluation_status.academic_year 
    current_semester = evaluation_status.semester


    evaluation_course = get_object_or_404(Course, pk=pk)
       # Count evaluations per course
    sections = Section.objects.filter(course=evaluation_course)
    section_evaluation_counts = []
    for section in sections: 
        evaluation_count = LikertEvaluation.objects.filter( user__student__Section=section, academic_year=current_academic_year, semester=current_semester).count() 
        section_evaluation_counts.append({ 'section': section, 'evaluation_count': evaluation_count })
    #pagination
    page_number = request.GET.get('page', 1)
    evaluation_paginator = Paginator(section_evaluation_counts, ITEMS_PER_PAGE)
    print(section_evaluation_counts, pk)

    try:
        page = evaluation_paginator.page(page_number)
    except EmptyPage:
        page = evaluation_paginator.page(evaluation_paginator.num_pages)
 
    context = {'evaluation': page.object_list, 'page_obj':page, 'is_paginated': True, 'paginator':evaluation_paginator,'is_admin': is_admin,'evaluation_course': evaluation_course }
    return render(request, 'pages/admin_faculty_evaluations_sections.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def admin_faculty_evaluations_sections_view_forms(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    evaluation_status = EvaluationStatus.objects.first()  # Assuming there's only one status entry
    current_academic_year = evaluation_status.academic_year 
    current_semester = evaluation_status.semester
    academic_year = request.GET.get('academic_year', current_academic_year)
    semester = request.GET.get('semester', current_semester)

    selected_section = get_object_or_404(Section, pk=pk)
       # Count evaluations per course
    evaluations = LikertEvaluation.objects.filter(academic_year=academic_year, semester=semester, user__student__Section=selected_section)
    filter_params = request.GET.copy()

    # Set defaults only when values are missing or empty
    if not filter_params.get('academic_year'):
        filter_params['academic_year'] = current_academic_year
    if not filter_params.get('semester'):
        filter_params['semester'] = current_semester

     #filter and search
    faculty_evaluation_filter = EvaluationFilter(filter_params, queryset=evaluations)
    evaluations = faculty_evaluation_filter.qs
    

    # ordering functionality
   
    ordering = request.GET.get('ordering', "")

     
    if ordering:
        evaluations = evaluations.order_by(ordering) 

    page_number = request.GET.get('page', 1)
    evaluation_paginator = Paginator(evaluations, ITEMS_PER_PAGE)

    try:
        page = evaluation_paginator.page(page_number)
    except EmptyPage:
        page = evaluation_paginator.page(evaluation_paginator.num_pages)
 
    context = {'evaluation': page.object_list, 'page_obj':page, 'is_paginated': True, 'paginator':evaluation_paginator,'is_admin': is_admin,'selected_section': selected_section, 'faculty_evaluation_filter': faculty_evaluation_filter, 'current_academic_year': current_academic_year, 'current_semester': current_semester }
    return render(request, 'pages/admin_faculty_evaluations_sections_view_forms.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def view_latest_faculty_evaluations(request):
    evaluation_status = EvaluationStatus.objects.first()
    current_academic_year = evaluation_status.academic_year 
    current_semester = evaluation_status.semester
    academic_year = request.GET.get('academic_year', current_academic_year)
    semester = request.GET.get('semester', current_semester)
    
    evaluation = LikertEvaluation.objects.filter(academic_year=academic_year, semester=semester)
    filter_params = request.GET.copy()

    # Set defaults only when values are missing or empty
    if not filter_params.get('academic_year'):
        filter_params['academic_year'] = current_academic_year
    if not filter_params.get('semester'):
        filter_params['semester'] = current_semester

    is_admin = request.user.groups.filter(name='admin').exists()
    total_evaluations = evaluation.count()
    #filter and search
    faculty_evaluation_filter = EvaluationFilter(filter_params, queryset=evaluation)
    evaluation = faculty_evaluation_filter.qs
    

    # ordering functionality
   
    ordering = request.GET.get('ordering', "")

     
    if ordering:
        evaluation = evaluation.order_by(ordering) 

    #pagination
    page_number = request.GET.get('page', 1)
    evaluation_paginator = Paginator(evaluation, ITEMS_PER_PAGE)
   

    try:
        page = evaluation_paginator.page(page_number)
    except EmptyPage:
        page = evaluation_paginator.page(evaluation_paginator.num_pages)
 
    context = {'evaluation': page.object_list,'total_evaluations': total_evaluations, 'faculty_evaluation_filter': faculty_evaluation_filter, 'page_obj':page, 'is_paginated': True, 'paginator':evaluation_paginator,'ordering': ordering, 'is_admin': is_admin, 'current_academic_year': current_academic_year, 'current_semester': current_semester }
    return render(request, 'pages/view_latest_faculty_evaluations.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def evaluations_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=faculty_evaluations.csv'

    # Create a csv writer
    writer = csv.writer(response)

    # Apply filters from the EvaluationFilter based on the request data
    evaluation_filter = EvaluationFilter(request.GET, queryset=LikertEvaluation.objects.all())

    # Get the filtered queryset
    filtered_evaluations = evaluation_filter.qs

    # Add column headings to csv file

    writer.writerow(['Subject', 'Faculty', 'Average', 'Rating', 'Overall Impression', 'Polarity', 'Academic Year', 'Semester'])

    # Loop thru and output
    for i in filtered_evaluations:
        writer.writerow([i.section_subject_faculty.subjects, i.section_subject_faculty.faculty, i.average_rating, i.get_rating_category(), i.comments, i.predicted_sentiment, i.academic_year, i.semester ])

    return response

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def faculty_evaluations_summary_report_pdf(request):
    """
    Generates a PDF summary report of faculty evaluations.

    Args:
        request: The HTTP request object.

    Returns:
        An HTTP response containing the PDF report.
    """
    image_path = os.path.join(settings.BASE_DIR, static('images/cvsulogo.png'))
    evaluation_status = EvaluationStatus.objects.first()  # Assuming there's only one status entry

    current_academic_year = evaluation_status.academic_year 
    current_semester = evaluation_status.semester
    academic_year =  request.GET.get('academic_year', current_academic_year)
    semester =  request.GET.get('semester', current_semester)

    # Apply filters from the EvaluationFilter based on the request data
    evaluation_filter = EvaluationFilter(request.GET, queryset=LikertEvaluation.objects.filter(academic_year=academic_year, semester=semester))
    filtered_evaluations = evaluation_filter.qs

    # Get all faculty
    faculties = SectionSubjectFaculty.objects.values('faculty').distinct()

    summary_data = []
    comments_data = []

  

    # Loop through each faculty and calculate the required aggregates
    for faculty in faculties:
        faculty_id = faculty['faculty']
        faculty_instance = Faculty.objects.get(id=faculty_id)
        faculty_name = faculty_instance.full_name()  
        evaluations = filtered_evaluations.filter(section_subject_faculty__faculty=faculty_id)
        num_evaluators = evaluations.count()

        if num_evaluators > 0:
            category_sums = {
                'Subject Matter Content': 0,
                'Organization': 0,
                'Teacher-Student Rapport': 0,
                'Teaching Methods': 0,
                'Presentation': 0,
                'Classroom Management': 0,
                'Sensitivity and Support to Students': 0,
                'Overall': 0
            }
            MAX_SCORE = 5
            NUM_CATEGORIES = len(category_sums) - 1  # Exclude 'Overall' key
            for evaluation in evaluations:
                category_averages = evaluation.calculate_category_averages()
                for category, average in category_averages.items():
                    if average is not None:
                        category_sums[category] += average
                category_sums['Overall'] += evaluation.average_rating

            # Calculate averages
            category_averages = {category: round(total / num_evaluators, 2) for category, total in category_sums.items()}
            avg_rating = category_averages['Overall']

            # Calculate overall percentage of category averages
            total_score = sum(category_averages[category] for category in category_sums if category != 'Overall')
            max_possible_score = MAX_SCORE * NUM_CATEGORIES
            overall_percentage = round((total_score / max_possible_score) * 100, 2)

            if avg_rating is not None:
                if 1.0 <= avg_rating <= 1.99:
                    rating_category = "Poor"
                elif 2.0 <= avg_rating <= 2.99:
                    rating_category = "Unsatisfactory"
                elif 3.0 <= avg_rating <= 3.99:
                    rating_category = "Satisfactory"
                elif 4.0 <= avg_rating <= 4.99:
                    rating_category = "Very Satisfactory"
                elif 5.0 <= avg_rating <= 5.0:
                    rating_category = "Outstanding"
                else:
                    rating_category = "No Rating"
            else:
                rating_category = "No Rating"

            summary_data.append({
                'faculty': faculty_name,
                'num_evaluators': num_evaluators,
                'subject_matter_content_avg': category_averages['Subject Matter Content'],
                'organization_avg': category_averages['Organization'],
                'teacher_student_rapport_avg': category_averages['Teacher-Student Rapport'],
                'teaching_methods_avg': category_averages['Teaching Methods'],
                'presentation_avg': category_averages['Presentation'],
                'classroom_management_avg': category_averages['Classroom Management'],
                'sensitivity_support_students_avg': category_averages['Sensitivity and Support to Students'],
                'overall_avg': category_averages['Overall'],
                'overall_percentage': overall_percentage,
                'rating_category': rating_category
            })

            for evaluation in evaluations: 
                comments_data.append({ 'faculty': faculty_name, 
                                      'requires_less_task_for_credit': evaluation.requires_less_task_for_credit, 
                                      'strengths_of_the_faculty': evaluation.strengths_of_the_faculty,
                                    'other_suggestions_for_improvement': evaluation.other_suggestions_for_improvement,
                                    'comments': evaluation.comments })

    # Render the summary data to an HTML template
    html = render_to_string('pages/faculty_evaluations_summary_report.html', {'summary_data': summary_data, 'image_path': image_path, 'comments_data': comments_data, 'academic_year': academic_year, 'semester': semester})

    # Create the PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="SET-Summary-Report.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response, link_callback=link_callback)

    if pisa_status.err:
        return HttpResponse('We had some errors with code %s' % pisa_status.err)
    return response

@login_required(login_url='signin')
@allowed_users(allowed_roles=['department head/program coordinator', 'head of OSAS'])
def peer_to_peer_summary_report_pdf(request):
    """
    Generates a PDF summary report of peer-to-peer evaluations.
    """
    # Build the path to the logo image
    image_path = os.path.join(settings.BASE_DIR, static('images/cvsulogo.png'))
    
    # Get current academic year and semester
    evaluation_status = EvaluationStatus.objects.first()  # Assuming there's only one status entry
    current_academic_year = evaluation_status.academic_year 
    current_semester = evaluation_status.semester
    academic_year =  request.GET.get('academic_year', current_academic_year)
    semester =  request.GET.get('semester', current_semester)

    # Get the department of the logged-in user
    user_department = request.user.faculty.department

    # Filter peer-to-peer evaluations for the current academic year, semester, and department
    evaluation_filter = EvaluationFilter(
        request.GET,
        queryset=PeertoPeerEvaluation.objects.filter(
            academic_year=academic_year,
            semester=semester,
            peer__department=user_department
        )
    )
    filtered_evaluations = evaluation_filter.qs

    # Get all faculties in the same department
    faculties = Faculty.objects.filter(department=user_department).distinct()

    summary_data = []
    comments_data = []

    # Loop through each faculty and calculate aggregates
    for faculty in faculties:
        faculty_id = faculty.id
        faculty_name = faculty.full_name()  # Assumes full_name() returns a string

        # Filter evaluations where this faculty is the peer
        evaluations = filtered_evaluations.filter(peer=faculty_id)
        num_evaluators = evaluations.count()

        if num_evaluators > 0:
            # Initialize sums for each category.
            # Note: These keys must match those returned by calculate_category_averages().
            category_sums = {
                'Subject Matter Content': 0,
                'Organization': 0,
                'Teacher-Student Rapport': 0,
                'Teaching Methods': 0,
                'Presentation': 0,
                'Classroom Management': 0,
                'Sensitivity and Support to Students': 0,
                
            }

            MAX_SCORE = 5
            NUM_CATEGORIES = len(category_sums)  # Exclude 'Overall' key

            # Loop through evaluations and sum the averages per category.
            for evaluation in evaluations:
                # Use the new function in the model that returns a dictionary of averages.
                result = evaluation.calculate_category_averages()
                # Debug print to check the structure of the returned result:
                # print("DEBUG: calculate_category_averages() returned:", result)

                if isinstance(result, dict):
                    for category, avg in result.items():
                        if avg is not None and category in category_sums:
                            category_sums[category] += avg

            # Calculate the average for each category for the current faculty.
            category_averages = {
                category: round(total / num_evaluators, 2)
                for category, total in category_sums.items()
            }
            # Optionally, compute an overall average as the mean of all category averages.
            overall_avg = round(sum(category_averages.values()) / len(category_averages), 2)
           
            # Calculate overall percentage of category averages
            total_score = sum(category_averages[category] for category in category_sums if category != 'Overall')
            max_possible_score = MAX_SCORE * NUM_CATEGORIES
            overall_percentage = round((total_score / max_possible_score) * 100, 2)
           
            # Determine the rating category based on the overall average rating.
            if overall_avg is not None:
                if 1.0 <= overall_avg <= 1.49:
                    rating_category = "Poor"
                elif 1.5 <= overall_avg <= 2.49:
                    rating_category = "Unsatisfactory"
                elif 2.5 <= overall_avg <= 3.49:
                    rating_category = "Satisfactory"
                elif 3.5 <= overall_avg <= 4.49:
                    rating_category = "Very Satisfactory"
                elif 4.5 <= overall_avg <= 5.0:
                    rating_category = "Outstanding"
                else:
                    rating_category = "No Rating"
            else:
                rating_category = "No Rating"

            # Append the summarized data for this faculty.
            summary_data.append({
                'faculty': faculty_name,
                'num_evaluators': num_evaluators,
                'subject_matter_content_avg': category_averages['Subject Matter Content'],
                'organization_avg': category_averages['Organization'],
                'teacher_student_rapport_avg': category_averages['Teacher-Student Rapport'],
                'teaching_methods_avg': category_averages['Teaching Methods'],
                'presentation_avg': category_averages['Presentation'],
                'classroom_management_avg': category_averages['Classroom Management'],
                'sensitivity_support_students_avg': category_averages['Sensitivity and Support to Students'],
                'overall_avg': overall_avg,
                'overall_percentage': overall_percentage,
                'rating_category': rating_category
            })

            # Collect comments data for each evaluation of this faculty.
            for evaluation in evaluations:
                comments_data.append({
                    'faculty': faculty_name,
                    'requires_less_task_for_credit': evaluation.requires_less_task_for_credit,
                    'strengths_of_the_faculty': evaluation.strengths_of_the_faculty,
                    'other_suggestions_for_improvement': evaluation.other_suggestions_for_improvement,
                    'comments': evaluation.comments
                })

    # Render the summary data to an HTML template.
    html = render_to_string('pages/peer_to_peer_summary_report.html', {
        'summary_data': summary_data,
        'image_path': image_path,
        'comments_data': comments_data,
        'academic_year': academic_year,
        'semester': semester,
        'user_department': user_department
    })

    # Create the PDF response using xhtml2pdf.
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="peer_to_peer_summary_report.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response, link_callback=link_callback)

    if pisa_status.err:
        return HttpResponse('We had some errors with code %s' % pisa_status.err)
    return response

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def admin_view_evaluation_form(request, pk):
    faculty_evaluation_form = get_object_or_404(LikertEvaluation, pk=pk)
    questions = FacultyEvaluationQuestions.objects.all().order_by('order')
    is_admin = request.user.groups.filter(name='admin').exists()
    
    rating_fields = [
        'command_and_knowledge_of_the_subject',
        'depth_of_mastery',
        'practice_in_respective_discipline',
        'up_to_date_knowledge',
        'integrates_subject_to_practical_circumstances',
        'organizes_the_subject_matter',
        'provides_orientation_on_course_content',
        'efforts_of_class_preparation',
        'summarizes_main_points',
        'monitors_online_class',
        'holds_interest_of_students',
        'provides_relevant_feedback',
        'encourages_participation',
        'shows_enthusiasm',
        'shows_sense_of_humor',
        'teaching_methods',
        'flexible_learning_strategies',
        'student_engagement',
        'clear_examples',
        'focused_on_objectives',
        'starts_with_motivating_activities',
        'speaks_in_clear_and_audible_manner',
        'uses_appropriate_medium_of_instruction',
        'establishes_online_classroom_environment',
        'observes_proper_classroom_etiquette',
        'uses_time_wisely',
        'gives_ample_time_for_students_to_prepare',
        'updates_the_students_of_their_progress',
        'demonstrates_leadership_and_professionalism',
        'understands_possible_distractions',
        'sensitivity_to_student_culture',
        'responds_appropriately',
        'assists_students_on_concerns',
        'guides_the_students_in_accomplishing_tasks',
        'extends_consideration_to_students'
    ]
    # Initialize counters
    outstanding_count = 0
    very_satisfactory_count = 0
    satisfactory_count = 0
    unsatisfactory_count = 0
    poor_count = 0

    # Calculate counts in memory
    for field in rating_fields:
        value = getattr(faculty_evaluation_form, field)
        if value == 5:
            outstanding_count += 1
        elif value == 4:
            very_satisfactory_count += 1
        elif value == 3:
            satisfactory_count += 1
        elif value == 2:
            unsatisfactory_count += 1
        elif value == 1:
            poor_count += 1

    return render(request, 'pages/admin_view_evaluation_form.html', {'faculty_evaluation_form': faculty_evaluation_form, 'faculty': faculty, 'questions': questions, 'outstanding_count': outstanding_count, 'very_satisfactory_count': very_satisfactory_count, 'satisfactory_count': satisfactory_count, 'unsatisfactory_count': unsatisfactory_count, 'poor_count': poor_count, 'is_admin': is_admin})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def deleteEvaluation(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    evaluation = LikertEvaluation.objects.get(pk=pk)
    if request.method == 'POST':
            evaluation.delete()
            return redirect('evaluations')

    return render(request, 'pages/delete.html', {'obj':evaluation, 'is_admin': is_admin})



@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def admin_event_list(request):
    is_admin = request.user.groups.filter(name='admin').exists()
    events = Event.objects.filter(admin_status='Approved').order_by('-updated') 
    event_filter = EventFilter(request.GET, queryset=events)
    events = event_filter.qs 
     # Annotate events with average rating and evaluator count depending on event type.
    events = events.annotate(
    avg_rating=Case(
        When(
            event_type__name__in=['School Event', 'Training Workshop'],
            then=Coalesce(Avg('schooleventmodel__average_rating'), Value(0))
        ),
        When(
            event_type__name='Webinar/Seminar',
            then=Coalesce(Avg('webinarseminarmodel__average_rating'), Value(0))
        ),
        default=Value(0),
        output_field=FloatField()
    ),
    evaluator_count=Case(
         When(
             event_type__name__in=['School Event', 'Training Workshop'],
             then=Count('schooleventmodel__id')
         ),
         When(
             event_type__name='Webinar/Seminar',
             then=Count('webinarseminarmodel__id')
         ),
         default=Value(0),
         output_field=IntegerField()
    )
)
    ordering = request.GET.get('ordering', "")

     
    if ordering:
        events = events.order_by(ordering) 

    paginator = Paginator(events, 10) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request,'pages/admin_event_list.html',{'events': events, 'page_obj': page_obj, 'event_filter': event_filter, 'is_admin': is_admin, })

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def admin_edit_event(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    event = Event.objects.get(pk=pk)
    form = EventCreationForm(instance=event)
    if request.method == 'POST':
        form = EventCreationForm(request.POST, request.FILES, instance=event)
        if form.is_valid():
            form.save(commit=True)
            return redirect('faculty_event_evaluations')

           
    context = {'event': event, 'form':form, 'is_admin': is_admin}
    return render(request, 'pages/admin_edit_event.html', context)


@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def admin_delete_event(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    event = Event.objects.get(pk=pk)
    if request.method == 'POST':
            event.delete()
            return redirect('admin_event_evaluations')

    return render(request, 'pages/delete.html', {'obj':event, 'is_admin': is_admin})


@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def admin_event_evaluations(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    # Fetch the event
    event = get_object_or_404(Event, pk=pk)

    # Determine which model and filter to use based on the event type
    if event.event_type.name in ['School Event', 'Training Workshop']:
        evaluations_queryset = SchoolEventModel.objects.filter(event=event)
        filter_class = SchoolEventFilter
        total_average_rating = SchoolEventModel.get_event_average_rating(event)
    elif event.event_type.name == 'Webinar/Seminar':
        evaluations_queryset = WebinarSeminarModel.objects.filter(event=event)
        filter_class = WebinarSeminarEventFilter
        total_average_rating = WebinarSeminarModel.get_event_average_rating(event)
    else:
        evaluations_queryset = None
        filter_class = None
        total_average_rating = 'Invalid event type'

    # Apply filters if applicable
    if filter_class and evaluations_queryset is not None:
        event_filter = filter_class(request.GET, queryset=evaluations_queryset)
        filtered_evaluations = event_filter.qs
    else:
        event_filter = None
        filtered_evaluations = evaluations_queryset

    # Ensure filtered_evaluations is not None before paginating
    if filtered_evaluations is None:
        filtered_evaluations = []

    paginator = Paginator(filtered_evaluations, 10) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'event': event,
        'filtered_evaluations': filtered_evaluations,
        'total_average_rating': total_average_rating,
        'page_obj': page_obj, 
        'is_admin': is_admin,
        'event_filter': event_filter
    }
    return render(request, 'pages/admin_event_evaluations.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin', 'society president', 'faculty', 'head of OSAS'])
def eventevaluations_excel(request):
    # Set up the Excel response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=event_evaluations.xlsx'

    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Event Evaluations"

    # Retrieve the event ID and filter parameters from the request
    event_id = request.GET.get('event_id')
    event = get_object_or_404(Event, pk=event_id)
    
    # Retrieve filter parameters from GET
    predicted_sentiment = request.GET.get('predicted_sentiment')
    academic_year = request.GET.get('academic_year')
    semester = request.GET.get('semester')
    rating_category = request.GET.get('rating_category')  # if you want to filter by rating category too

    # Determine which model to query based on event type
    evaluations = SchoolEventModel.objects.filter(event=event)
    event_type = 'school_event'
    if not evaluations.exists():
        evaluations = WebinarSeminarModel.objects.filter(event=event)
        event_type = 'webinar_event'

    # If no evaluations are found at all, return an error
    if not evaluations.exists():
        return HttpResponse("No evaluations found for this event.")

    # Apply filters if provided
    if predicted_sentiment:
        evaluations = evaluations.filter(predicted_sentiment=predicted_sentiment)
    if academic_year:
        evaluations = evaluations.filter(academic_year=academic_year)
    if semester:
        evaluations = evaluations.filter(semester=semester)
    
    # If you have a rating category filter and the models use a method like get_rating_category(),
    # you might need to filter in Python rather than at the database level:
    if rating_category:
        evaluations = [e for e in evaluations if e.get_rating_category() == rating_category]

    # Write the headers
    headers = [
        'Event', 'Suggestions and Comments', 'Average Rating', 'Rating Category',
        'Academic Year', 'Semester', 'Date Submitted'
    ]
    ws.append(headers)

    # Adjust column widths
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 20

    # Write the evaluation data
    # If evaluations is a QuerySet, we can iterate; if we filtered by rating_category above, it is now a list.
    for evaluation in evaluations:
        ws.append([
            evaluation.event.title,
            evaluation.suggestions_and_comments,
            evaluation.average_rating,
            evaluation.get_rating_category(),
            evaluation.academic_year,
            evaluation.semester,
            evaluation.created.strftime('%Y-%m-%d %H:%M:%S')
        ])

    # Center-align the header row
    for cell in ws["1:1"]:
        cell.alignment = Alignment(horizontal='center')

    # Save the workbook into the response and return it
    wb.save(response)
    return response

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def view_admin_schoolevent_evaluations(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    # Try to get the event from SchoolEventModel or WebinarSeminarModel
    try:
        event_form_details = SchoolEventModel.objects.get(pk=pk)
        event_type = 'school_event'
    except SchoolEventModel.DoesNotExist:
        try:
            event_form_details = WebinarSeminarModel.objects.get(pk=pk)
            event_type = 'webinar_event'
        except WebinarSeminarModel.DoesNotExist:
            # If no event is found in either model, handle the error
            return HttpResponse("Event not found.")

    # Render different templates based on the event type
    if event_type == 'school_event':
        questions = SchoolEventQuestions.objects.all().order_by('order')

        excellent_count = SchoolEventModel.objects.filter(pk=pk).filter(
            meeting_expectation=5
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            attainment_of_the_objectives=5
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(pk=pk).filter(
            topics_discussed=5
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            input_presentation=5
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            management_team=5
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            venue_and_physical_arrangement=5
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            overall_assessment=5
        ).count()

        very_satisfactory_count = SchoolEventModel.objects.filter(pk=pk).filter(
            meeting_expectation=4
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            attainment_of_the_objectives=4
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(pk=pk).filter(
            topics_discussed=4
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            input_presentation=4
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            management_team=4
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            venue_and_physical_arrangement=4
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            overall_assessment=4
        ).count()

        satisfactory_count = SchoolEventModel.objects.filter(pk=pk).filter(
            meeting_expectation=3
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            attainment_of_the_objectives=3
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(pk=pk).filter(
            topics_discussed=3
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            input_presentation=3
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            management_team=3
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            venue_and_physical_arrangement=3
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            overall_assessment=3
        ).count()

        fair_count = SchoolEventModel.objects.filter(pk=pk).filter(
            meeting_expectation=2
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            attainment_of_the_objectives=2
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(pk=pk).filter(
            topics_discussed=2
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            input_presentation=2
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            management_team=2
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            venue_and_physical_arrangement=2
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            overall_assessment=2
        ).count()

        poor_count = SchoolEventModel.objects.filter(pk=pk).filter(
            meeting_expectation=1
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            attainment_of_the_objectives=1
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(pk=pk).filter(
            topics_discussed=1
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            input_presentation=1
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            management_team=1
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            venue_and_physical_arrangement=1
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            overall_assessment=1
        ).count()

        template_name = 'pages/view_admin_schoolevent_evaluations.html'
        
    elif event_type == 'webinar_event':
        questions = WebinarSeminarQuestions.objects.all().order_by('order')

        excellent_count = WebinarSeminarModel.objects.filter(pk=pk).filter(
            relevance_of_the_activity=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(pk=pk).filter(
            quality_of_the_activity=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness=5
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            attainment_of_the_objective=5
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_to_attain_the_objective=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods_used=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_in_the_present_time=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            usefulness_of_the_topic_discusssed_in_the_activity=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            displayed_a_thorough_knowledge_of_the_topic=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            explained_activities=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_create_a_good_learning_environment=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_manage_her_time_well=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            demonstrated_keenness_to_the_participant_needs=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness_or_suitability_of_service=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            overall_satisfaction=5
        ).count()

        very_satisfactory_count = WebinarSeminarModel.objects.filter(pk=pk).filter(
            relevance_of_the_activity=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(pk=pk).filter(
            quality_of_the_activity=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness=4
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            attainment_of_the_objective=4
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_to_attain_the_objective=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods_used=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_in_the_present_time=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            usefulness_of_the_topic_discusssed_in_the_activity=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            displayed_a_thorough_knowledge_of_the_topic=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            explained_activities=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_create_a_good_learning_environment=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_manage_her_time_well=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            demonstrated_keenness_to_the_participant_needs=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness_or_suitability_of_service=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            overall_satisfaction=4
        ).count()

        satisfactory_count = WebinarSeminarModel.objects.filter(pk=pk).filter(
            relevance_of_the_activity=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(pk=pk).filter(
            quality_of_the_activity=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness=3
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            attainment_of_the_objective=3
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_to_attain_the_objective=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods_used=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_in_the_present_time=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            usefulness_of_the_topic_discusssed_in_the_activity=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            displayed_a_thorough_knowledge_of_the_topic=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            explained_activities=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_create_a_good_learning_environment=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_manage_her_time_well=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            demonstrated_keenness_to_the_participant_needs=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness_or_suitability_of_service=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            overall_satisfaction=3
        ).count()

        fair_count = WebinarSeminarModel.objects.filter(pk=pk).filter(
            relevance_of_the_activity=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(pk=pk).filter(
            quality_of_the_activity=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness=2
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            attainment_of_the_objective=2
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_to_attain_the_objective=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods_used=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_in_the_present_time=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            usefulness_of_the_topic_discusssed_in_the_activity=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            displayed_a_thorough_knowledge_of_the_topic=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            explained_activities=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_create_a_good_learning_environment=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_manage_her_time_well=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            demonstrated_keenness_to_the_participant_needs=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness_or_suitability_of_service=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            overall_satisfaction=2
        ).count()

        poor_count = WebinarSeminarModel.objects.filter(pk=pk).filter(
            relevance_of_the_activity=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(pk=pk).filter(
            quality_of_the_activity=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness=1
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            attainment_of_the_objective=1
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_to_attain_the_objective=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods_used=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_in_the_present_time=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            usefulness_of_the_topic_discusssed_in_the_activity=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            displayed_a_thorough_knowledge_of_the_topic=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            explained_activities=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_create_a_good_learning_environment=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_manage_her_time_well=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            demonstrated_keenness_to_the_participant_needs=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness_or_suitability_of_service=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            overall_satisfaction=1
        ).count()

       
        template_name = 'pages/view_admin_webinar_evaluations.html'
    else:
        return HttpResponse("Invalid event type.")


    return render(request, template_name, {'event_form_details': event_form_details, 'faculty': faculty, 'questions': questions, 'excellent_count': excellent_count, 'very_satisfactory_count': very_satisfactory_count, 'satisfactory_count': satisfactory_count, 'fair_count': fair_count, 'poor_count': poor_count, 'is_admin': is_admin})


@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin', 'HR admin'])
def stakeholderevaluations(request):
    is_admin = request.user.groups.filter(name='admin').exists()
    evaluation_status = EvaluationStatus.objects.first()
    current_academic_year = evaluation_status.academic_year 
    selected_academic_year = request.GET.get('academic_year', current_academic_year)
    evaluations = StakeholderFeedbackModel.objects.all()
    

           #filter and search
    stakeholder_evaluation_filter = StakeholderFilter(request.GET, queryset=evaluations)
    evaluations = stakeholder_evaluation_filter.qs
    total_entries = evaluations.count()
       
        # ordering functionality
    ordering = request.GET.get('ordering', "")  
    if ordering:
        evaluations = evaluations.order_by(ordering) 
    
    paginator = Paginator(evaluations, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {'evaluations': evaluations, 'page_obj': page_obj, 'stakeholder_evaluation_filter': stakeholder_evaluation_filter, 'is_admin': is_admin, 'selected_academic_year': selected_academic_year, 'total_entries': total_entries}
    return render(request, 'pages/stakeholderevaluations.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin', 'HR admin'])
def stakeholder_evaluations_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=stakeholder_evaluations.xlsx'

    # Create a workbook and a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Stakeholder Evaluations"

    # Apply filters from the EvaluationFilter based on the request data
    evaluation_filter = StakeholderFilter(request.GET, queryset=StakeholderFeedbackModel.objects.all())

    # Get the filtered queryset
    filtered_evaluations = evaluation_filter.qs

    # Add column headings to the worksheet
    columns = [
        'Name', 'Agency', 'Email Address', 'Purpose of Visit', 'Date of Visit',
        'Attending Staff', 'Courtesy', 'Quality', 'Timeliness', 'Efficiency',
        'Cleanliness', 'Comfort', 'Average', 'Rating', 'Comments/Suggestions',
        'Predicted Sentiment', 'Academic Year', 'Semester'
    ]
    ws.append(columns)

    # Loop through and output the data
    for evaluation in filtered_evaluations:
        row = [
            str(evaluation.name),
            str(evaluation.agency),
            str(evaluation.email),
            str(evaluation.purpose),
            str(evaluation.date),
            str(evaluation.staff),
            str(evaluation.courtesy),
            str(evaluation.quality),
            str(evaluation.timeliness),
            str(evaluation.efficiency),
            str(evaluation.cleanliness),
            str(evaluation.comfort),
            str(evaluation.average_rating),
            str(evaluation.get_rating_category()),
            str(evaluation.suggestions_and_comments),
            str(evaluation.predicted_sentiment),
            str(evaluation.academic_year),
            str(evaluation.semester)
        ]
        ws.append(row)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2  # Add some padding
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook to a BytesIO stream
    output = BytesIO()
    wb.save(output)
    response.write(output.getvalue())

    return response

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin', 'HR admin'])
def admin_view_stakeholder_form(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    stakeholder_feedback_form = StakeholderFeedbackModel.objects.get(pk=pk)
    questions = StakeholderFeedbackQuestions.objects.all().order_by('order')

    highly_satisfied_count = StakeholderFeedbackModel.objects.filter(pk=pk).filter(
            courtesy=5
        ).count() + StakeholderFeedbackModel.objects.filter(pk=pk).filter(
            quality=5
        ).count() + StakeholderFeedbackModel.objects.filter(pk=pk).filter(pk=pk).filter(
            timeliness=5
        ).count() + StakeholderFeedbackModel.objects.filter(pk=pk).filter(
            efficiency=5
        ).count()  + StakeholderFeedbackModel.objects.filter(pk=pk).filter(
            cleanliness=5
        ).count()  + StakeholderFeedbackModel.objects.filter(pk=pk).filter(
            comfort=5
        ).count()
    
    very_satisfied_count = StakeholderFeedbackModel.objects.filter(pk=pk).filter(
            courtesy=4
        ).count() + StakeholderFeedbackModel.objects.filter(pk=pk).filter(
            quality=4
        ).count() + StakeholderFeedbackModel.objects.filter(pk=pk).filter(pk=pk).filter(
            timeliness=4
        ).count() + StakeholderFeedbackModel.objects.filter(pk=pk).filter(
            efficiency=4
        ).count()  + StakeholderFeedbackModel.objects.filter(pk=pk).filter(
            cleanliness=4
        ).count()  + StakeholderFeedbackModel.objects.filter(pk=pk).filter(
            comfort=4
        ).count()
    
    moderately_satisfied_count = StakeholderFeedbackModel.objects.filter(pk=pk).filter(
            courtesy=3
        ).count() + StakeholderFeedbackModel.objects.filter(pk=pk).filter(
            quality=3
        ).count() + StakeholderFeedbackModel.objects.filter(pk=pk).filter(pk=pk).filter(
            timeliness=3
        ).count() + StakeholderFeedbackModel.objects.filter(pk=pk).filter(
            efficiency=3
        ).count()  + StakeholderFeedbackModel.objects.filter(pk=pk).filter(
            cleanliness=3
        ).count()  + StakeholderFeedbackModel.objects.filter(pk=pk).filter(
            comfort=3
        ).count() 
    
    barely_satisfied_count = StakeholderFeedbackModel.objects.filter(pk=pk).filter(
            courtesy=2
        ).count() + StakeholderFeedbackModel.objects.filter(pk=pk).filter(
            quality=2
        ).count() + StakeholderFeedbackModel.objects.filter(pk=pk).filter(pk=pk).filter(
            timeliness=2
        ).count() + StakeholderFeedbackModel.objects.filter(pk=pk).filter(
            efficiency=2
        ).count()  + StakeholderFeedbackModel.objects.filter(pk=pk).filter(
            cleanliness=2
        ).count()  + StakeholderFeedbackModel.objects.filter(pk=pk).filter(
            comfort=2
        ).count() 
    
    not_satisfied_count = StakeholderFeedbackModel.objects.filter(pk=pk).filter(
            courtesy=1
        ).count() + StakeholderFeedbackModel.objects.filter(pk=pk).filter(
            quality=1
        ).count() + StakeholderFeedbackModel.objects.filter(pk=pk).filter(pk=pk).filter(
            timeliness=1
        ).count() + StakeholderFeedbackModel.objects.filter(pk=pk).filter(
            efficiency=1
        ).count()  + StakeholderFeedbackModel.objects.filter(pk=pk).filter(
            cleanliness=1
        ).count()  + StakeholderFeedbackModel.objects.filter(pk=pk).filter(
            comfort=1
        ).count()
    
    return render(request, 'pages/admin_view_stakeholder_form.html', {'stakeholder_feedback_form': stakeholder_feedback_form, 'questions': questions, 'highly_satisfied_count': highly_satisfied_count, 'very_satisfied_count': very_satisfied_count, 'moderately_satisfied_count': moderately_satisfied_count, 'barely_satisfied_count': barely_satisfied_count, 'not_satisfied_count': not_satisfied_count, 'is_admin': is_admin})


@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def forms(request):
    is_admin = request.user.groups.filter(name='admin').exists()
    context = {'is_admin': is_admin}
    return render(request, 'pages/forms.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def edit_faculty_evaluation_form(request):
    is_admin = request.user.groups.filter(name='admin').exists()
    questions = FacultyEvaluationQuestions.objects.all()
    return render(request, 'pages/edit_faculty_evaluation_form.html', {'questions': questions, 'is_admin': is_admin} )

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def edit_faculty_evaluation_form_question(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    question = FacultyEvaluationQuestions.objects.get(pk=pk)
    if request.method == 'POST':
        form = EditQuestionForm(request.POST, instance=question)  # Use EditQuestionForm
        if form.is_valid():
            form.save()
            messages.success(request, 'Form updated successfully')
            return redirect('edit_faculty_evaluation_form')  # Redirect to success view
    else:
        form = EditQuestionForm(instance=question)  # Pre-populate form with existing data
    context = {'form': form, 'is_admin': is_admin}
    return render(request, 'pages/edit_faculty_evaluation_form_question.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def edit_school_event_form(request):
    is_admin = request.user.groups.filter(name='admin').exists()
    questions = SchoolEventQuestions.objects.all()
    return render(request, 'pages/edit_school_event_form.html', {'questions': questions, 'is_admin': is_admin} )

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def edit_school_event_form_question(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    question = SchoolEventQuestions.objects.get(pk=pk)
    if request.method == 'POST':
        form = EditSchoolEventQuestionForm(request.POST, instance=question)  # Use EditQuestionForm
        if form.is_valid():
            form.save()
            messages.success(request, 'Form updated successfully')
            return redirect('edit_school_event_form')  # Redirect to success view
    else:
        form = EditSchoolEventQuestionForm(instance=question)  # Pre-populate form with existing data
    context = {'form': form, 'is_admin': is_admin}
    return render(request, 'pages/edit_school_event_form_question.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def edit_webinar_seminar_form(request):
    is_admin = request.user.groups.filter(name='admin').exists()
    questions = WebinarSeminarQuestions.objects.all()
    return render(request, 'pages/edit_webinar_seminar_form.html', {'questions': questions, 'is_admin': is_admin} )

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def edit_webinar_seminar_form_question(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    question = WebinarSeminarQuestions.objects.get(pk=pk)
    if request.method == 'POST':
        form = EditWebinarSeminarQuestionForm(request.POST, instance=question)  # Use EditQuestionForm
        if form.is_valid():
            form.save()
            messages.success(request, 'Form updated successfully')
            return redirect('edit_webinar_seminar_form')  # Redirect to success view
    else:
        form = EditWebinarSeminarQuestionForm(instance=question)  # Pre-populate form with existing data
    context = {'form': form, 'is_admin': is_admin}
    return render(request, 'pages/edit_webinar_seminar_form_question.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def edit_stakeholders_feedback_form(request):
    is_admin = request.user.groups.filter(name='admin').exists()
    questions = StakeholderFeedbackQuestions.objects.all()
    return render(request, 'pages/edit_stakeholders_feedback_form.html', {'questions': questions, 'is_admin': is_admin})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def edit_stakeholders_feedback_form_question(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    question = StakeholderFeedbackQuestions.objects.get(pk=pk)
    if request.method == 'POST':
        form = EditStakeholdersQuestionForm(request.POST, instance=question)  # Use EditQuestionForm
        if form.is_valid():
            form.save()
            messages.success(request, 'Form updated successfully')
            return redirect('edit_webinar_seminar_form')  # Redirect to success view
    else:
        form = EditStakeholdersQuestionForm(instance=question)  # Pre-populate form with existing data
    context = {'form': form, 'is_admin': is_admin}
    return render(request, 'pages/edit_stakeholders_feedback_form_question.html', context)


@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def faculty(request):
    is_admin = request.user.groups.filter(name='admin').exists()
    evaluation_status = EvaluationStatus.objects.first()
    current_academic_year = evaluation_status.academic_year 
    current_semester = evaluation_status.semester
    faculty = Faculty.objects.all().annotate(
    num_of_evaluators=Count(
        'sectionsubjectfaculty__likertevaluation',
        filter=Q(sectionsubjectfaculty__likertevaluation__academic_year=current_academic_year, sectionsubjectfaculty__likertevaluation__semester=current_semester)
    ),
    average_rating=Round(
        Coalesce(
            Avg(
                'sectionsubjectfaculty__likertevaluation__average_rating',
                filter=Q(sectionsubjectfaculty__likertevaluation__academic_year=current_academic_year, sectionsubjectfaculty__likertevaluation__semester=current_semester)
            ),
            Value(0.0, output_field=FloatField())
        ),
        2,
        output_field=FloatField()
    ),
).annotate(
    rating_category=Case(
        # If average_rating is null, return "Not yet evaluated"
        When(average_rating__isnull=True, then=Value("Not yet evaluated")),
        # If average_rating <= 1.99 -> "Poor"
        When(average_rating__lte=Value(1.99, output_field=FloatField()), then=Value("Poor")),
        # If 2.0 <= average_rating <= 2.99 -> "Unsatisfactory"
        When(
            average_rating__gte=Value(2.0, output_field=FloatField()),
            average_rating__lte=Value(2.99, output_field=FloatField()),
            then=Value("Unsatisfactory")
        ),
        # If 3.0 <= average_rating <= 3.99 -> "Satisfactory"
        When(
            average_rating__gte=Value(3.0, output_field=FloatField()),
            average_rating__lte=Value(3.99, output_field=FloatField()),
            then=Value("Satisfactory")
        ),
        # If 4.0 <= average_rating <= 4.99 -> "Very Satisfactory"
        When(
            average_rating__gte=Value(4.0, output_field=FloatField()),
            average_rating__lte=Value(4.99, output_field=FloatField()),
            then=Value("Very Satisfactory")
        ),
        # If average_rating >= 5.0 -> "Outstanding"
        When(
            average_rating__gte=Value(5.0, output_field=FloatField()),
            then=Value("Outstanding")
        ),
        default=Value("Not yet evaluated"),
        output_field=CharField()
    )
)
      
        #filtering functionality
    faculty_filter = FacultyFilter(request.GET, queryset=faculty)
    filtered_faculty = faculty_filter.qs

    # ordering functionality
   
    ordering = request.GET.get('ordering', "")
     
    if ordering:
        filtered_faculty = faculty.order_by(ordering) 

    form = TeacherForm()
    if request.method == 'POST':
        form = TeacherForm(request.POST, request.FILES)
             # Check for file upload
        created_count = 0
        updated_count = 0
        if form.is_valid():
            form.save()
            messages.success(request, 'Faculty added successfully')
            return redirect('faculty')
        import_faculty = request.FILES.get('facultyfile')

        if import_faculty and import_faculty.name.endswith('xlsx'):
            dataset = Dataset()
            imported_data = dataset.load(import_faculty.read(), format='xlsx')

            for data in imported_data:
                last_name = data[0]
                first_name = data[1]
                middle_name = data[2]
                gender = data[3]
                birthday = data[4]
                contact_number = data[5]
                email = data[6]
                employment_status = data[7]
                academic_rank = data[8]
                date_of_employment = data[9]
                years_in_service = data[10]
                department_name = data[11]
                no_of_workload = data[12]
                educational_attainment = data[13]
                eligibility = data[14]


                department = Department.objects.filter(name=department_name).first()
                if not department:
                    messages.error(request, f"Department with name '{department_name}' not found.")
                    continue  # Skip this row if course doesn't exist

     
                faculty, created = Faculty.objects.update_or_create(
                    email=email,  
                    defaults={
                        'first_name': first_name,
                        'last_name': last_name,
                        'middle_name': middle_name,
                        'gender': gender,
                        'birthday': birthday,
                        'contact_number': contact_number,
                        'employment_status': employment_status,
                        'academic_rank': academic_rank,
                        'date_of_employment': date_of_employment,
                        'years_in_service': years_in_service,
                        'department': department,
                        'no_of_workload': no_of_workload,
                        'educational_attainment': educational_attainment,
                        'eligibility': eligibility,
                    }
                )
                                # Increment counters based on whether the student was created or updated
                if created:
                    created_count += 1
                else:
                    updated_count += 1

            # After processing all data, show the success message
            if created_count > 0:
                messages.success(request, f"Created {created_count} new faculty(s).")
            if updated_count > 0:
                messages.info(request, f"Updated {updated_count} faculty(s).")

        else:
            pass
        

   
    paginator = Paginator(filtered_faculty, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {'faculty': faculty, 'form': form, 'page_obj': page_obj, 'faculty_filter': faculty_filter, 'is_admin': is_admin}

    return render(request, 'pages/faculty.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def facultyevaluations(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    teacher = get_object_or_404(Faculty, pk=pk)
    teacher_evaluations = LikertEvaluation.objects.filter(section_subject_faculty__faculty=teacher)


    context = {'teacher': teacher, 'teacher_evaluations':  teacher_evaluations, 'is_admin': is_admin}

    return render(request, 'pages/facultyevaluations.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def editteacher(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    faculty = Faculty.objects.get(pk=pk)
    form = TeacherForm(instance=faculty)
    if request.method == 'POST':
        form = TeacherForm(request.POST, instance=faculty)
        if form.is_valid():
            form.save()
            messages.success(request, 'Faculty updated successfully')
            return redirect('faculty')

    context = {'form':form, 'is_admin': is_admin}
    return render(request, 'pages/editteacher.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def send_message(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    user = request.user
    faculty = get_object_or_404(Faculty, pk=pk)
    faculty_user = faculty.user
    if request.method == 'POST':
        form = MessageForm(request.POST, request.FILES)
        if form.is_valid():
            message = form.save(commit=False)
            message.sender = user
            message.recipient = faculty_user
            message.save()

            # Send email with attachment if it exists
            email = EmailMessage(
                subject=f'New Message: {message.subject}',
                body=message.content,
                from_email='your-email@example.com',
                to=[faculty.email],
            )
            if message.attachment:
                email.attach(message.attachment.name, message.attachment.read())
            email.send()

            notify.send(
                sender=user,
                recipient=faculty_user,
                verb=message.subject,
                description=message.content
            )
            return redirect('faculty')
    else:
        form = MessageForm()
    return render(request, 'pages/send_message.html', {'form': form, 'faculty': faculty, 'is_admin': is_admin})


@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def deleteTeacher(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    faculty = Faculty.objects.get(pk=pk)
    if request.method == 'POST':
            faculty.delete()
            messages.success(request, 'Faculty deleted successfully')
            return redirect('faculty')

    return render(request, 'pages/delete.html', {'obj':faculty, 'is_admin': is_admin})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def department(request):
    is_admin = request.user.groups.filter(name='admin').exists()
    department = Department.objects.all()
    form = DepartmentForm()
    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Department added successfully')
            return redirect('department')
        
    paginator = Paginator(department, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {'department': department, 'form': form, 'page_obj': page_obj, 'is_admin': is_admin}

    return render(request, 'pages/departments.html', context)


@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def view_department(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    department = Department.objects.get(pk=pk)
    faculties = department.faculty_set.all()  # Retrieve all faculties in the department
    paginator = Paginator(faculties, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'department': department,
        'faculties': faculties,
        'page_obj': page_obj, 
        'is_admin': is_admin
    }

    return render(request, 'pages/view_department.html', context)


@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def edit_department(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    department = Department.objects.get(pk=pk)
    form = DepartmentForm(instance=department)
    if request.method == 'POST':
        form = DepartmentForm(request.POST, instance=department)
        if form.is_valid():
            form.save()
            messages.success(request, 'Department updated successfully')
            return redirect('department')

    context = {'form':form, 'is_admin': is_admin}
    return render(request, 'pages/edit_department.html', context)


@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def delete_department(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    department = Department.objects.get(pk=pk)
    if request.method == 'POST':
            department.delete()
            messages.success(request, 'Department deleted successfully')
            return redirect('department')

    return render(request, 'pages/delete.html', {'obj':department, 'is_admin': is_admin})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def students(request):
        is_admin = request.user.groups.filter(name='admin').exists()
        students = Student.objects.filter(is_active=True)
        student_filter = StudentFilter(request.GET, queryset=students)
        students = student_filter.qs
        form = StudentForm()

        evaluation_status = EvaluationStatus.objects.first() 
        current_academic_year = evaluation_status.academic_year 
        current_semester = evaluation_status.semester

        if request.method == 'POST':
            form = StudentForm(request.POST, request.FILES)
                # Initialize counters
            created_count = 0
            updated_count = 0
            if form.is_valid():
                form.save(commit=True)  # Ensure commit is set to True to save to the database
                messages.success(request, 'Student added successfully')
                return redirect('students')
            
            
        # Check for file upload
        new_student_file = request.FILES.get('studentfile')

        if new_student_file and new_student_file.name.endswith('xlsx'):
            dataset = Dataset()
            imported_data = dataset.load(new_student_file.read(), format='xlsx')
            
            for data in imported_data:
                student_number = data[0]
                last_name = data[1]
                first_name = data[2]
                middle_name = data[3]
                course_name = data[4]
                year = data[5]
                address = data[6]
                semester = data[7]
                school_year = data[8]  
                date_enrolled = data[9]  
                major = data[10]  
                section_name = data[11]  
                old_or_new_student = data[12]  
                status = data[13]  
                birthdate = data[14]  
                gender = data[15]  
                contact_no = data[16]  
                email = data[17]  




                course = Course.objects.filter(name=course_name).first()
                section = Section.objects.filter(name=section_name).first()

                if not course:
                    messages.error(request, f"Course with name '{course_name}' not found.")
                    break  # Skip this row if course doesn't exist

                if not section:
                    messages.error(request, f"Section with name '{section_name}' not found.")
                    break  # Skip this row if section doesn't exist

                # Update the existing student or create a new one
                student, created = Student.objects.update_or_create(
                    student_number=student_number,
                    defaults={
                        'first_name': first_name,
                        'last_name': last_name,
                        'middle_name': middle_name,
                        'email': email,
                        'year': year,  
                        'address': address,
                        'semester': semester,
                        'school_year': school_year,
                        'date_enrolled': date_enrolled,
                        'major': major,
                        'Section': section,
                        'old_or_new_student': old_or_new_student,
                        'status': status,
                        'birthdate': birthdate,
                        'gender': gender,
                        'contact_no': contact_no,
                        'Course': course,
                    }
                )
                                # Increment counters based on whether the student was created or updated
                if created:
                    created_count += 1
                else:
                    updated_count += 1

            # After processing all data, show the success message
            if created_count > 0:
                messages.success(request, f"Created {created_count} new student(s).")
            if updated_count > 0:
                messages.info(request, f"Updated {updated_count} student(s).")


       
        paginator = Paginator(students, 20) 
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context = {'students': students, 'form': form, 'page_obj': page_obj, 'student_filter': student_filter, 'is_admin': is_admin}
        return render(request, 'pages/students.html',  context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])   
def archive_student(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    student = get_object_or_404(Student, pk=pk)
    if request.method == "POST":
        student.is_active = False
        student.save()
        if student.user:
            student.user.is_active = False
            student.user.save()

        messages.success(request, f"Student {student.full_name()} has been archived.")
        return redirect('students')
    
    return render(request, 'pages/archive.html', {'obj':student, 'is_admin': is_admin})
    
@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])   
def unarchive_student(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    student = get_object_or_404(Student, pk=pk)
    if request.method == "POST":
        student.is_active = True
        student.save()
        if student.user:
            student.user.is_active = True
            student.user.save()

        messages.success(request, f"Student {student.full_name()} has been unarchived.")
        return redirect('archived_students')
    
    return render(request, 'pages/unarchive.html', {'obj':student, 'is_admin': is_admin})
    


@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])    
def archived_students(request):
    is_admin = request.user.groups.filter(name='admin').exists()
    archived_students = Student.objects.filter(is_active=False)
    student_filter = StudentFilter(request.GET, queryset=archived_students)
    archived_students = student_filter.qs

    paginator = Paginator(archived_students, 20) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'is_admin': is_admin,
        'archived_students': archived_students,
        'student_filter': student_filter,
        'page_obj': page_obj
    }

    return render(request, 'pages/archived_students.html', context)


@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def editstudent(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    student = Student.objects.get(pk=pk)
    form = StudentForm(instance=student)
    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, 'Student updated successfully')
            return redirect('students')

    context = {'form':form, 'is_admin': is_admin}
    return render(request, 'pages/editstudent.html', context)


@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def deleteStudent(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    student = Student.objects.get(pk=pk)
    if request.method == 'POST':
            student.delete()
            messages.success(request, 'Student deleted successfully')
            return redirect('students')

    return render(request, 'pages/delete.html', {'obj':student, 'is_admin': is_admin})


@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def courses(request):
    is_admin = request.user.groups.filter(name='admin').exists()
    course = Course.objects.all()
    student = Student.objects.all()
    
    total_students = student.count()

    form = CourseForm()
    if request.method == 'POST':
        form = CourseForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Course added successfully')
            return redirect('courses')
    context = {'course': course,
               'total_students': total_students,
              'student': student,
              'form':form, 
              'is_admin': is_admin}
   

   
    return render(request, 'pages/courses.html',  context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def editcourse(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    course = Course.objects.get(pk=pk)

    form = CourseForm(instance=course)
    if request.method == 'POST':
        form = CourseForm(request.POST, instance=course)
        if form.is_valid():
            form.save()
            messages.success(request, 'Course updated successfully')
            return redirect('courses')

    context = {'form':form, 'course': course, 'is_admin': is_admin}
    return render(request, 'pages/editcourse.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def deleteCourse(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    course = Course.objects.get(pk=pk)
    if request.method == 'POST':
            course.delete()
            messages.success(request, 'Course deleted successfully')
            return redirect('courses')

    return render(request, 'pages/delete.html', {'obj':course, 'is_admin': is_admin})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def sections(request):
    is_admin = request.user.groups.filter(name='admin').exists()
    sections = Section.objects.all()

    section_filter = SectionFilter(request.GET, queryset=sections)
    sections = section_filter.qs
    # Create a list to store section data along with student counts
    section_data = []
    
    for section in sections:
        student_count = Student.objects.filter(Section=section).count()  # Count students in each section
        subject_count = SectionSubjectFaculty.objects.filter(section=section).count()
        section_data.append({
            'section': section,
            'student_count': student_count,
            'subject_count': subject_count
        })

    form = SectionForm()
    if request.method == 'POST':
        form = SectionForm(request.POST)
        if form.is_valid():
            section_name = form.cleaned_data['name']  # Assuming 'name' is the field in your form
            # Check if the section already exists
            if Section.objects.filter(name=section_name).exists():
                messages.warning(request, f"The section '{section_name}' already exists.")
            else:
                form.save()
                messages.success(request, f"The section '{section_name}' has been created successfully.")
                return redirect('sections')
            
    import_subject_faculty = request.FILES.get('subjectfacultyfile')

    if import_subject_faculty and import_subject_faculty.name.endswith('xlsx'):
            dataset = Dataset()
            imported_data = dataset.load(import_subject_faculty.read(), format='xlsx')
            created_count = 0 
            updated_count = 0 
            sections_with_new_subjects_faculty = set() 
            sections_with_updated_subjects_faculty = set()

            for data in imported_data:
                section_name = data[0]
                subject_name = data[1]
                faculty_full_name = data[2]  # Concatenated first and last name

                section = Section.objects.filter(name=section_name).first()
                subject = Subject.objects.filter(subject_name=subject_name).first()

                # Use Concat function to concatenate first_name and last_name in the filter
                faculty = Faculty.objects.annotate(
                    full_name=Concat('first_name', Value(' '), 'last_name')
                ).filter(full_name=faculty_full_name).first()


                if not section:
                            messages.error(request, f"Section with name '{section_name}' not found.")
                            continue
                
                if not subject:
                    messages.error(request, f"Subject with name '{subject_name}' not found.")
                    continue  

                
                
                if not faculty:
                    messages.error(request, f"Faculty with name '{faculty_full_name}' not found.")
                    continue  
                
                sub_faculty, created = SectionSubjectFaculty.objects.update_or_create(
                section=section,
                subjects=subject,
                defaults={
                    'faculty': faculty
                }
            )
                if created:
                    created_count += 1
                    sections_with_new_subjects_faculty.add(section.name)
                else:
                    updated_count += 1
                    sections_with_updated_subjects_faculty.add(section.name)
                

            new_sections_count = len(sections_with_new_subjects_faculty) 
            updated_sections_count = len(sections_with_updated_subjects_faculty)
            # After processing all data, show the success message
            if created_count > 0:
                messages.success(request, f"Created {created_count} new subject(s) for {new_sections_count} section(s).")
            if updated_count > 0:
                messages.info(request, f"Updated {updated_count} subject(s) for {updated_sections_count} section(s).")
        
    # Ordering functionality
    ordering = request.GET.get('ordering', "")

    if ordering == "name":  
        section_data = sorted(section_data, key=lambda x: x['section'].name)
    elif ordering == "-name":
        section_data = sorted(section_data, key=lambda x: x['section'].name, reverse=True)
    elif ordering == "student_count":
        section_data = sorted(section_data, key=lambda x: x['student_count'])
    elif ordering == "-student_count":
        section_data = sorted(section_data, key=lambda x: x['student_count'], reverse=True)

    paginator = Paginator(section_data, 10) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    
    context = {'sections': sections, 'form':form, 'page_obj': page_obj, 'section_filter': section_filter, 'is_admin': is_admin }
   
    return render(request, 'pages/sections.html',  context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def editsection(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    section = Section.objects.get(pk=pk)

    form = SectionForm(instance=section)
    if request.method == 'POST':
        form = SectionForm(request.POST, instance=section)
        if form.is_valid():
            form.save()
            messages.success(request, 'Section updated successfully')
            return redirect('sections')

    context = {'form':form, 'is_admin': is_admin}
    return render(request, 'pages/editsection.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def deleteSection(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    section = Section.objects.get(pk=pk)
    if request.method == 'POST':
            section.delete()
            messages.success(request, 'Section deleted successfully')
            return redirect('sections')

    return render(request, 'pages/delete.html', {'obj':section, 'is_admin': is_admin})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def section_details(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    section = Section.objects.get(pk=pk)
    section = get_object_or_404(Section, pk=pk) #get the primary key of the selected section
    form = SectionSubjectFacultyForm(initial={'section': section}) # pass the instance of the section attribute to the section that is selected to add subjects
    if request.method == 'POST':
        form = SectionSubjectFacultyForm(request.POST, request.FILES)
        created_count = 0
        updated_count = 0
         # Check for file upload
        if form.is_valid():
            form.save(commit=True)
            messages.success(request, 'Subject and Faculty added successfully')
            return redirect('section_details', pk=pk)
    subjects_faculty = SectionSubjectFaculty.objects.filter(section=section)
    return render(request, 'pages/section_details.html', {'section': section, 'subjects_faculty': subjects_faculty, 'form': form, 'is_admin': is_admin})
    
@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def deleteSub_Section(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    subject = SectionSubjectFaculty.objects.get(pk=pk)
    section_pk = subject.section.pk  # Get the primary key of the related section
    
    if request.method == 'POST':
            subject.delete()
            messages.success(request, 'Subject and Faculty deleted successfully')
            return redirect('section_details', pk=section_pk)

    return render(request, 'pages/delete.html', {'obj':subject, 'is_admin': is_admin})


@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def subjects(request):
    is_admin = request.user.groups.filter(name='admin').exists()
    subject = Subject.objects.all()

    #filtering functionality
    subject_filter = SubjectFilter(request.GET, queryset=subject)
    subject = subject_filter.qs

    # ordering functionality  
    ordering = request.GET.get('ordering', "")

     
    if ordering:
        subject = subject.order_by(ordering) 

    #add subject form
    form = SubjectForm()
    if request.method == 'POST':
        form = SubjectForm(request.POST)

        if form.is_valid():
            form.save()
            messages.success(request, 'Subject added successfully')
            return redirect('subjects')
  

    
    import_subjects = request.FILES.get('subjectfile')

    if import_subjects and import_subjects.name.endswith('xlsx'):
        dataset = Dataset()
        imported_data = dataset.load(import_subjects.read(), format='xlsx')
        created_count = 0
        updated_count = 0
        for data in imported_data:
            subject_code = data[0]
            subject_name = data[1]

            # Update the existing student or create a new one
            subject_instance, created = Subject.objects.update_or_create(
                subject_code=subject_code,
                defaults={
                    'subject_name': subject_name,
                }
            )

            if created:
                created_count += 1
            else:
                updated_count += 1

        # After processing all data, show the success message
        if created_count > 0:
            messages.success(request, f"Created {created_count} new subject(s).")
        if updated_count > 0:
            messages.info(request, f"Updated {updated_count} subject(s).")


            
    paginator = Paginator(subject, 20) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {'form': form, 'page_obj': page_obj, 'subject_filter': subject_filter,
            'is_admin': is_admin}
    return render(request, 'pages/subjects.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def editsubject(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    subject = Subject.objects.get(pk=pk)

    form = SubjectForm(instance=subject)
    if request.method == 'POST':
        form = SubjectForm(request.POST, instance=subject)
        if form.is_valid():
            form.save()
            messages.success(request, 'Subject updated successfully')
            return redirect('subjects')

    context = {'form':form,
            'is_admin': is_admin}
    return render(request, 'pages/editsubject.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def deleteSubject(request, pk):
    is_admin = request.user.groups.filter(name='admin').exists()
    subject = Subject.objects.get(pk=pk)
    if request.method == 'POST':
            subject.delete()
            messages.success(request, 'Subject deleted successfully')
            return redirect('subjects')

    return render(request, 'pages/delete.html', {'obj':subject,
            'is_admin': is_admin})


@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def users(request):
    is_admin = request.user.groups.filter(name='admin').exists()
# Get all users
    users = User.objects.filter(is_active=True).order_by('-date_joined')
    user_filter = UserFilter(request.GET, queryset=users)
    users = user_filter.qs

        # Annotate users with full names from Student or Faculty
    users = users.annotate(
        student_name=Concat(
            'student__first_name', Value(' '), 'student__last_name',
            output_field=CharField()
        ),
        faculty_name=Concat(
            'faculty__first_name', Value(' '), 'faculty__last_name',
            output_field=CharField()
        )
    )

    # Create a dictionary to store user groups
    user_groups = {}

    # Iterate over each user
    for user in users:
        # Get the groups the user belongs to
        groups = user.groups.all()
        
        # Store user groups in the dictionary with user id as key
        user_groups[user.username] = groups


        

    paginator = Paginator(users, 10) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {'user_groups': user_groups, 'page_obj': page_obj, 'user_filter': user_filter,
            'is_admin': is_admin}
    return render(request, 'pages/users.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def edit_user_group(request, user_id):
    is_admin = request.user.groups.filter(name='admin').exists()
    user = User.objects.get(username=user_id)
    user_groups = user.groups.all()

    # Define the allowed groups for the user
    if user_groups.filter(name__in=['student', 'society president']).exists():
        allowed_groups = Group.objects.filter(name__in=['student', 'society president'])
        allow_multiple_selection = False  # Only single selection allowed for students
    elif user_groups.filter(name__in=['faculty', 'head of OSAS', 'department head/program coordinator']).exists():
        allowed_groups = Group.objects.filter(name__in=['faculty', 'head of OSAS', 'department head/program coordinator'])
        allow_multiple_selection = True  # Multiple selection allowed for faculty and head of OSAS
    elif user_groups.filter(name__in=['admin', 'HR admin']).exists():
        allowed_groups = Group.objects.filter(name__in=['admin', 'HR admin'])
        allow_multiple_selection = False  # Only single selection allowed for admins
    else:
        # Handle any other cases as needed
        allowed_groups = Group.objects.all()
        allow_multiple_selection = False

    if request.method == 'POST':
        if allow_multiple_selection:
            group_names = request.POST.getlist('group_name')  # Get multiple selected groups
            groups = Group.objects.filter(name__in=group_names)  # Retrieve the groups from the database
        else:
            group_name = request.POST.get('group_name')  # Get a single group name
            groups = Group.objects.filter(name=group_name)  # Retrieve the group

        user.groups.clear()  # Clear existing groups
        user.groups.add(*groups)  # Add the selected groups

        messages.success(request, 'User role updated successfully.')
        return redirect('users')  # Redirect to the users page after updating groups

    else:
        context = {
            'user_id': user_id,
            'user_groups': user_groups,
            'allowed_groups': allowed_groups,
            'allow_multiple_selection': allow_multiple_selection,
            'is_admin': is_admin
        }
        return render(request, 'pages/edit_user_group.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['HR admin'])
def hr_dashboard(request):
    user=request.user
    # Get the current month and year
    current_month = datetime.now().month
    current_year = datetime.now().year
    current_month_name = datetime.now().strftime('%B')  # Example: 'November'r

    months = [datetime(current_year, i, 1).strftime('%B') for i in range(1, 13)]
    evaluation_status = EvaluationStatus.objects.first()
    current_academic_year = evaluation_status.academic_year 
    current_semester = evaluation_status.semester
    current_evaluation_status = evaluation_status.evaluation_status
        # Initialize the filterset and filter data
    filterset = StakeholderFilter(request.GET, queryset=StakeholderFeedbackModel.objects.all())
    if not request.GET or (request.GET.get('academic_year') == '' and request.GET.get('semester') == ''):
        data = StakeholderFeedbackModel.objects.filter(date__month=current_month)
    else:
        data = filterset.qs

    total_clients = data.count()
        # Compute evaluation counts
    positive_evaluations = data.filter(predicted_sentiment='Positive').count()
    negative_evaluations = data.filter(predicted_sentiment='Negative').count()
    neutral_evaluations = data.filter(predicted_sentiment='Neutral').count()
    total_evaluations = data.count()

    # Sentiment score calculation
    sentiment_score = (positive_evaluations - negative_evaluations) / total_evaluations if total_evaluations else 0
    rounded_sentiment_score = round(sentiment_score, 2)
    # Average rating calculation
    avg_rating = data.aggregate(average_rating=Avg('average_rating'))['average_rating']
    avg_rating = round(avg_rating, 2) if avg_rating else None

      # Fetch recent comments
    recent_comments = data.values('suggestions_and_comments', 'predicted_sentiment')[:3]

        # Calculate average ratings for each agency
    agencies = (
        data.values('agency__name')
        .annotate(avg_rating=Avg('average_rating'))
        .order_by('-avg_rating')
    )

    agency_ratings = {
        agency['agency__name']: round(agency['avg_rating'], 1)
        for agency in agencies if agency['avg_rating'] is not None
    }
 

     # Fetch evaluations and categorize them
    # positive_wordclouds = StakeholderFeedbackModel.objects.filter(predicted_sentiment='Positive')
    # negative_wordclouds = StakeholderFeedbackModel.objects.filter(predicted_sentiment='Negative')
    # neutral_wordclouds = StakeholderFeedbackModel.objects.filter(predicted_sentiment='Neutral')
    # all_wordclouds = StakeholderFeedbackModel.objects.all()

    # Combine texts for word clouds
    # positive_text = " ".join([eval.suggestions_and_comments for eval in positive_wordclouds])
    # negative_text = " ".join([eval.suggestions_and_comments for eval in negative_wordclouds])
    # neutral_text = " ".join([eval.suggestions_and_comments for eval in neutral_wordclouds])
    # all_text = " ".join([eval.suggestions_and_comments for eval in all_wordclouds])

    # Generate word clouds
    # wordclouds = {
    #     'Positive': generate_wordcloud(positive_text, colormap='Greens'),
    #     'Negative': generate_wordcloud(negative_text, colormap='Reds'),
    #     'Neutral': generate_wordcloud(neutral_text, colormap='Blues'),
    #     'All': generate_wordcloud(all_text, colormap='viridis'),
    # }
    context = {
        'user': user,
        'evaluation_status': evaluation_status,
        'current_evaluation_status': current_evaluation_status,
        'current_academic_year': current_academic_year,
        'current_semester': current_semester,
        'current_month': current_month,
        'data': data,
        'total_clients': total_clients,
        'months': months,
        'filterset': filterset,
        'avg_rating': avg_rating,
        'rounded_sentiment_score': rounded_sentiment_score,
        'avg_rating': avg_rating,
        'current_month_name': current_month_name,
        'current_year': current_year,
        'recent_comments': recent_comments,
        'positive_evaluations': positive_evaluations,
        'negative_evaluations': negative_evaluations,
        'neutral_evaluations': neutral_evaluations,
        'agency_ratings': agency_ratings,
    }

    return render(request, 'pages/hr_dashboard.html', context)

def link_callback(uri, rel):
    """ Convert HTML URIs to absolute system paths so xhtml2pdf can access those resources """
    if uri.startswith(settings.MEDIA_URL):
        path = os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ""))
    elif uri.startswith(settings.STATIC_URL):
        path = os.path.join(settings.BASE_DIR, 'Feedbacksystem/static', uri.replace(settings.STATIC_URL, ""))
    else:
        raise Exception(f"Unsupported URI: {uri}. It must start with {settings.MEDIA_URL} or {settings.STATIC_URL}.")
    
    if not os.path.isfile(path):
        raise Exception(f"File does not exist: {path}")
    return path

def stakeholders_generate_summary_report(request):
    months = request.GET.getlist('months')
    filtered_evaluations = StakeholderFeedbackModel.objects.all()

    if months:
        months = [int(month) for month in months]  # Convert months to integers
        filtered_evaluations = filtered_evaluations.filter(date__month__in=months)

        if not filtered_evaluations.exists(): 
            messages.error(request, 'No evaluations found for the selected month(s).') 
            return redirect('hr_dashboard')

        if len(months) == 1:
            year = filtered_evaluations.first().date.year if filtered_evaluations.exists() else ''
            title_period = f"Stakeholders' Feedback Monthly Report\n{calendar.month_name[months[0]]} {year}"
        elif len(months) == 3:
            start_month = calendar.month_name[months[0]]
            end_month = calendar.month_name[months[-1]]
            year = filtered_evaluations.first().date.year if filtered_evaluations.exists() else ''
            title_period = f"Stakeholders' Feedback Quarterly Report\n{start_month} - {end_month} {year}"
        elif len(months) == 12:
            year = filtered_evaluations.first().date.year if filtered_evaluations.exists() else ''
            title_period = f"Stakeholders' Feedback Yearly Report - {year}"
        
        else:
            start_month = calendar.month_name[months[0]]
            end_month = calendar.month_name[months[-1]]
            year = filtered_evaluations.first().date.year if filtered_evaluations.exists() else ''
            title_period = f"Stakeholders' Feedback Summary Report\n{start_month} - {end_month} {year}"

    else:
        messages.error(request, 'Please select the month/s to generate report.')
        return redirect('hr_dashboard')

    summary_data = {}
    category_totals = {
        'courtesy': 0,
        'quality': 0,
        'timeliness': 0,
        'efficiency': 0,
        'cleanliness': 0,
        'comfort': 0,
    }
    total_clients = filtered_evaluations.count()

    for evaluation in filtered_evaluations:
        agency = evaluation.agency.name
        if agency not in summary_data:
            summary_data[agency] = {
                'courtesy': 0,
                'quality': 0,
                'timeliness': 0,
                'efficiency': 0,
                'cleanliness': 0,
                'comfort': 0,
                'total': 0
            }
        summary_data[agency]['courtesy'] += evaluation.courtesy
        summary_data[agency]['quality'] += evaluation.quality
        summary_data[agency]['timeliness'] += evaluation.timeliness
        summary_data[agency]['efficiency'] += evaluation.efficiency
        summary_data[agency]['cleanliness'] += evaluation.cleanliness
        summary_data[agency]['comfort'] += evaluation.comfort
        summary_data[agency]['total'] += 1

        category_totals['courtesy'] += evaluation.courtesy
        category_totals['quality'] += evaluation.quality
        category_totals['timeliness'] += evaluation.timeliness
        category_totals['efficiency'] += evaluation.efficiency
        category_totals['cleanliness'] += evaluation.cleanliness
        category_totals['comfort'] += evaluation.comfort

    category_averages = {k: round(v / total_clients, 2) for k, v in category_totals.items()}
    overall_average = round(sum(category_averages.values()) / len(category_averages), 2)

    satisfaction_level = get_satisfaction_level(overall_average)

    for agency, data in summary_data.items():
        data['average'] = round((data['courtesy'] + data['quality'] + data['timeliness'] + data['efficiency'] + data['cleanliness'] + data['comfort']) / 6, 2)
        data['satisfaction_level'] = get_satisfaction_level(data['average'])

    html = render_to_string('pages/stakeholders_generate_summary_report.html', {
        'title_period': title_period,
        'summary_data': summary_data,
        'total_clients': total_clients,
        'category_averages': category_averages,
        'overall_average': overall_average,
        'satisfaction_level': satisfaction_level,
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=Stakeholders-Feedback-Summary-Report.pdf'
    pisa_status = pisa.CreatePDF(html, dest=response,
                                  default_font=DEFAULT_FONT, 
                                  encoding='utf-8', 
                                  link_callback=link_callback, 
                                  how_error_as_pdf=False, 
                                  context=None, 
                                  default_css=None,
                                  path=None, 
                                  pagesize='landscape' )

    if pisa_status.err:
        return HttpResponse('We had some errors with generating your PDF.')

    return response

def get_satisfaction_level(average):
    if average >= 4.5:
        return 'Highly Satisfied'
    elif average >= 3.5:
        return 'Very Satisfied'
    elif average >= 2.5:
        return 'Moderately Satisfied'
    elif average >= 1.5:
        return 'Barely Satisfied'
    else:
        return 'Not Satisfied'    


def adminlogout(request):
    logout(request)
    return redirect('adminlogin')

@login_required(login_url='signin')
@allowed_users(allowed_roles=['admin'])
def delete_user(request, user_id):
    is_admin = request.user.groups.filter(name='admin').exists()
    user = User.objects.get(username=user_id)
    if request.method == 'POST':
            user.delete()
            return redirect('users')

    return render(request, 'pages/delete.html', {'obj':user, 'is_admin': is_admin})


            # ------------------------------------------------------
            #                Faculty-Page Views
            # ------------------------------------------------------


@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty', 'head of OSAS', 'department head/program coordinator'])
def facultydashboard(request):
    is_head_of_osas = request.user.groups.filter(name='head of OSAS').exists() 
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()
    user=request.user
    faculty = Faculty.objects.filter(email=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    # Filter for unread event notifications (level=success) and unread message notifications (level=info)
    unread_event_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_message_notifications = Notification.objects.filter(recipient=user, level='info', unread=True)

    # Count the unread notifications
    unread_event_notifications_count = unread_event_notifications.count()
    unread_message_notifications_count = unread_message_notifications.count()

    # Combine both counts for the total unread notifications
    total_unread_notifications_count = unread_event_notifications_count + unread_message_notifications_count
    # Combine and order the notifications
    all_notifications = Notification.objects.filter(
        Q(level='success') | Q(level='info') | Q(level='warning')
    ).filter(recipient=user).order_by('-timestamp')[:5]


    evaluation_status = EvaluationStatus.objects.first()
    current_academic_year = evaluation_status.academic_year 
    current_semester = evaluation_status.semester
    current_evaluation_status = evaluation_status.evaluation_status

    

    # Initialize the filterset and filter data
    filterset = LikertEvaluationFilter(request.GET, queryset=LikertEvaluation.objects.filter(section_subject_faculty__faculty=faculty, admin_status='Approved'))
    if not request.GET or (request.GET.get('academic_year') == '' and request.GET.get('semester') == ''):
        data = LikertEvaluation.objects.filter(section_subject_faculty__faculty=faculty, 
                                               academic_year=current_academic_year, semester=current_semester, admin_status='Approved')
        is_filtered = False
    else:
        data = filterset.qs
        is_filtered = True

    # Get the filtered academic year and semester
    filtered_academic_year = request.GET.get('academic_year', current_academic_year)
    filtered_semester = request.GET.get('semester', current_semester)

    # Compute evaluation counts
    positive_evaluations = data.filter(predicted_sentiment='Positive').count()
    negative_evaluations = data.filter(predicted_sentiment='Negative').count()
    neutral_evaluations = data.filter(predicted_sentiment='Neutral').count()
    total_evaluations = data.count()

    # Sentiment score calculation
    sentiment_score = (positive_evaluations - negative_evaluations) / total_evaluations if total_evaluations else 0
    rounded_sentiment_score = round(sentiment_score, 1)
    # Average rating calculation
    avg_rating = data.aggregate(average_rating=Avg('average_rating'))['average_rating']
    avg_rating = round(avg_rating, 1) if avg_rating else None

    # Fetch recent comments
    recent_comments = data.values('comments', 'predicted_sentiment')[:3]

   
     # Categorize the evaluations for Chart.js
    categories = {
        'Subject Matter Content': {
            'command_and_knowledge_of_the_subject': 0,
            'depth_of_mastery': 0,
            'practice_in_respective_discipline': 0,
            'up_to_date_knowledge': 0,
            'integrates_subject_to_practical_circumstances': 0,
        },
        'Organization': {
            'organizes_the_subject_matter': 0,
            'provides_orientation_on_course_content': 0,
            'efforts_of_class_preparation': 0,
            'summarizes_main_points': 0,
            'monitors_online_class': 0,
        },
        'Teacher-Student Rapport': {
            'holds_interest_of_students': 0,
            'provides_relevant_feedback': 0,
            'encourages_participation': 0,
            'shows_enthusiasm': 0,
            'shows_sense_of_humor': 0,
        },
        'Teaching Methods': {
            'teaching_methods': 0,
            'flexible_learning_strategies': 0,
            'student_engagement': 0,
            'clear_examples': 0,
            'focused_on_objectives': 0,
        },
        'Presentation': {
            'starts_with_motivating_activities': 0,
            'speaks_in_clear_and_audible_manner': 0,
            'uses_appropriate_medium_of_instruction': 0,
            'establishes_online_classroom_environment': 0,
            'observes_proper_classroom_etiquette': 0,
        },
        'Classroom Management': {
            'uses_time_wisely': 0,
            'gives_ample_time_for_students_to_prepare': 0,
            'updates_the_students_of_their_progress': 0,
            'demonstrates_leadership_and_professionalism': 0,
            'understands_possible_distractions': 0,
        },
        'Sensitivity and Support to Students': {
            'sensitivity_to_student_culture': 0,
            'responds_appropriately': 0,
            'assists_students_on_concerns': 0,
            'guides_the_students_in_accomplishing_tasks': 0,
            'extends_consideration_to_students': 0,
        }
    }

    count = {key: 0 for key in categories}  # Track rating counts per category

    # Populate category data from evaluations
    for evaluation in data:
        # Subject Matter Content
        categories['Subject Matter Content']['command_and_knowledge_of_the_subject'] += evaluation.command_and_knowledge_of_the_subject
        categories['Subject Matter Content']['depth_of_mastery'] += evaluation.depth_of_mastery
        categories['Subject Matter Content']['practice_in_respective_discipline'] += evaluation.practice_in_respective_discipline
        categories['Subject Matter Content']['up_to_date_knowledge'] += evaluation.up_to_date_knowledge
        categories['Subject Matter Content']['integrates_subject_to_practical_circumstances'] += evaluation.integrates_subject_to_practical_circumstances
        
        count['Subject Matter Content'] += 5
        
        # Organization
        categories['Organization']['organizes_the_subject_matter'] += evaluation.organizes_the_subject_matter
        categories['Organization']['provides_orientation_on_course_content'] += evaluation.provides_orientation_on_course_content
        categories['Organization']['efforts_of_class_preparation'] += evaluation.efforts_of_class_preparation
        categories['Organization']['summarizes_main_points'] += evaluation.summarizes_main_points
        categories['Organization']['monitors_online_class'] += evaluation.monitors_online_class
        
        count['Organization'] += 5

       
        # Teacher-Student Rapport
        categories['Teacher-Student Rapport']['holds_interest_of_students'] += evaluation.holds_interest_of_students
        categories['Teacher-Student Rapport']['provides_relevant_feedback'] += evaluation.provides_relevant_feedback
        categories['Teacher-Student Rapport']['encourages_participation'] += evaluation.encourages_participation
        categories['Teacher-Student Rapport']['shows_enthusiasm'] += evaluation.shows_enthusiasm
        categories['Teacher-Student Rapport']['shows_sense_of_humor'] += evaluation.shows_sense_of_humor
        
        count['Teacher-Student Rapport'] += 5

       
        # Teaching Methods
        categories['Teaching Methods']['teaching_methods'] += evaluation.teaching_methods
        categories['Teaching Methods']['flexible_learning_strategies'] += evaluation.flexible_learning_strategies
        categories['Teaching Methods']['student_engagement'] += evaluation.student_engagement
        categories['Teaching Methods']['clear_examples'] += evaluation.clear_examples
        categories['Teaching Methods']['focused_on_objectives'] += evaluation.focused_on_objectives
        
        count['Teaching Methods'] += 5

       
        # Presentation
        categories['Presentation']['starts_with_motivating_activities'] += evaluation.starts_with_motivating_activities
        categories['Presentation']['speaks_in_clear_and_audible_manner'] += evaluation.speaks_in_clear_and_audible_manner
        categories['Presentation']['uses_appropriate_medium_of_instruction'] += evaluation.uses_appropriate_medium_of_instruction
        categories['Presentation']['establishes_online_classroom_environment'] += evaluation.establishes_online_classroom_environment
        categories['Presentation']['observes_proper_classroom_etiquette'] += evaluation.observes_proper_classroom_etiquette
        
        count['Presentation'] += 5

       
        # Classroom Management
        categories['Classroom Management']['uses_time_wisely'] += evaluation.uses_time_wisely
        categories['Classroom Management']['gives_ample_time_for_students_to_prepare'] += evaluation.gives_ample_time_for_students_to_prepare
        categories['Classroom Management']['updates_the_students_of_their_progress'] += evaluation.updates_the_students_of_their_progress
        categories['Classroom Management']['demonstrates_leadership_and_professionalism'] += evaluation.demonstrates_leadership_and_professionalism
        categories['Classroom Management']['understands_possible_distractions'] += evaluation.understands_possible_distractions
        
        count['Classroom Management'] += 5

       
        # Sensitivity and Support to Students
        categories['Sensitivity and Support to Students']['sensitivity_to_student_culture'] += evaluation.sensitivity_to_student_culture
        categories['Sensitivity and Support to Students']['responds_appropriately'] += evaluation.responds_appropriately
        categories['Sensitivity and Support to Students']['assists_students_on_concerns'] += evaluation.assists_students_on_concerns
        categories['Sensitivity and Support to Students']['guides_the_students_in_accomplishing_tasks'] += evaluation.guides_the_students_in_accomplishing_tasks
        categories['Sensitivity and Support to Students']['extends_consideration_to_students'] += evaluation.extends_consideration_to_students
        
        count['Sensitivity and Support to Students'] += 5

       

    # Average each field
    for category, fields in categories.items():
        for field in fields:
            fields[field] = round(fields[field] / total_evaluations, 1) if total_evaluations else 0
    
     # Compute evaluation counts
    positive_evaluations = data.filter(predicted_sentiment='Positive').count()
    negative_evaluations = data.filter(predicted_sentiment='Negative').count()
    neutral_evaluations = data.filter(predicted_sentiment='Neutral').count()
    context = {
        'faculty': faculty,
        'evaluation_status': evaluation_status,
        'avg_rating': avg_rating,
        'recent_comments': recent_comments,
        'categorized_data': categories,  # Pass the categorized data to the template
        'total_evaluations': total_evaluations,
        'event_notifications': event_notifications,
        'all_notifications': all_notifications,
        'total_unread_notifications_count': total_unread_notifications_count,
        'messages_notifications': messages_notifications,
        'sentiment_score': sentiment_score,
        'current_evaluation_status': current_evaluation_status,
        'filterset': filterset,
        'rounded_sentiment_score': rounded_sentiment_score,
        'positive_evaluations': positive_evaluations,
        'negative_evaluations': negative_evaluations,
        'neutral_evaluations': neutral_evaluations,
        'current_academic_year': current_academic_year,
        'current_semester': current_semester,
        'is_head_of_osas': is_head_of_osas,
        'is_department_head': is_department_head,
        'filtered_academic_year': filtered_academic_year,
        'filtered_semester': filtered_semester,
        'is_filtered': is_filtered
    }

    return render(request, 'pages/facultydashboard.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty', 'head of OSAS', 'department head/program coordinator'])
def facultydashboard_peer_to_peer(request):
    is_head_of_osas = request.user.groups.filter(name='head of OSAS').exists() 
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()
    user=request.user
    faculty = Faculty.objects.filter(email=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    # Filter for unread event notifications (level=success) and unread message notifications (level=info)
    unread_event_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_message_notifications = Notification.objects.filter(recipient=user, level='info', unread=True)

    # Count the unread notifications
    unread_event_notifications_count = unread_event_notifications.count()
    unread_message_notifications_count = unread_message_notifications.count()

    # Combine both counts for the total unread notifications
    total_unread_notifications_count = unread_event_notifications_count + unread_message_notifications_count
    # Combine and order the notifications
    all_notifications = Notification.objects.filter(
        Q(level='success') | Q(level='info') | Q(level='warning')
    ).filter(recipient=user).order_by('-timestamp')[:5]


    evaluation_status = EvaluationStatus.objects.first()
    current_academic_year = evaluation_status.academic_year 
    current_semester = evaluation_status.semester
    current_evaluation_status = evaluation_status.evaluation_status

    

    # Initialize the filterset and filter data
    filterset = PeertoPeerEvaluationFilter(request.GET, queryset=PeertoPeerEvaluation.objects.filter(peer=faculty))
    if not request.GET or (request.GET.get('academic_year') == '' and request.GET.get('semester') == ''):
        data = PeertoPeerEvaluation.objects.filter(peer=faculty, 
                                               academic_year=current_academic_year, semester=current_semester)
        is_filtered = False
    else:
        data = filterset.qs
        is_filtered = True

    # Get the filtered academic year and semester
    filtered_academic_year = request.GET.get('academic_year', current_academic_year)
    filtered_semester = request.GET.get('semester', current_semester)

    # Compute evaluation counts
    positive_evaluations = data.filter(predicted_sentiment='Positive').count()
    negative_evaluations = data.filter(predicted_sentiment='Negative').count()
    neutral_evaluations = data.filter(predicted_sentiment='Neutral').count()
    total_evaluations = data.count()

    # Sentiment score calculation
    sentiment_score = (positive_evaluations - negative_evaluations) / total_evaluations if total_evaluations else 0
    rounded_sentiment_score = round(sentiment_score, 1)
    # Average rating calculation
    avg_rating = data.aggregate(average_rating=Avg('average_rating'))['average_rating']
    avg_rating = round(avg_rating, 1) if avg_rating else None

    # Fetch recent comments
    recent_comments = data.values('comments', 'predicted_sentiment')[:3]

   
     # Categorize the evaluations for Chart.js
    categories = {
        'Subject Matter Content': {
            'command_and_knowledge_of_the_subject': 0,
            'depth_of_mastery': 0,
            'practice_in_respective_discipline': 0,
            'up_to_date_knowledge': 0,
            'integrates_subject_to_practical_circumstances': 0,
        },
        'Organization': {
            'organizes_the_subject_matter': 0,
            'provides_orientation_on_course_content': 0,
            'efforts_of_class_preparation': 0,
            'summarizes_main_points': 0,
            'monitors_online_class': 0,
        },
        'Teacher-Student Rapport': {
            'holds_interest_of_students': 0,
            'provides_relevant_feedback': 0,
            'encourages_participation': 0,
            'shows_enthusiasm': 0,
            'shows_sense_of_humor': 0,
        },
        'Teaching Methods': {
            'teaching_methods': 0,
            'flexible_learning_strategies': 0,
            'student_engagement': 0,
            'clear_examples': 0,
            'focused_on_objectives': 0,
        },
        'Presentation': {
            'starts_with_motivating_activities': 0,
            'speaks_in_clear_and_audible_manner': 0,
            'uses_appropriate_medium_of_instruction': 0,
            'establishes_online_classroom_environment': 0,
            'observes_proper_classroom_etiquette': 0,
        },
        'Classroom Management': {
            'uses_time_wisely': 0,
            'gives_ample_time_for_students_to_prepare': 0,
            'updates_the_students_of_their_progress': 0,
            'demonstrates_leadership_and_professionalism': 0,
            'understands_possible_distractions': 0,
        },
        'Sensitivity and Support to Students': {
            'sensitivity_to_student_culture': 0,
            'responds_appropriately': 0,
            'assists_students_on_concerns': 0,
            'guides_the_students_in_accomplishing_tasks': 0,
            'extends_consideration_to_students': 0,
        }
    }

    count = {key: 0 for key in categories}  # Track rating counts per category

    # Populate category data from evaluations
    for evaluation in data:
        # Subject Matter Content
        categories['Subject Matter Content']['command_and_knowledge_of_the_subject'] += evaluation.command_and_knowledge_of_the_subject
        categories['Subject Matter Content']['depth_of_mastery'] += evaluation.depth_of_mastery
        categories['Subject Matter Content']['practice_in_respective_discipline'] += evaluation.practice_in_respective_discipline
        categories['Subject Matter Content']['up_to_date_knowledge'] += evaluation.up_to_date_knowledge
        categories['Subject Matter Content']['integrates_subject_to_practical_circumstances'] += evaluation.integrates_subject_to_practical_circumstances
        
        count['Subject Matter Content'] += 5
        
        # Organization
        categories['Organization']['organizes_the_subject_matter'] += evaluation.organizes_the_subject_matter
        categories['Organization']['provides_orientation_on_course_content'] += evaluation.provides_orientation_on_course_content
        categories['Organization']['efforts_of_class_preparation'] += evaluation.efforts_of_class_preparation
        categories['Organization']['summarizes_main_points'] += evaluation.summarizes_main_points
        categories['Organization']['monitors_online_class'] += evaluation.monitors_online_class
        
        count['Organization'] += 5

       
        # Teacher-Student Rapport
        categories['Teacher-Student Rapport']['holds_interest_of_students'] += evaluation.holds_interest_of_students
        categories['Teacher-Student Rapport']['provides_relevant_feedback'] += evaluation.provides_relevant_feedback
        categories['Teacher-Student Rapport']['encourages_participation'] += evaluation.encourages_participation
        categories['Teacher-Student Rapport']['shows_enthusiasm'] += evaluation.shows_enthusiasm
        categories['Teacher-Student Rapport']['shows_sense_of_humor'] += evaluation.shows_sense_of_humor
        
        count['Teacher-Student Rapport'] += 5

       
        # Teaching Methods
        categories['Teaching Methods']['teaching_methods'] += evaluation.teaching_methods
        categories['Teaching Methods']['flexible_learning_strategies'] += evaluation.flexible_learning_strategies
        categories['Teaching Methods']['student_engagement'] += evaluation.student_engagement
        categories['Teaching Methods']['clear_examples'] += evaluation.clear_examples
        categories['Teaching Methods']['focused_on_objectives'] += evaluation.focused_on_objectives
        
        count['Teaching Methods'] += 5

       
        # Presentation
        categories['Presentation']['starts_with_motivating_activities'] += evaluation.starts_with_motivating_activities
        categories['Presentation']['speaks_in_clear_and_audible_manner'] += evaluation.speaks_in_clear_and_audible_manner
        categories['Presentation']['uses_appropriate_medium_of_instruction'] += evaluation.uses_appropriate_medium_of_instruction
        categories['Presentation']['establishes_online_classroom_environment'] += evaluation.establishes_online_classroom_environment
        categories['Presentation']['observes_proper_classroom_etiquette'] += evaluation.observes_proper_classroom_etiquette
        
        count['Presentation'] += 5

       
        # Classroom Management
        categories['Classroom Management']['uses_time_wisely'] += evaluation.uses_time_wisely
        categories['Classroom Management']['gives_ample_time_for_students_to_prepare'] += evaluation.gives_ample_time_for_students_to_prepare
        categories['Classroom Management']['updates_the_students_of_their_progress'] += evaluation.updates_the_students_of_their_progress
        categories['Classroom Management']['demonstrates_leadership_and_professionalism'] += evaluation.demonstrates_leadership_and_professionalism
        categories['Classroom Management']['understands_possible_distractions'] += evaluation.understands_possible_distractions
        
        count['Classroom Management'] += 5

       
        # Sensitivity and Support to Students
        categories['Sensitivity and Support to Students']['sensitivity_to_student_culture'] += evaluation.sensitivity_to_student_culture
        categories['Sensitivity and Support to Students']['responds_appropriately'] += evaluation.responds_appropriately
        categories['Sensitivity and Support to Students']['assists_students_on_concerns'] += evaluation.assists_students_on_concerns
        categories['Sensitivity and Support to Students']['guides_the_students_in_accomplishing_tasks'] += evaluation.guides_the_students_in_accomplishing_tasks
        categories['Sensitivity and Support to Students']['extends_consideration_to_students'] += evaluation.extends_consideration_to_students
        
        count['Sensitivity and Support to Students'] += 5

       

    # Average each field
    for category, fields in categories.items():
        for field in fields:
            fields[field] = round(fields[field] / total_evaluations, 1) if total_evaluations else 0
    
     # Compute evaluation counts
    positive_evaluations = data.filter(predicted_sentiment='Positive').count()
    negative_evaluations = data.filter(predicted_sentiment='Negative').count()
    neutral_evaluations = data.filter(predicted_sentiment='Neutral').count()
    context = {
        'faculty': faculty,
        'evaluation_status': evaluation_status,
        'avg_rating': avg_rating,
        'recent_comments': recent_comments,
        'categorized_data': categories,  # Pass the categorized data to the template
        'total_evaluations': total_evaluations,
        'event_notifications': event_notifications,
        'all_notifications': all_notifications,
        'total_unread_notifications_count': total_unread_notifications_count,
        'messages_notifications': messages_notifications,
        'sentiment_score': sentiment_score,
        'current_evaluation_status': current_evaluation_status,
        'filterset': filterset,
        'rounded_sentiment_score': rounded_sentiment_score,
        'positive_evaluations': positive_evaluations,
        'negative_evaluations': negative_evaluations,
        'neutral_evaluations': neutral_evaluations,
        'current_academic_year': current_academic_year,
        'current_semester': current_semester,
        'is_head_of_osas': is_head_of_osas,
        'is_department_head': is_department_head,
        'filtered_academic_year': filtered_academic_year,
        'filtered_semester': filtered_semester,
        'is_filtered': is_filtered
    }

    return render(request, 'pages/facultydashboard_peer_to_peer.html', context)

def get_evaluation_data(request):
    faculty = Faculty.objects.filter(email=request.user.username).first()
    evaluation_status = EvaluationStatus.objects.first()
    current_academic_year = evaluation_status.academic_year 
    current_semester = evaluation_status.semester
    
    teacher_evaluations = LikertEvaluation.objects.filter(section_subject_faculty__faculty=faculty,academic_year=current_academic_year, semester=current_semester, admin_status='Approved' )
    positive_evaluations = teacher_evaluations.filter(predicted_sentiment='Positive').count()
    negative_evaluations = teacher_evaluations.filter(predicted_sentiment='Negative').count()
    neutral_evaluations = teacher_evaluations.filter(predicted_sentiment='Neutral').count()

    data = {
        'positive_evaluations': positive_evaluations,
        'negative_evaluations': negative_evaluations,
        'neutral_evaluations': neutral_evaluations,
    }
    return JsonResponse(data)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty', 'head of OSAS', 'department head/program coordinator'])
def faculty_evaluations_individual_summary_report_pdf(request):
    """
    Generates a PDF summary report of faculty evaluations for the currently logged-in faculty.

    Args:
        request: The HTTP request object.

    Returns:
        An HTTP response containing the PDF report.
    """
    current_user = request.user

    image_path = os.path.join(settings.BASE_DIR, static('images/cvsulogo.png'))
    evaluation_status = EvaluationStatus.objects.first()  # Assuming there's only one status entry
    current_academic_year = evaluation_status.academic_year
    current_semester = evaluation_status.semester

    # Apply filters from the EvaluationFilter based on the request data
    evaluation_filter = EvaluationFilter(request.GET, queryset=LikertEvaluation.objects.filter(
        academic_year=current_academic_year, semester=current_semester, section_subject_faculty__faculty=current_user.faculty))
    filtered_evaluations = evaluation_filter.qs

    summary_data = []
    comments_data = []

    faculty_name = current_user.faculty.full_name()
    evaluations = filtered_evaluations.filter(section_subject_faculty__faculty=current_user.faculty)
    num_evaluators = evaluations.count()

    if num_evaluators > 0:
        category_sums = {
            'Subject Matter Content': 0,
            'Organization': 0,
            'Teacher-Student Rapport': 0,
            'Teaching Methods': 0,
            'Presentation': 0,
            'Classroom Management': 0,
            'Sensitivity and Support to Students': 0,
            'Overall': 0
        }

        MAX_SCORE = 5
        NUM_CATEGORIES = len(category_sums) - 1  # Exclude 'Overall' key

        for evaluation in evaluations:
            category_averages = evaluation.calculate_category_averages()
            for category, average in category_averages.items():
                if average is not None:
                    category_sums[category] += average
            category_sums['Overall'] += evaluation.average_rating

        # Calculate averages
        category_averages = {category: round(total / num_evaluators, 2) for category, total in category_sums.items()}
        avg_rating = category_averages['Overall']

        # Calculate overall percentage of category averages
        total_score = sum(category_averages[category] for category in category_sums if category != 'Overall')
        max_possible_score = MAX_SCORE * NUM_CATEGORIES
        overall_percentage = round((total_score / max_possible_score) * 100, 2)

        if avg_rating is not None:
            if 1.0 <= avg_rating <= 1.49:
                rating_category = "Poor"
            elif 1.5 <= avg_rating <= 2.49:
                rating_category = "Unsatisfactory"
            elif 2.5 <= avg_rating <= 3.49:
                rating_category = "Satisfactory"
            elif 3.5 <= avg_rating <= 4.49:
                rating_category = "Very Satisfactory"
            elif 4.5 <= avg_rating <= 5.0:
                rating_category = "Outstanding"
            else:
                rating_category = "No Rating"
        else:
            rating_category = "No Rating"

        summary_data.append({
            'faculty': faculty_name,
            'num_evaluators': num_evaluators,
            'subject_matter_content_avg': category_averages['Subject Matter Content'],
            'organization_avg': category_averages['Organization'],
            'teacher_student_rapport_avg': category_averages['Teacher-Student Rapport'],
            'teaching_methods_avg': category_averages['Teaching Methods'],
            'presentation_avg': category_averages['Presentation'],
            'classroom_management_avg': category_averages['Classroom Management'],
            'sensitivity_support_students_avg': category_averages['Sensitivity and Support to Students'],
            'overall_avg': category_averages['Overall'],
            'overall_percentage': overall_percentage,
            'rating_category': rating_category
        })

        for evaluation in evaluations:
            comments_data.append({
                'faculty': faculty_name,
                'requires_less_task_for_credit': evaluation.requires_less_task_for_credit,
                'strengths_of_the_faculty': evaluation.strengths_of_the_faculty,
                'other_suggestions_for_improvement': evaluation.other_suggestions_for_improvement,
                'comments': evaluation.comments
            })

    # Render the summary data to an HTML template
    html = render_to_string('pages/faculty_evaluations_individual_summary_report.html', {'summary_data': summary_data, 'image_path': image_path, 'comments_data': comments_data, 'current_academic_year': current_academic_year, 'current_semester': current_semester, 'faculty_name': faculty_name})

    # Create the PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="SET-Individual-Summary-Report.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response, link_callback=link_callback)

    if pisa_status.err:
        return HttpResponse('We had some errors with code %s' % pisa_status.err)
    return response

@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty', 'head of OSAS', 'department head/program coordinator'])
def peer_to_peer_individual_summary_report_pdf(request):
    """
    Generates a PDF summary report of peer-to-peer evaluations for the currently logged in faculty.
    """
    # Build the path to the logo image.
    image_path = os.path.join(settings.BASE_DIR, static('images/cvsulogo.png'))
    
    # Get current academic year and semester (or fallback to GET parameters).
    evaluation_status = EvaluationStatus.objects.first()  # Assuming a single status entry exists.
    current_academic_year = evaluation_status.academic_year 
    current_semester = evaluation_status.semester
    academic_year = request.GET.get('academic_year', current_academic_year)
    semester = request.GET.get('semester', current_semester)

    # Get the current faculty (logged in user) and its department.
    current_faculty = Faculty.objects.filter(email=request.user.username).first()   

    # Filter peer-to-peer evaluations for the current academic year, semester, and department.
    # Note: We filter so that the evaluations are for which the current faculty is the "peer".
    evaluation_filter = EvaluationFilter(
        request.GET,
        queryset=PeertoPeerEvaluation.objects.filter(
            academic_year=academic_year,
            semester=semester,
            peer=current_faculty  # current faculty as the evaluated peer
        )
    )
    filtered_evaluations = evaluation_filter.qs

    # Use the current faculty's full name.
    faculty_name = current_faculty.full_name()  # Assumes this returns a string.
    
    # Get the evaluations for the current faculty.
    evaluations = filtered_evaluations
    num_evaluators = evaluations.count()

    summary_data = []
    comments_data = []

    if num_evaluators > 0:
        category_sums = {
            'Subject Matter Content': 0,
            'Organization': 0,
            'Teacher-Student Rapport': 0,
            'Teaching Methods': 0,
            'Presentation': 0,
            'Classroom Management': 0,
            'Sensitivity and Support to Students': 0,
            'Overall': 0
        }

        MAX_SCORE = 5
        NUM_CATEGORIES = len(category_sums) - 1  # Exclude 'Overall' key

        for evaluation in evaluations:
            category_averages = evaluation.calculate_category_averages()
            for category, average in category_averages.items():
                if average is not None:
                    category_sums[category] += average
            category_sums['Overall'] += evaluation.average_rating

        # Calculate averages
        category_averages = {category: round(total / num_evaluators, 2) for category, total in category_sums.items()}
        avg_rating = category_averages['Overall']

        # Calculate overall percentage of category averages
        total_score = sum(category_averages[category] for category in category_sums if category != 'Overall')
        max_possible_score = MAX_SCORE * NUM_CATEGORIES
        overall_percentage = round((total_score / max_possible_score) * 100, 2)

        if avg_rating is not None:
            if 1.0 <= avg_rating <= 1.49:
                rating_category = "Poor"
            elif 1.5 <= avg_rating <= 2.49:
                rating_category = "Unsatisfactory"
            elif 2.5 <= avg_rating <= 3.49:
                rating_category = "Satisfactory"
            elif 3.5 <= avg_rating <= 4.49:
                rating_category = "Very Satisfactory"
            elif 4.5 <= avg_rating <= 5.0:
                rating_category = "Outstanding"
            else:
                rating_category = "No Rating"
        else:
            rating_category = "No Rating"

        summary_data.append({
            'faculty': faculty_name,
            'num_evaluators': num_evaluators,
            'subject_matter_content_avg': category_averages['Subject Matter Content'],
            'organization_avg': category_averages['Organization'],
            'teacher_student_rapport_avg': category_averages['Teacher-Student Rapport'],
            'teaching_methods_avg': category_averages['Teaching Methods'],
            'presentation_avg': category_averages['Presentation'],
            'classroom_management_avg': category_averages['Classroom Management'],
            'sensitivity_support_students_avg': category_averages['Sensitivity and Support to Students'],
            'overall_avg': category_averages['Overall'],
            'overall_percentage': overall_percentage,
            'rating_category': rating_category
        })
        # Collect comments data for each evaluation.
        for evaluation in evaluations:
            comments_data.append({
                'faculty': faculty_name,
                'requires_less_task_for_credit': evaluation.requires_less_task_for_credit,
                'strengths_of_the_faculty': evaluation.strengths_of_the_faculty,
                'other_suggestions_for_improvement': evaluation.other_suggestions_for_improvement,
                'comments': evaluation.comments
            })

    # Render the summary data to an HTML template.
    html = render_to_string('pages/peer_to_peer_individual_summary_report_pdf.html', {
        'summary_data': summary_data,
        'image_path': image_path,
        'comments_data': comments_data,
        'academic_year': academic_year,
        'semester': semester,
        'faculty_name': faculty_name
    })

    # Create the PDF response using xhtml2pdf.
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="peer_to_peer_individual_summary_report_pdf.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response, link_callback=link_callback)

    if pisa_status.err:
        return HttpResponse('We had some errors with code %s' % pisa_status.err)
    return response

@login_required(login_url='signin')
@allowed_users(allowed_roles=['student', 'faculty', 'head of OSAS', 'department head/program coordinator'])
def mark_notifications_read(request):
    if request.method == 'POST':
        # Mark all notifications as read for the current user
        Notification.objects.filter(recipient=request.user, unread=True).update(unread=False)
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'}, status=400)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty', 'head of OSAS', 'department head/program coordinator'])
def faculty_notifications(request):
    user=request.user
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()
    faculty = Faculty.objects.filter(email=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()
    notifications = Notification.objects.filter(recipient=user, level='success')
  
    return render(request, 'pages/faculty_notifications.html', {'faculty': faculty, 'notifications': notifications, 'event_notifications': event_notifications,
        'notifications_unread_count': notifications_unread_count,
        'messages_notifications': messages_notifications,
        'messages_unread_count': messages_unread_count, 'is_department_head': is_department_head })  # Return the response

@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty', 'head of OSAS', 'department head/program coordinator'])
def inbox(request):
    user=request.user
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()
    faculty = Faculty.objects.filter(email=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()
    inbox = Notification.objects.filter(recipient=user, level='info')
  
    return render(request, 'pages/inbox.html', {'faculty': faculty, 'inbox': inbox, 'event_notifications': event_notifications,
        'notifications_unread_count': notifications_unread_count,
        'messages_notifications': messages_notifications,
        'messages_unread_count': messages_unread_count,'is_department_head': is_department_head })  # Return the response

@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty', 'head of OSAS', 'department head/program coordinator'])
def view_message(request, notification_id):
    user=request.user
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()
    faculty = Faculty.objects.filter(email=request.user.username).first()   

    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()


     # Get the specific message by ID, ensuring it's intended for the current user
    message = get_object_or_404(Notification, id=notification_id, recipient=request.user)
    message.mark_as_read()

    return render(request, 'pages/view_message.html', {'faculty': faculty, 'message': message,'event_notifications': event_notifications,
        'notifications_unread_count': notifications_unread_count,
        'messages_notifications': messages_notifications,
        'messages_unread_count': messages_unread_count, 'is_department_head': is_department_head })  # Return the response


@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty', 'head of OSAS', 'department head/program coordinator'])
def facultyprofile(request):
    user=request.user
    faculty = Faculty.objects.filter(email=request.user.username).first()   
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()

    context = {'faculty': faculty, 'event_notifications': event_notifications,
        'notifications_unread_count': notifications_unread_count,
        'messages_notifications': messages_notifications,
        'messages_unread_count': messages_unread_count,'is_department_head': is_department_head}
    return render(request, 'pages/facultyprofile.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty', 'head of OSAS', 'department head/program coordinator'])
def edit_faculty_profile(request):
    user=request.user.faculty
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()
    user_faculty = request.user
    faculty = Faculty.objects.filter(email=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user_faculty, level='success')
    unread_notifications = Notification.objects.filter(recipient=user_faculty, level='success', unread=True)
    notifications_unread_count = unread_notifications.count()
    form = FacultyProfileForm(instance = user)
    if request.method == 'POST':
        form = FacultyProfileForm(request.POST, request.FILES, instance = user)
        if form.is_valid():
            form.save(commit=True)  # Ensure commit is set to True to save to the database
            messages.success(request, 'Profile Updated Successfully')
            return redirect('facultydashboard')

    return render(request, 'pages/edit_faculty_profile.html', {'faculty': faculty, 'form': form, 'event_notifications': event_notifications, 'notifications_unread_count': notifications_unread_count, 'is_department_head': is_department_head})

  
@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty', 'head of OSAS', 'department head/program coordinator'])   
def facultyfeedbackandevaluations(request):
    user=request.user
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()
    faculty = Faculty.objects.filter(email=request.user.username).first()   
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()
    evaluation_status = EvaluationStatus.objects.first()
    current_academic_year = evaluation_status.academic_year 
    current_semester = evaluation_status.semester
    current_status = evaluation_status.evaluation_status
    
    try:
        # Query the Faculty model using the email address
        teacher = Faculty.objects.get(user=user)
    except Faculty.DoesNotExist:
        # Handle the case where the faculty with the given email does not exist
        raise Http404("Faculty does not exist for the logged-in user")
    
     # Create mutable copy of GET parameters
    filter_params = request.GET.copy()

    # Set defaults only when values are missing or empty
    if not filter_params.get('academic_year'):
        filter_params['academic_year'] = current_academic_year
    if not filter_params.get('semester'):
        filter_params['semester'] = current_semester

     #filter and search
    faculty_evaluation_filter = EvaluationFilter(
        filter_params,
        queryset=LikertEvaluation.objects.filter(section_subject_faculty__faculty=faculty, admin_status='Approved'), 
        faculty=faculty, 
        )

    if not request.GET or (request.GET.get('academic_year') == '' and request.GET.get('semester') == ''):
        teacher_evaluations = LikertEvaluation.objects.filter(section_subject_faculty__faculty=faculty, admin_status='Approved', academic_year=current_academic_year, semester=current_semester)
    else:    
        teacher_evaluations = faculty_evaluation_filter.qs
    

    # ordering functionality
   
    ordering = request.GET.get('ordering', "")

     
    if ordering:
        teacher_evaluations = teacher_evaluations.order_by(ordering) 

    #pagination
    page_number = request.GET.get('page', 1)
    evaluation_paginator = Paginator(teacher_evaluations, 10)
   

    try:
        page = evaluation_paginator.page(page_number)
    except EmptyPage:
        page = evaluation_paginator.page(evaluation_paginator.num_pages)
 
    context = {'faculty': faculty, 'teacher': teacher, 'teacher_evaluations': page.object_list, 'faculty_evaluation_filter': faculty_evaluation_filter, 'page_obj':page, 'is_paginated': True, 'paginator':evaluation_paginator,'ordering': ordering, 'event_notifications': event_notifications,
        'notifications_unread_count': notifications_unread_count,
        'messages_notifications': messages_notifications,
        'messages_unread_count': messages_unread_count, 'is_department_head': is_department_head, 'current_status': current_status, 'current_academic_year': current_academic_year, 'current_semester': current_semester}

    return render(request, 'pages/facultyfeedbackandevaluations.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty', 'head of OSAS', 'department head/program coordinator'])
def faculty_evaluations_excel(request):
    faculty = Faculty.objects.filter(email=request.user.username).first()
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=faculty_evaluations.xlsx'

    # Create a workbook and a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Faculty Evaluations"
    
    evaluation_status = EvaluationStatus.objects.first()
    current_academic_year = evaluation_status.academic_year 
    current_semester = evaluation_status.semester
    
    academic_year = request.GET.get('academic_year', current_academic_year)
    semester = request.GET.get('semester', current_semester)
    # Apply filters from the EvaluationFilter based on the request data
    evaluation_filter = EvaluationFilter(request.GET, queryset=LikertEvaluation.objects.filter(section_subject_faculty__faculty=faculty, academic_year=academic_year, semester=semester))

    # Get the filtered queryset
    filtered_evaluations = evaluation_filter.qs

    # Add column headings to the worksheet
    columns = ['Subject', 'Faculty', 'Average', 'Rating', 'Overall Impression', 'Polarity', 'Academic Year', 'Semester']
    ws.append(columns)

    # Loop through and output the data
    for evaluation in filtered_evaluations:
        row = [
            str(evaluation.section_subject_faculty.subjects),
            str(evaluation.section_subject_faculty.faculty),
            str(evaluation.average_rating),
            str(evaluation.get_rating_category()),
            str(evaluation.comments),
            str(evaluation.predicted_sentiment),
            str(evaluation.academic_year),
            str(evaluation.semester)
        ]
        ws.append(row)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook to a BytesIO stream
    output = BytesIO()
    wb.save(output)
    response.write(output.getvalue())

    return response

@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty', 'head of OSAS', 'department head/program coordinator'])
def view_evaluation_form(request, pk):
    user=request.user
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()
    faculty = Faculty.objects.filter(email=request.user.username).first()   

    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()
    
    questions = FacultyEvaluationQuestions.objects.all().order_by('order')
    faculty_evaluation_form = LikertEvaluation.objects.get(pk=pk)

        # List of all rating fields to check
    rating_fields = [
        'command_and_knowledge_of_the_subject',
        'depth_of_mastery',
        'practice_in_respective_discipline',
        'up_to_date_knowledge',
        'integrates_subject_to_practical_circumstances',
        'organizes_the_subject_matter',
        'provides_orientation_on_course_content',
        'efforts_of_class_preparation',
        'summarizes_main_points',
        'monitors_online_class',
        'holds_interest_of_students',
        'provides_relevant_feedback',
        'encourages_participation',
        'shows_enthusiasm',
        'shows_sense_of_humor',
        'teaching_methods',
        'flexible_learning_strategies',
        'student_engagement',
        'clear_examples',
        'focused_on_objectives',
        'starts_with_motivating_activities',
        'speaks_in_clear_and_audible_manner',
        'uses_appropriate_medium_of_instruction',
        'establishes_online_classroom_environment',
        'observes_proper_classroom_etiquette',
        'uses_time_wisely',
        'gives_ample_time_for_students_to_prepare',
        'updates_the_students_of_their_progress',
        'demonstrates_leadership_and_professionalism',
        'understands_possible_distractions',
        'sensitivity_to_student_culture',
        'responds_appropriately',
        'assists_students_on_concerns',
        'guides_the_students_in_accomplishing_tasks',
        'extends_consideration_to_students'
    ]
    # Initialize counters
    outstanding_count = 0
    very_satisfactory_count = 0
    satisfactory_count = 0
    unsatisfactory_count = 0
    poor_count = 0

    # Calculate counts in memory
    for field in rating_fields:
        value = getattr(faculty_evaluation_form, field)
        if value == 5:
            outstanding_count += 1
        elif value == 4:
            very_satisfactory_count += 1
        elif value == 3:
            satisfactory_count += 1
        elif value == 2:
            unsatisfactory_count += 1
        elif value == 1:
            poor_count += 1

    return render(request, 'pages/view_evaluation_form.html', {'faculty_evaluation_form': faculty_evaluation_form, 'faculty': faculty, 'outstanding_count': outstanding_count, 'very_satisfactory_count': very_satisfactory_count, 'satisfactory_count': satisfactory_count, 'unsatisfactory_count': unsatisfactory_count, 'poor_count': poor_count, 'questions': questions, 'event_notifications': event_notifications,
        'notifications_unread_count': notifications_unread_count,
        'messages_notifications': messages_notifications,
        'messages_unread_count': messages_unread_count,'is_department_head': is_department_head})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty', 'head of OSAS', 'department head/program coordinator'])
def faculty_event_evaluations(request):
     user=request.user
     is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()
     faculty = Faculty.objects.filter(email=request.user.username).first()   
     is_head_of_osas = request.user.groups.filter(name='head of OSAS').exists() 
     event_notifications = Notification.objects.filter(recipient=user, level='success')
     messages_notifications = Notification.objects.filter(recipient=user, level='info')
     unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
     unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

     notifications_unread_count = unread_notifications.count()
     messages_unread_count = unread_messages.count()    

     pending_events = Event.objects.filter(admin_status='Pending')
     pending_count = pending_events.count()   
     if is_head_of_osas == True:
          event = Event.objects.filter(admin_status='Approved').select_related('author').order_by('-updated')

     else:
        event = Event.objects.filter(author=user, admin_status='Approved').order_by('-updated')

        # ordering functionality
     event_filter = EventFilter(request.GET, queryset=event)
     event = event_filter.qs 
     ordering = request.GET.get('ordering', "")
     event = event.annotate(
     avg_rating=Case(
        When(
            event_type__name__in=['School Event', 'Training Workshop'],
            then=Coalesce(Avg('schooleventmodel__average_rating'), Value(0))
        ),
        When(
            event_type__name='Webinar/Seminar',
            then=Coalesce(Avg('webinarseminarmodel__average_rating'), Value(0))
        ),
        default=Value(0),
        output_field=FloatField()
    ),
    evaluator_count=Case(
         When(
             event_type__name__in=['School Event', 'Training Workshop'],
             then=Count('schooleventmodel__id')
         ),
         When(
             event_type__name='Webinar/Seminar',
             then=Count('webinarseminarmodel__id')
         ),
         default=Value(0),
         output_field=IntegerField()
    )
)
        
     if ordering:
        event = event.order_by(ordering) 

     form = EventCreationForm()
     if request.method == 'POST':
        form = EventCreationForm(request.POST, request.FILES)


        today = datetime.today().date()


        if form.is_valid():
            event = form.save(commit=False)
            event.author = request.user  # Set the author to the currently logged-in user
            if request.user.groups.filter(name='head of OSAS').exists(): 
                event.admin_status = 'Approved'
                form.save()
                messages.success(request, "The event has been successfully posted.")
                saved_event = Event.objects.get(pk=event.pk)

                    #Notifications      
                student_users = User.objects.filter(student__Course__in=saved_event.course_attendees.all())

                faculty_users = User.objects.filter(faculty__department__in=saved_event.department_attendees.all())

                notification_description = f"A new event '{event.title}' is scheduled for {event.date}. Please evaluate and provide your feedback!"

                # Send notifications to student attendees
                notify.send(
                    sender=user,
                    recipient=student_users,
                    verb="New Event Created",
                    description=notification_description,
                    level='success'
                )

                # Send notifications to faculty attendees
                notify.send(
                    sender=user,
                    recipient=faculty_users,
                    verb="New Event Created",
                    description=notification_description,
                    level='success'
                )
               
            form.save()
            messages.success(request, "The event has been submitted for approval to the head of OSAS.")

            head_of_osas_group = Group.objects.get(name='head of OSAS')
            head_of_osas_users = User.objects.filter(groups=head_of_osas_group)

            notification_description = f"A new event '{event.title}' has been submitted for approval."

            # Send notification to the head of OSAS
            notify.send(
                sender=request.user,
                recipient=head_of_osas_users,
                verb="Event Approval Needed",
                description=notification_description,
                level='success'
            )
            evaluation_start_date_str = request.POST.get("evaluation_start_datetime")  
            evaluation_end_date_str = request.POST.get("evaluation_end_datetime") 

            event_pk = event.pk
            print(event_pk)
            if evaluation_start_date_str:

                start_date = datetime.strptime(evaluation_start_date_str, "%Y-%m-%d").date()

                # Check if the end date exceeds today's date
                if start_date > today:
                    # Schedule the task if the date is valid
                    trigger = DateTrigger(run_date=start_date) 
                    scheduler.add_job(start_event_evaluations, trigger=trigger, args=[event_pk])
                elif start_date == today:
                    # Schedule the task if the date is valid
                    run_date = datetime.now() + timedelta(minutes=1)
                    trigger = DateTrigger(run_date=run_date) 
                    scheduler.add_job(start_event_evaluations, trigger=trigger, args=[event_pk])
                else:
                    # Display an error if the date exceeds today
                    messages.error(request, "The start date cannot be in the past.")
                    return redirect('faculty_event_evaluations')

            if evaluation_end_date_str:
                end_date = datetime.strptime(evaluation_end_date_str, "%Y-%m-%d").date()

                # Check if the release date exceeds today's date
                if end_date > today:
                    # Schedule the task if the date is valid
                    trigger = DateTrigger(run_date=end_date)
                    scheduler.add_job(end_event_evaluations, trigger=trigger, args=[event_pk])
                else:
                    # Display an error if the date exceeds today
                    messages.error(request, "The end date cannot be in the past.")
                    return redirect('faculty_event_evaluations')
       
            return redirect('faculty_event_evaluations')
        else:
            # Add form errors to messages
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Error in {field}: {error}")

     for i in event:
        # Get combined courses and departments names
        attendees_text = ", ".join(course.name for course in i.course_attendees.all()) + " | " + ", ".join(dept.name for dept in i.department_attendees.all())
        # Truncate combined text to 50 characters (or your chosen length)
        i.attendees_summary = Truncator(attendees_text).chars(50)
        

     paginator = Paginator(event, 10) 
     page_number = request.GET.get('page')
     page_obj = paginator.get_page(page_number)


     context = {'event': event, 'faculty': faculty, 'form':form, 'event_notifications': event_notifications,
        'notifications_unread_count': notifications_unread_count,
        'messages_notifications': messages_notifications,
        'messages_unread_count': messages_unread_count, 'is_head_of_osas': is_head_of_osas, 'page_obj': page_obj, 'event_filter': event_filter, 'is_department_head': is_department_head, 'pending_count': pending_count}
     return render(request, 'pages/faculty_event_evaluations.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['head of OSAS', 'department head/program coordinator'])
def pending_events(request):
    user = request.user
    faculty = Faculty.objects.filter(email=request.user.username).first()   
    events = Event.objects.filter(admin_status='Pending').order_by('-updated')
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()
    is_head_of_osas = request.user.groups.filter(name='head of OSAS').exists() 
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()    
    paginator = Paginator(events, 10) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request,'pages/pending_events.html',{'events': events, 'page_obj': page_obj,'is_department_head': is_department_head,'faculty': faculty, 'event_notifications': event_notifications,
        'notifications_unread_count': notifications_unread_count,
        'messages_notifications': messages_notifications,
        'messages_unread_count': messages_unread_count, 'is_head_of_osas': is_head_of_osas, })

@login_required(login_url='signin')
@allowed_users(allowed_roles=['head of OSAS', 'department head/program coordinator'])
def approve_event(request, event_id):
    user=request.user
    event = get_object_or_404(Event, id=event_id, admin_status='Pending')
    
    event.admin_status = 'Approved'
    event.save()

    #Notifications      
    student_users = User.objects.filter(student__Course__in=event.course_attendees.all())

    faculty_users = User.objects.filter(faculty__department__in=event.department_attendees.all())

    notification_description = f"A new event '{event.title}' is scheduled for {event.date}. Please evaluate and provide your feedback!"

    # Send notifications to student attendees
    notify.send(
        sender=user,
        recipient=student_users,
        verb="New Event Created",
        description=notification_description,
        level='success'
    )

    # Send notifications to faculty attendees
    notify.send(
        sender=user,
        recipient=faculty_users,
        verb="New Event Created",
        description=notification_description,
        level='success'
    )
    messages.success(request, 'Event approved successfully')
    return redirect('pending_events')

@login_required(login_url='signin')
@allowed_users(allowed_roles=['head of OSAS', 'department head/program coordinator'])
def reject_event(request, event_id):
    event = get_object_or_404(Event, id=event_id, admin_status='Pending')
    event.admin_status = 'Rejected'
    event.save()
    messages.success(request, 'Event rejected successfully')
    return redirect('pending_events')

@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty', 'head of OSAS', 'department head/program coordinator'])
def view_faculty_event_evaluations(request, pk):
    user = request.user
    faculty = Faculty.objects.filter(email=request.user.username).first()
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()

    # Notifications
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()

    # Fetch the event
    event = get_object_or_404(Event, pk=pk)

    # Determine which model and filter to use based on the event type
    if event.event_type.name in ['School Event', 'Training Workshop']:
        evaluations_queryset = SchoolEventModel.objects.filter(event=event)
        filter_class = SchoolEventFilter
        total_average_rating = SchoolEventModel.get_event_average_rating(event)
    elif event.event_type.name == 'Webinar/Seminar':
        evaluations_queryset = WebinarSeminarModel.objects.filter(event=event)
        filter_class = WebinarSeminarEventFilter
        total_average_rating = WebinarSeminarModel.get_event_average_rating(event)
    else:
        evaluations_queryset = None
        filter_class = None
        total_average_rating = 'Invalid event type'

    # Apply filters if applicable
    if filter_class and evaluations_queryset is not None:
        event_filter = filter_class(request.GET, queryset=evaluations_queryset)
        filtered_evaluations = event_filter.qs
    else:
        event_filter = None
        filtered_evaluations = evaluations_queryset

    # Ensure filtered_evaluations is not None before paginating
    if filtered_evaluations is None:
        filtered_evaluations = []

    paginator = Paginator(filtered_evaluations, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'faculty': faculty,
        'event': event,
        'filtered_evaluations': filtered_evaluations,
        'event_filter': event_filter,
        'total_average_rating': total_average_rating,
        'is_department_head': is_department_head,
        'notifications_unread_count': notifications_unread_count,
        'messages_unread_count': messages_unread_count,
        'event_notifications': event_notifications,
        'messages_notifications': messages_notifications,
        'page_obj': page_obj,
    }

    return render(request, 'pages/view_faculty_event_evaluations.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty', 'head of OSAS', 'department head/program coordinator'])
def faculty_view_event_evaluations(request, pk):
    user=request.user
    faculty = Faculty.objects.filter(email=request.user.username).first()   
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()

    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()  
     # Try to get the event from SchoolEventModel or WebinarSeminarModel
    try:
        event_form_details = SchoolEventModel.objects.get(pk=pk)
 
        
        event_type = 'school_event'
    except SchoolEventModel.DoesNotExist:
        try:
            event_form_details = WebinarSeminarModel.objects.get(pk=pk)
            event_type = 'webinar_event'
        except WebinarSeminarModel.DoesNotExist:
            # If no event is found in either model, handle the error
            return HttpResponse("Event not found.")

    # Render different templates based on the event type
    if event_type == 'school_event':
        user=request.user
        faculty = Faculty.objects.filter(email=request.user.username).first()   

        event_notifications = Notification.objects.filter(recipient=user, level='success')
        messages_notifications = Notification.objects.filter(recipient=user, level='info')
        unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
        unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

        notifications_unread_count = unread_notifications.count()
        messages_unread_count = unread_messages.count()  
        questions = SchoolEventQuestions.objects.all().order_by('order')

        excellent_count = SchoolEventModel.objects.filter(pk=pk).filter(
            meeting_expectation=5
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            attainment_of_the_objectives=5
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(pk=pk).filter(
            topics_discussed=5
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            input_presentation=5
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            management_team=5
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            venue_and_physical_arrangement=5
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            overall_assessment=5
        ).count()

        very_satisfactory_count = SchoolEventModel.objects.filter(pk=pk).filter(
            meeting_expectation=4
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            attainment_of_the_objectives=4
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(pk=pk).filter(
            topics_discussed=4
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            input_presentation=4
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            management_team=4
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            venue_and_physical_arrangement=4
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            overall_assessment=4
        ).count()

        satisfactory_count = SchoolEventModel.objects.filter(pk=pk).filter(
            meeting_expectation=3
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            attainment_of_the_objectives=3
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(pk=pk).filter(
            topics_discussed=3
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            input_presentation=3
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            management_team=3
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            venue_and_physical_arrangement=3
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            overall_assessment=3
        ).count()

        fair_count = SchoolEventModel.objects.filter(pk=pk).filter(
            meeting_expectation=2
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            attainment_of_the_objectives=2
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(pk=pk).filter(
            topics_discussed=2
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            input_presentation=2
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            management_team=2
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            venue_and_physical_arrangement=2
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            overall_assessment=2
        ).count()

        poor_count = SchoolEventModel.objects.filter(pk=pk).filter(
            meeting_expectation=1
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            attainment_of_the_objectives=1
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(pk=pk).filter(
            topics_discussed=1
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            input_presentation=1
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            management_team=1
        ).count()  + SchoolEventModel.objects.filter(pk=pk).filter(
            venue_and_physical_arrangement=1
        ).count() + SchoolEventModel.objects.filter(pk=pk).filter(
            overall_assessment=1
        ).count()

        template_name = 'pages/view_faculty_schoolevent_evaluations.html'
        
    elif event_type == 'webinar_event':
        user=request.user
        faculty = Faculty.objects.filter(email=request.user.username).first()   

        event_notifications = Notification.objects.filter(recipient=user, level='success')
        messages_notifications = Notification.objects.filter(recipient=user, level='info')
        unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
        unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

        notifications_unread_count = unread_notifications.count()
        messages_unread_count = unread_messages.count()
        questions = WebinarSeminarQuestions.objects.all().order_by('order')

        excellent_count = WebinarSeminarModel.objects.filter(pk=pk).filter(
            relevance_of_the_activity=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(pk=pk).filter(
            quality_of_the_activity=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness=5
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            attainment_of_the_objective=5
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_to_attain_the_objective=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods_used=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_in_the_present_time=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            usefulness_of_the_topic_discusssed_in_the_activity=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            displayed_a_thorough_knowledge_of_the_topic=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            explained_activities=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_create_a_good_learning_environment=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_manage_her_time_well=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            demonstrated_keenness_to_the_participant_needs=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness_or_suitability_of_service=5
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            overall_satisfaction=5
        ).count()

        very_satisfactory_count = WebinarSeminarModel.objects.filter(pk=pk).filter(
            relevance_of_the_activity=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(pk=pk).filter(
            quality_of_the_activity=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness=4
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            attainment_of_the_objective=4
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_to_attain_the_objective=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods_used=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_in_the_present_time=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            usefulness_of_the_topic_discusssed_in_the_activity=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            displayed_a_thorough_knowledge_of_the_topic=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            explained_activities=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_create_a_good_learning_environment=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_manage_her_time_well=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            demonstrated_keenness_to_the_participant_needs=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness_or_suitability_of_service=4
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            overall_satisfaction=4
        ).count()

        satisfactory_count = WebinarSeminarModel.objects.filter(pk=pk).filter(
            relevance_of_the_activity=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(pk=pk).filter(
            quality_of_the_activity=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness=3
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            attainment_of_the_objective=3
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_to_attain_the_objective=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods_used=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_in_the_present_time=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            usefulness_of_the_topic_discusssed_in_the_activity=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            displayed_a_thorough_knowledge_of_the_topic=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            explained_activities=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_create_a_good_learning_environment=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_manage_her_time_well=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            demonstrated_keenness_to_the_participant_needs=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness_or_suitability_of_service=3
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            overall_satisfaction=3
        ).count()

        fair_count = WebinarSeminarModel.objects.filter(pk=pk).filter(
            relevance_of_the_activity=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(pk=pk).filter(
            quality_of_the_activity=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness=2
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            attainment_of_the_objective=2
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_to_attain_the_objective=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods_used=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_in_the_present_time=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            usefulness_of_the_topic_discusssed_in_the_activity=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            displayed_a_thorough_knowledge_of_the_topic=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            explained_activities=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_create_a_good_learning_environment=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_manage_her_time_well=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            demonstrated_keenness_to_the_participant_needs=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness_or_suitability_of_service=2
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            overall_satisfaction=2
        ).count()

        poor_count = WebinarSeminarModel.objects.filter(pk=pk).filter(
            relevance_of_the_activity=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(pk=pk).filter(
            quality_of_the_activity=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness=1
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            attainment_of_the_objective=1
        ).count()  + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_to_attain_the_objective=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods_used=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_topic_in_the_present_time=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            usefulness_of_the_topic_discusssed_in_the_activity=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            appropriateness_of_the_searching_methods=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            displayed_a_thorough_knowledge_of_the_topic=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            explained_activities=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_create_a_good_learning_environment=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            able_to_manage_her_time_well=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            demonstrated_keenness_to_the_participant_needs=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            timeliness_or_suitability_of_service=1
        ).count() + WebinarSeminarModel.objects.filter(pk=pk).filter(
            overall_satisfaction=1
        ).count()

       
        template_name = 'pages/view_faculty_webinar_evaluations.html'
    else:
        return HttpResponse("Invalid event type.")


    return render(request, template_name, {'event_form_details': event_form_details, 'faculty': faculty, 'questions': questions, 'excellent_count': excellent_count, 'very_satisfactory_count': very_satisfactory_count, 'satisfactory_count': satisfactory_count, 'fair_count': fair_count, 'poor_count': poor_count, 'faculty': faculty, 'event_notifications': event_notifications,
        'notifications_unread_count': notifications_unread_count,
        'messages_notifications': messages_notifications,
        'messages_unread_count': messages_unread_count,'is_department_head': is_department_head})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty' , 'head of OSAS', 'department head/program coordinator'])
def edit_faculty_events(request, pk):
    user=request.user
    faculty = Faculty.objects.filter(email=request.user.username).first()   
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()

    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()    
    event = Event.objects.get(pk=pk)
    form = EventCreationForm(instance=event)
    today = datetime.today().date()
    if request.method == 'POST':
        form = EventCreationForm(request.POST, request.FILES, instance=event)
        if form.is_valid():
            form.save(commit=True)
            event_pk = event.pk

            try:
                scheduler.remove_job(f'start_event_{event_pk}')
            except JobLookupError:
                pass

            try:
                scheduler.remove_job(f'end_event_{event_pk}')
            except JobLookupError:
                pass

            evaluation_start_date_str = request.POST.get("evaluation_start_datetime")
            evaluation_end_date_str = request.POST.get("evaluation_end_datetime")

            # Log date strings
            print(f"Start Date: {evaluation_start_date_str}, End Date: {evaluation_end_date_str}")

            if evaluation_start_date_str:
                start_date = datetime.strptime(evaluation_start_date_str, "%Y-%m-%d").date()
                if start_date >= today:
                    run_date = datetime.now() + timedelta(minutes=1) if start_date == today else start_date
                    try:
                        trigger = DateTrigger(run_date=run_date)
                        scheduler.add_job(start_event_evaluations, trigger=trigger, args=[event_pk], id=f'start_event_{event_pk}')
                    except Exception as e:
                        print(f"Scheduler Error (start): {e}")
                else:
                    messages.error(request, "The start date cannot be in the past.")

            if evaluation_end_date_str:
                end_date = datetime.strptime(evaluation_end_date_str, "%Y-%m-%d").date()
                if end_date >= today:
                    run_date = datetime.now() + timedelta(minutes=1) if end_date == today else end_date
                    try:
                        trigger = DateTrigger(run_date=run_date)
                        scheduler.add_job(end_event_evaluations, trigger=trigger, args=[event_pk], id=f'end_event_{event_pk}')
                    except Exception as e:
                        print(f"Scheduler Error (end): {e}")
                else:
                    messages.error(request, "The end date cannot be in the past.")

            # Default redirect after processing
            return redirect('faculty_event_evaluations')
        else:
            print(form.errors)  # Debug form errors   

           
    context = {'event': event, 'faculty': faculty, 'form':form, 'event_notifications': event_notifications,
        'notifications_unread_count': notifications_unread_count,
        'messages_notifications': messages_notifications,
        'messages_unread_count': messages_unread_count,'is_department_head': is_department_head}
    return render(request, 'pages/edit_faculty_events.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty', 'head of OSAS', 'department head/program coordinator'])
def delete_faculty_events(request, pk):
    user=request.user
    faculty = Faculty.objects.filter(email=request.user.username).first()   
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()

    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()    
    event = Event.objects.get(pk=pk)
    if request.method == 'POST':
            event.delete()
            return redirect('faculty_event_evaluations')

    return render(request, 'pages/faculty_delete_form.html', {'obj':event, 'faculty': faculty, 'event_notifications': event_notifications,
        'notifications_unread_count': notifications_unread_count,
        'messages_notifications': messages_notifications,
        'messages_unread_count': messages_unread_count,'is_department_head': is_department_head})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty' , 'head of OSAS', 'department head/program coordinator'])
def faculty_events(request):
    is_head_of_osas = request.user.groups.filter(name='head of OSAS').exists() 
    user=request.user
    faculty = Faculty.objects.filter(email=request.user.username).first()
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()

    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()   
    faculty_department = faculty.department  # Get the department the faculty is a part of
    events = Event.objects.filter(department_attendees=faculty_department).distinct()  # Get events related to those department
    # Get evaluated event IDs by the current user
     # Get evaluated event IDs by the current user for SchoolEventModel
    evaluated_school_event_ids = SchoolEventModel.objects.filter(user=user).values_list('event_id', flat=True)
    
    # Get evaluated event IDs by the current user for WebinarSeminarModel
    evaluated_webinar_event_ids = WebinarSeminarModel.objects.filter(user=user).values_list('event_id', flat=True)

    # Combine evaluated event IDs from both models
    evaluated_event_ids = list(evaluated_school_event_ids) + list(evaluated_webinar_event_ids)
     # Get current time
    current_time = timezone.now()
    upcoming_events = events.filter(date__gt=current_time, evaluation_status='Closed')  # Events in the future
    # Past events with closed evaluation
    past_events = events.filter(date__lt=current_time, evaluation_status='Closed')  # Past events with evaluation closed   
    # Exclude events that have been evaluated
    unevaluated_events = events.exclude(id__in=evaluated_event_ids).exclude(id__in=past_events).exclude(id__in=upcoming_events).order_by('-date') 
    for event in unevaluated_events:
        event.attended = Attendance.objects.filter(user=user, event=event, attended=True).exists()

    return render(request, 'pages/faculty_events.html', {'faculty': faculty, 'unevaluated_events': unevaluated_events, 'past_events': past_events, 'upcoming_events': upcoming_events, 'event_notifications': event_notifications,
        'notifications_unread_count': notifications_unread_count,
        'messages_notifications': messages_notifications,
        'messages_unread_count': messages_unread_count,'is_department_head': is_department_head, 'is_head_of_osas': is_head_of_osas,})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty', 'head of OSAS', 'department head/program coordinator'])
def faculty_events_upcoming(request):
    is_head_of_osas = request.user.groups.filter(name='head of OSAS').exists() 
    user=request.user
    faculty = Faculty.objects.filter(email=request.user.username).first()   
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()

    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()  
    department = faculty.department  # Get the department the faculty is a part of
    events = Event.objects.filter(department_attendees=department, admin_status='Approved').distinct()  # Get events related to those department
    
    current_time = timezone.now()
    upcoming_events = events.filter(date__gt=current_time, evaluation_status='Closed')  # Events in the future

    return render(request, 'pages/faculty_events_upcoming.html', {'faculty': faculty,'upcoming_events': upcoming_events, 'event_notifications': event_notifications,
        'notifications_unread_count': notifications_unread_count,
        'messages_notifications': messages_notifications,
        'messages_unread_count': messages_unread_count,'is_department_head': is_department_head, 'is_head_of_osas': is_head_of_osas,})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty', 'head of OSAS', 'department head/program coordinator'])
def faculty_events_evaluated(request):
    is_head_of_osas = request.user.groups.filter(name='head of OSAS').exists() 
    user=request.user
    faculty = Faculty.objects.filter(email=request.user.username).first()   
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()

    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()  
    department = faculty.department  # Get the department the faculty is a part of
    events = Event.objects.filter(department_attendees=department).distinct()  # Get events related to those department
      # Get evaluated event IDs by the current user
     # Get evaluated event IDs by the current user for SchoolEventModel
    evaluated_school_event_ids = SchoolEventModel.objects.filter(user=user).values_list('event_id', flat=True)
    
    # Get evaluated event IDs by the current user for WebinarSeminarModel
    evaluated_webinar_event_ids = WebinarSeminarModel.objects.filter(user=user).values_list('event_id', flat=True)

    # Combine evaluated event IDs from both models
    evaluated_event_ids = list(evaluated_school_event_ids) + list(evaluated_webinar_event_ids)
    
    # Exclude events that have been evaluated
    evaluated_events = events.filter(id__in=evaluated_event_ids).order_by('-date')[:10]
    return render(request, 'pages/faculty_events_evaluated.html', {'faculty': faculty, 'evaluated_events': evaluated_events, 'event_notifications': event_notifications,
        'notifications_unread_count': notifications_unread_count,
        'messages_notifications': messages_notifications,
        'messages_unread_count': messages_unread_count, 'is_department_head': is_department_head, 'is_head_of_osas': is_head_of_osas,})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty', 'head of OSAS', 'department head/program coordinator'])
def faculty_events_closed(request):
    is_head_of_osas = request.user.groups.filter(name='head of OSAS').exists() 
    user=request.user
    faculty = Faculty.objects.filter(email=request.user.username).first()
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()

    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()  
    department = faculty.department  # Get the department the faculty is a part of
    events = Event.objects.filter(department_attendees=department).distinct()  # Get events related to those department

    # Get current time
    current_time = timezone.now()
    # Past events with closed evaluation
    past_events = events.filter(date__lt=current_time, evaluation_status='Closed')  # Past events with evaluation closed   
    return render(request, 'pages/faculty_events_closed.html', {'faculty': faculty, 'past_events': past_events, 'event_notifications': event_notifications,
        'notifications_unread_count': notifications_unread_count,
        'messages_notifications': messages_notifications,
        'messages_unread_count': messages_unread_count, 'is_department_head': is_department_head, 'is_head_of_osas': is_head_of_osas,})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty', 'head of OSAS', 'department head/program coordinator'])
def faculty_event_detail(request, pk):
    is_head_of_osas = request.user.groups.filter(name='head of OSAS').exists() 
    user=request.user
    faculty = Faculty.objects.filter(email=request.user.username).first()   
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()

    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count() 
    event = Event.objects.get(pk=pk)
    questions = SchoolEventQuestions.objects.all().order_by('order')
    if event.event_type.name == 'School Event' or event.event_type.name == 'Training Workshop':
        form = SchoolEventForm()
        if request.method == 'POST':
            form = SchoolEventForm(request.POST)
            if form.is_valid():
                 # Process the evaluation form
                meeting_expectation = form.cleaned_data['meeting_expectation']
                attainment_of_the_objectives = form.cleaned_data['attainment_of_the_objectives']
                topics_discussed = form.cleaned_data['topics_discussed']
                input_presentation = form.cleaned_data['input_presentation']
                management_team = form.cleaned_data['management_team']
                venue_and_physical_arrangement = form.cleaned_data['venue_and_physical_arrangement']
                overall_assessment = form.cleaned_data['overall_assessment']
                suggestions_and_comments = form.cleaned_data['suggestions_and_comments']
                predicted_sentiment = get_sentiment(suggestions_and_comments)
                sentiment_label = 'None'
                if predicted_sentiment and isinstance(predicted_sentiment, list):
                    try:
                        # Access the first list in the response, then sort by score
                        sentiment_list = predicted_sentiment[0]
                        # Get the prediction with the highest score (sorted descending)
                        top_prediction = max(sentiment_list, key=lambda x: x['score'])
                        
                        # Directly use the label from the API response
                        sentiment_label = top_prediction['label'].capitalize()  # "Negative", "Neutral", or "Positive"
                    
                    except (IndexError, KeyError, TypeError) as e:
                        print(f"Error processing sentiment: {e}")
                        sentiment_label = "None"
                # Save the data to database
                form = SchoolEventModel(
                    user=user,
                    event=event,
                    meeting_expectation=meeting_expectation,
                    attainment_of_the_objectives=attainment_of_the_objectives,
                    topics_discussed=topics_discussed,
                    input_presentation=input_presentation,
                    management_team=management_team,
                    venue_and_physical_arrangement=venue_and_physical_arrangement,
                    overall_assessment=overall_assessment,
                    suggestions_and_comments=suggestions_and_comments,
                    predicted_sentiment=sentiment_label
                )
                form.save()
            messages.success(request, 'Evaluation submitted successfully.')
            return redirect('faculty_events')
        return render(request, 'pages/faculty_school_event_form.html', context = {'event': event, 'form': form, 'faculty': faculty, 'questions': questions, 'event_notifications': event_notifications,
        'notifications_unread_count': notifications_unread_count,
        'messages_notifications': messages_notifications,
        'messages_unread_count': messages_unread_count, 'is_department_head': is_department_head, 'is_head_of_osas': is_head_of_osas,})
    
    elif event.event_type.name == 'Webinar/Seminar':
        questions = WebinarSeminarQuestions.objects.all().order_by('order')
        form = WebinarSeminarForm()  # Assuming WebinarSeminarForm is similar to SchoolEventForm
        if request.method == 'POST':
            form = WebinarSeminarForm(request.POST)
            if form.is_valid():
                # Process Webinar/Seminar evaluation
                relevance_of_the_activity = form.cleaned_data['relevance_of_the_activity']

                quality_of_the_activity = form.cleaned_data['quality_of_the_activity']

                timeliness = form.cleaned_data['timeliness']

                suggestions_and_comments = form.cleaned_data['suggestions_and_comments']

                attainment_of_the_objective = form.cleaned_data['attainment_of_the_objective']

                appropriateness_of_the_topic_to_attain_the_objective = form.cleaned_data['appropriateness_of_the_topic_to_attain_the_objective']

                appropriateness_of_the_searching_methods_used = form.cleaned_data['appropriateness_of_the_searching_methods_used']
                
                topics_to_be_included = form.cleaned_data['topics_to_be_included']

                appropriateness_of_the_topic_in_the_present_time = form.cleaned_data['appropriateness_of_the_topic_in_the_present_time']

                usefulness_of_the_topic_discusssed_in_the_activity = form.cleaned_data['usefulness_of_the_topic_discusssed_in_the_activity']

                appropriateness_of_the_searching_methods = form.cleaned_data['appropriateness_of_the_searching_methods']

                displayed_a_thorough_knowledge_of_the_topic = form.cleaned_data['displayed_a_thorough_knowledge_of_the_topic']

                explained_activities = form.cleaned_data['explained_activities']

                able_to_create_a_good_learning_environment = form.cleaned_data['able_to_create_a_good_learning_environment']

                able_to_manage_her_time_well = form.cleaned_data['able_to_manage_her_time_well']

                demonstrated_keenness_to_the_participant_needs = form.cleaned_data['demonstrated_keenness_to_the_participant_needs']

                timeliness_or_suitability_of_service = form.cleaned_data['timeliness_or_suitability_of_service']

                overall_satisfaction = form.cleaned_data['overall_satisfaction']

                predicted_sentiment = get_sentiment(suggestions_and_comments)
                sentiment_label = 'None'
                if predicted_sentiment and isinstance(predicted_sentiment, list):
                    try:
                        # Access the first list in the response, then sort by score
                        sentiment_list = predicted_sentiment[0]
                        # Get the prediction with the highest score (sorted descending)
                        top_prediction = max(sentiment_list, key=lambda x: x['score'])
                        
                        # Directly use the label from the API response
                        sentiment_label = top_prediction['label'].capitalize()  # "Negative", "Neutral", or "Positive"
                    
                    except (IndexError, KeyError, TypeError) as e:
                        print(f"Error processing sentiment: {e}")
                        sentiment_label = "None"

                # Save the data to the database
                form = WebinarSeminarModel(
                    user=user,
                    
                    event=event,

                    relevance_of_the_activity=relevance_of_the_activity,
                    
                    quality_of_the_activity=quality_of_the_activity,

                    timeliness=timeliness,

                    suggestions_and_comments=suggestions_and_comments,

                    attainment_of_the_objective=attainment_of_the_objective,

                    appropriateness_of_the_topic_to_attain_the_objective = appropriateness_of_the_topic_to_attain_the_objective,

                    appropriateness_of_the_searching_methods_used=appropriateness_of_the_searching_methods_used,

                    topics_to_be_included=topics_to_be_included,

                    appropriateness_of_the_topic_in_the_present_time=appropriateness_of_the_topic_in_the_present_time,

                    usefulness_of_the_topic_discusssed_in_the_activity=usefulness_of_the_topic_discusssed_in_the_activity,

                    appropriateness_of_the_searching_methods=appropriateness_of_the_searching_methods,

                    displayed_a_thorough_knowledge_of_the_topic=displayed_a_thorough_knowledge_of_the_topic,

                    explained_activities=explained_activities,

                    able_to_create_a_good_learning_environment=able_to_create_a_good_learning_environment,

                    able_to_manage_her_time_well=able_to_manage_her_time_well,

                    demonstrated_keenness_to_the_participant_needs=demonstrated_keenness_to_the_participant_needs,

                    timeliness_or_suitability_of_service=timeliness_or_suitability_of_service,

                    overall_satisfaction=overall_satisfaction,

                    predicted_sentiment=sentiment_label

                )
                form.save()
                messages.success(request, 'Evaluation submitted successfully.')
                return redirect('faculty_events')

        return render(request, 'pages/faculty_webinar_seminar_form.html', context={'event': event, 'form': form, 'faculty': faculty, 'questions': questions, 'event_notifications': event_notifications,
        'notifications_unread_count': notifications_unread_count,
        'messages_notifications': messages_notifications,
        'messages_unread_count': messages_unread_count, 'is_department_head': is_department_head , 'is_head_of_osas': is_head_of_osas,})
    else:
        # Handle other event types
        pass

@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty', 'head of OSAS', 'department head/program coordinator'])
def peer_to_peer_evaluation(request):
    evaluation_status = EvaluationStatus.objects.first()
    user=request.user
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()
    faculty = Faculty.objects.filter(email=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)
    

    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()
    if faculty.is_supervisor:  
        # If supervisor, display all faculty except other supervisors
        faculty_department = Faculty.objects.filter(
            department=faculty.department, 
            is_supervisor=False  # Exclude supervisors
        ).order_by('first_name')
    else:
        # If not a supervisor, display only non-supervisors
        faculty_department = Faculty.objects.filter(
            department=faculty.department, 
            is_supervisor=False  # Include only non-supervisors
        ).order_by('first_name')
    user_evaluations = PeertoPeerEvaluation.objects.filter(
        user=request.user,
        academic_year=evaluation_status.academic_year,
        semester=evaluation_status.semester
    ).values_list('peer__pk', flat=True)  # Get evaluated faculty IDs

    evaluated_count = user_evaluations.count() 
    minimum_evaluations_required = 6 
    has_met_minimum_evaluations = evaluated_count >= minimum_evaluations_required

    return render(request, 'pages/peer_to_peer_evaluation.html', {'faculty': faculty, 'messages_notifications': messages_notifications, 'evaluation_status': evaluation_status, 'messages_unread_count': messages_unread_count, 'event_notifications': event_notifications, 'notifications_unread_count': notifications_unread_count, 'faculty_department': faculty_department, 'evaluated_faculty_ids': user_evaluations, 'has_met_minimum_evaluations': has_met_minimum_evaluations, 'evaluated_count': evaluated_count, 'minimum_evaluations_required': minimum_evaluations_required, 'is_department_head': is_department_head})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty', 'head of OSAS', 'department head/program coordinator'])
def peer_to_peer_evaluation_form(request,pk):
    user=request.user
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()
    faculty = Faculty.objects.filter(email=request.user.username).first()   
    peer = get_object_or_404(Faculty, pk=pk)
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()
    faculty_department = Faculty.objects.filter(department=faculty.department).order_by('first_name')
    questions = FacultyEvaluationQuestions.objects.all().order_by('order')
    user = request.user
    evaluation_status = EvaluationStatus.objects.first()
    user_evaluations = PeertoPeerEvaluation.objects.filter(
        user=request.user,
        peer=peer,
        academic_year=evaluation_status.academic_year,
        semester=evaluation_status.semester
    ).values_list('peer__pk', flat=True)  # Get evaluated faculty IDs
   
    if request.method == 'POST':
        form = PeertoPeerEvaluationForm(request.POST)
        if form.is_valid():
            # Process the evaluation form
            command_and_knowledge_of_the_subject = form.cleaned_data['command_and_knowledge_of_the_subject']
            depth_of_mastery = form.cleaned_data['depth_of_mastery']
            practice_in_respective_discipline = form.cleaned_data['practice_in_respective_discipline']
            up_to_date_knowledge = form.cleaned_data['up_to_date_knowledge']
            integrates_subject_to_practical_circumstances = form.cleaned_data['integrates_subject_to_practical_circumstances']

            organizes_the_subject_matter = form.cleaned_data['organizes_the_subject_matter']
            provides_orientation_on_course_content = form.cleaned_data['provides_orientation_on_course_content']
            efforts_of_class_preparation = form.cleaned_data['efforts_of_class_preparation']
            summarizes_main_points = form.cleaned_data['summarizes_main_points']
            monitors_online_class = form.cleaned_data['monitors_online_class']

            holds_interest_of_students = form.cleaned_data['holds_interest_of_students']
            provides_relevant_feedback = form.cleaned_data['provides_relevant_feedback']
            encourages_participation = form.cleaned_data['encourages_participation']
            shows_enthusiasm = form.cleaned_data['shows_enthusiasm']
            shows_sense_of_humor = form.cleaned_data['shows_sense_of_humor']

            teaching_methods = form.cleaned_data['teaching_methods']
            flexible_learning_strategies = form.cleaned_data['flexible_learning_strategies']
            student_engagement = form.cleaned_data['student_engagement']
            clear_examples = form.cleaned_data['clear_examples']
            focused_on_objectives = form.cleaned_data['focused_on_objectives']

            starts_with_motivating_activities = form.cleaned_data['starts_with_motivating_activities']
            speaks_in_clear_and_audible_manner = form.cleaned_data['speaks_in_clear_and_audible_manner']
            uses_appropriate_medium_of_instruction = form.cleaned_data['uses_appropriate_medium_of_instruction']
            establishes_online_classroom_environment = form.cleaned_data['establishes_online_classroom_environment']
            observes_proper_classroom_etiquette = form.cleaned_data['observes_proper_classroom_etiquette']

            uses_time_wisely = form.cleaned_data['uses_time_wisely']
            gives_ample_time_for_students_to_prepare = form.cleaned_data['gives_ample_time_for_students_to_prepare']
            updates_the_students_of_their_progress = form.cleaned_data['updates_the_students_of_their_progress']
            demonstrates_leadership_and_professionalism = form.cleaned_data['demonstrates_leadership_and_professionalism']
            understands_possible_distractions = form.cleaned_data['understands_possible_distractions']

            sensitivity_to_student_culture = form.cleaned_data['sensitivity_to_student_culture']
            responds_appropriately = form.cleaned_data['responds_appropriately']
            assists_students_on_concerns = form.cleaned_data['assists_students_on_concerns']
            guides_the_students_in_accomplishing_tasks = form.cleaned_data['guides_the_students_in_accomplishing_tasks']
            extends_consideration_to_students = form.cleaned_data['extends_consideration_to_students']


           # Extract the cleaned data from the form
            credit_task_preference = form.cleaned_data['credit_task_preference']
            
            # Convert the choice to a boolean value
            requires_less_task_for_credit = credit_task_preference == 'True'

            strengths_of_the_faculty = form.cleaned_data['strengths_of_the_faculty']
            other_suggestions_for_improvement =  form.cleaned_data['other_suggestions_for_improvement']
            comments = form.cleaned_data['comments']
            predicted_sentiment = get_sentiment(comments)
            sentiment_label = 'None'
            if predicted_sentiment and isinstance(predicted_sentiment, list):
                try:
                    # Access the first list in the response, then sort by score
                    sentiment_list = predicted_sentiment[0]
                    # Get the prediction with the highest score (sorted descending)
                    top_prediction = max(sentiment_list, key=lambda x: x['score'])
                    
                    # Directly use the label from the API response
                    sentiment_label = top_prediction['label'].capitalize()  # "Negative", "Neutral", or "Positive"
                
                except (IndexError, KeyError, TypeError) as e:
                    print(f"Error processing sentiment: {e}")
                    sentiment_label = "None"

          
            # Save the data to the database
            # Create the LikertEvaluation instance
            evaluation_instance = PeertoPeerEvaluation(
                peer=peer,
                user=user,
                command_and_knowledge_of_the_subject=command_and_knowledge_of_the_subject,
                depth_of_mastery=depth_of_mastery,
                practice_in_respective_discipline=practice_in_respective_discipline,
                up_to_date_knowledge=up_to_date_knowledge,
                integrates_subject_to_practical_circumstances=integrates_subject_to_practical_circumstances,
                organizes_the_subject_matter=organizes_the_subject_matter,
                provides_orientation_on_course_content=provides_orientation_on_course_content,
                efforts_of_class_preparation=efforts_of_class_preparation,
                summarizes_main_points=summarizes_main_points,
                monitors_online_class=monitors_online_class,
                holds_interest_of_students=holds_interest_of_students,
                provides_relevant_feedback=provides_relevant_feedback,
                encourages_participation=encourages_participation,
                shows_enthusiasm=shows_enthusiasm,
                shows_sense_of_humor=shows_sense_of_humor,
                teaching_methods=teaching_methods,
                flexible_learning_strategies=flexible_learning_strategies,
                student_engagement=student_engagement,
                clear_examples=clear_examples,
                focused_on_objectives=focused_on_objectives,
                starts_with_motivating_activities=starts_with_motivating_activities,
                speaks_in_clear_and_audible_manner=speaks_in_clear_and_audible_manner,
                uses_appropriate_medium_of_instruction=uses_appropriate_medium_of_instruction,
                establishes_online_classroom_environment=establishes_online_classroom_environment,
                observes_proper_classroom_etiquette=observes_proper_classroom_etiquette,
                uses_time_wisely=uses_time_wisely,
                gives_ample_time_for_students_to_prepare=gives_ample_time_for_students_to_prepare,
                updates_the_students_of_their_progress=updates_the_students_of_their_progress,
                demonstrates_leadership_and_professionalism=demonstrates_leadership_and_professionalism,
                understands_possible_distractions=understands_possible_distractions,
                sensitivity_to_student_culture=sensitivity_to_student_culture,
                responds_appropriately=responds_appropriately,
                assists_students_on_concerns=assists_students_on_concerns,
                guides_the_students_in_accomplishing_tasks=guides_the_students_in_accomplishing_tasks,
                extends_consideration_to_students=extends_consideration_to_students,
                requires_less_task_for_credit=requires_less_task_for_credit,
                strengths_of_the_faculty=strengths_of_the_faculty,
                other_suggestions_for_improvement=other_suggestions_for_improvement,
                comments=comments,
                predicted_sentiment=sentiment_label
            )

            # Attempt to save the instance
            try:
                evaluation_instance.save()
                messages.success(request, 'Evaluation submitted successfully.')
                return redirect('peer_to_peer_evaluation')  # Redirect to a success page
            except Exception as e:
                # If the save fails, print the error message and return a failure message
                print(f"Error saving evaluation: {e}")
                messages.error(request, 'Failed to submit evaluation. Please try again.')

        else:
            # If the form is invalid, print form errors
            print(form.errors)
            messages.error(request, 'There was an error with your submission. Please check your input.')
            return redirect('peer_to_peer_evaluation')  # Redirect to a success page

    else:
        form = PeertoPeerEvaluationForm()

    context = { 'form': form, 'faculty_department': faculty_department, 'messages_notifications': messages_notifications, 'faculty': faculty, 'questions': questions, 'evaluated_faculty_ids': user_evaluations, 'event_notifications': event_notifications, 'notifications_unread_count': notifications_unread_count, 'messages_unread_count': messages_unread_count, 'peer': peer, 'is_department_head': is_department_head}
    return render(request, 'pages/peer_to_peer_evaluation_form.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty', 'head of OSAS', 'department head/program coordinator'])
def peer_to_peer_evaluations(request):
    evaluation_status = EvaluationStatus.objects.first()
    user=request.user
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()
    faculty = Faculty.objects.filter(email=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()

    peer_to_peer_evaluations = PeertoPeerEvaluation.objects.filter(peer=faculty)

    #filter and search
    faculty_evaluation_filter = PeertoPeerEvaluationFilter(request.GET, queryset=peer_to_peer_evaluations)
    peer_to_peer_evaluations = faculty_evaluation_filter.qs
    

    # ordering functionality
   
    ordering = request.GET.get('ordering', "")

     
    if ordering:
        peer_to_peer_evaluations = peer_to_peer_evaluations.order_by(ordering) 

    #pagination
    page_number = request.GET.get('page', 1)
    evaluation_paginator = Paginator(peer_to_peer_evaluations, ITEMS_PER_PAGE)
   

    try:
        page = evaluation_paginator.page(page_number)
    except EmptyPage:
        page = evaluation_paginator.page(evaluation_paginator.num_pages)

    return render(request, 'pages/peer_to_peer_evaluations.html', {'faculty': faculty, 'messages_notifications': messages_notifications, 'evaluation_status': evaluation_status, 'messages_unread_count': messages_unread_count, 'event_notifications': event_notifications, 'notifications_unread_count': notifications_unread_count, 'peer_to_peer_evaluations': peer_to_peer_evaluations, 'faculty_evaluation_filter': faculty_evaluation_filter, 'page_obj':page, 'is_paginated': True, 'paginator':evaluation_paginator,'ordering': ordering, 'is_department_head': is_department_head})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty', 'head of OSAS', 'department head/program coordinator'])
def peer_to_peer_evaluations_csv(request):
    faculty = Faculty.objects.filter(email=request.user.username).first()
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=pper_to_peer_evaluations.xlsx'

    # Create a workbook and a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Peer-to-Peer Evaluations"

    evaluation_status = EvaluationStatus.objects.first()
    current_academic_year = evaluation_status.academic_year 
    current_semester = evaluation_status.semester

    academic_year = request.GET.get('academic_year', current_academic_year)
    semester = request.GET.get('semester', current_semester)
    # Apply filters from the EvaluationFilter based on the request data
    evaluation_filter = EvaluationFilter(request.GET, queryset=PeertoPeerEvaluation.objects.filter(peer=faculty, academic_year=academic_year, semester=semester))

    # Get the filtered queryset
    filtered_evaluations = evaluation_filter.qs

    # Add column headings to the worksheet
    columns = ['Faculty', 'Average', 'Rating', 'Overall Impression', 'Polarity', 'Academic Year', 'Semester']
    ws.append(columns)

    # Loop through and output the data
    for evaluation in filtered_evaluations:
        row = [
            str(evaluation.peer),
            str(evaluation.average_rating),
            str(evaluation.get_rating_category()),
            str(evaluation.comments),
            str(evaluation.predicted_sentiment),
            str(evaluation.academic_year),
            str(evaluation.semester)
        ]
        ws.append(row)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook to a BytesIO stream
    output = BytesIO()
    wb.save(output)
    response.write(output.getvalue())

    return response

@login_required(login_url='signin')
@allowed_users(allowed_roles=['faculty', 'head of OSAS', 'department head/program coordinator'])
def view_peer_to_peer_evaluation_form(request, pk):
    user=request.user
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()
    faculty = Faculty.objects.filter(email=request.user.username).first()   

    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()
    
    questions = PeertoPeerEvaluationQuestions.objects.all().order_by('order')
    faculty_evaluation_form = PeertoPeerEvaluation.objects.get(pk=pk)
    rating_fields = [
        'command_and_knowledge_of_the_subject',
        'depth_of_mastery',
        'practice_in_respective_discipline',
        'up_to_date_knowledge',
        'integrates_subject_to_practical_circumstances',
        'organizes_the_subject_matter',
        'provides_orientation_on_course_content',
        'efforts_of_class_preparation',
        'summarizes_main_points',
        'monitors_online_class',
        'holds_interest_of_students',
        'provides_relevant_feedback',
        'encourages_participation',
        'shows_enthusiasm',
        'shows_sense_of_humor',
        'teaching_methods',
        'flexible_learning_strategies',
        'student_engagement',
        'clear_examples',
        'focused_on_objectives',
        'starts_with_motivating_activities',
        'speaks_in_clear_and_audible_manner',
        'uses_appropriate_medium_of_instruction',
        'establishes_online_classroom_environment',
        'observes_proper_classroom_etiquette',
        'uses_time_wisely',
        'gives_ample_time_for_students_to_prepare',
        'updates_the_students_of_their_progress',
        'demonstrates_leadership_and_professionalism',
        'understands_possible_distractions',
        'sensitivity_to_student_culture',
        'responds_appropriately',
        'assists_students_on_concerns',
        'guides_the_students_in_accomplishing_tasks',
        'extends_consideration_to_students'
    ]
    # Initialize counters
    outstanding_count = 0
    very_satisfactory_count = 0
    satisfactory_count = 0
    unsatisfactory_count = 0
    poor_count = 0

    # Calculate counts in memory
    for field in rating_fields:
        value = getattr(faculty_evaluation_form, field)
        if value == 5:
            outstanding_count += 1
        elif value == 4:
            very_satisfactory_count += 1
        elif value == 3:
            satisfactory_count += 1
        elif value == 2:
            unsatisfactory_count += 1
        elif value == 1:
            poor_count += 1
    
    return render(request, 'pages/view_peer_to_peer_evaluation_form.html', {'faculty_evaluation_form': faculty_evaluation_form, 'faculty': faculty, 'outstanding_count': outstanding_count, 'very_satisfactory_count': very_satisfactory_count, 'satisfactory_count': satisfactory_count, 'unsatisfactory_count': unsatisfactory_count, 'poor_count': poor_count, 'questions': questions, 'event_notifications': event_notifications,
        'notifications_unread_count': notifications_unread_count,
        'messages_notifications': messages_notifications,
        'messages_unread_count': messages_unread_count,'is_department_head': is_department_head})


@login_required(login_url='signin')
@allowed_users(allowed_roles=['head of OSAS', 'department head/program coordinator'])
def department_head_view_department(request):
    user=request.user
    evaluation_status = EvaluationStatus.objects.first()
    current_academic_year = evaluation_status.academic_year 
    current_semester = evaluation_status.semester
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()    
    faculty = Faculty.objects.filter(email=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)
    
    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()

    department = Department.objects.get(name=faculty.department)
    
    first_academic_year = LikertEvaluation.objects.aggregate(Min('academic_year'))['academic_year__min']
    if first_academic_year:
        start_year, _ = map(int, first_academic_year.split('-'))
        current_start_year, _ = map(int, current_academic_year.split('-'))
        academic_years = [f"{year} - {year + 1}" for year in range(start_year, current_start_year + 1)]
    else:
        academic_years = [current_academic_year]  # Fallback to the current academic year if no data exists

    semesters = ["1st Semester", "2nd Semester"]

    # Get query parameters for filtering
    academic_year = request.GET.get('academic_year', current_academic_year)
    semester = request.GET.get('semester', current_semester)
    
        # Parse the academic year to determine the previous academic year
    start_year, end_year = map(int, academic_year.split('-'))

    if semester == "1st Semester":
        # Previous semester is the "2nd" semester of the previous academic year
        previous_academic_year = f"{start_year - 1} - {end_year - 1}"
        previous_semester = "2nd Semester"
    else:
        # Previous semester is the "1st" semester of the same academic year
        previous_academic_year = academic_year
        previous_semester = "1st Semester"

    # Query the EvaluationStatus model for the previous semester's data
    previous_evaluation_status = EvaluationStatus.objects.filter(
        Q(academic_year=previous_academic_year) & Q(semester=previous_semester)
    ).first()

    if previous_evaluation_status:
        print(f"Previous Semester: {previous_evaluation_status.academic_year}, {previous_evaluation_status.semester}")
    else:
        print(start_year, end_year, previous_academic_year, previous_semester)
        print("No previous semester data found.")

    # Base queryset for SET evaluations
    faculties = Faculty.objects.filter(department=request.user.faculty.department).annotate(
       average_rating=Round(
           Coalesce(
        Avg(
            'sectionsubjectfaculty__likertevaluation__average_rating',
            filter=Q(
                sectionsubjectfaculty__likertevaluation__academic_year=academic_year,
                sectionsubjectfaculty__likertevaluation__semester=semester,
                sectionsubjectfaculty__likertevaluation__admin_status='Approved'
            ),
        ),
        Value(0), output_field=FloatField()  # Default value for null average_rating
       ), 2
      ),
       previous_average_rating=Round(
           Coalesce(
        Avg(
            'sectionsubjectfaculty__likertevaluation__average_rating',
            filter=Q(
                sectionsubjectfaculty__likertevaluation__academic_year=previous_academic_year,
                sectionsubjectfaculty__likertevaluation__semester=previous_semester,
            ),
        ),
        Value(0), output_field=FloatField()  # Default value for null average_rating
       ), 2
      ),        
      previous_rank=Window(
            expression=Rank(),
            order_by=F('previous_average_rating').desc(nulls_last=True)
        ),
      num_of_evaluators=Count(
            'sectionsubjectfaculty__likertevaluation',
            filter=Q(
                sectionsubjectfaculty__likertevaluation__academic_year=academic_year,
                sectionsubjectfaculty__likertevaluation__semester=semester,
                sectionsubjectfaculty__likertevaluation__admin_status='Approved'
            ),
        ),  
      previous_num_of_evaluators=Count(
            'sectionsubjectfaculty__likertevaluation',
            filter=Q(
                sectionsubjectfaculty__likertevaluation__academic_year=previous_academic_year,
                sectionsubjectfaculty__likertevaluation__semester=previous_semester,
                sectionsubjectfaculty__likertevaluation__admin_status='Approved'
            ),
        ),  
    ).order_by('-average_rating')  

    
    is_supervisor = faculty.is_supervisor
    department_faculty_filter = FacultyFilter(request.GET, queryset=faculties)
    faculties = department_faculty_filter.qs
    



    # Handle sorting
    ordering = request.GET.get('sort')
    if ordering in ['last_name', '-last_name', 'average_rating', '-average_rating', 'previous_average_rating', '-previous_average_rating']:
        faculties = faculties.order_by(ordering)
 


    form = TeacherForm()
    if request.method == 'POST':
        form = TeacherForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Faculty added successfully')
            return redirect('department_head_view_department')



    paginator = Paginator(faculties, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'department': department,
        'faculties': faculties,
        'page_obj': page_obj, 
        'is_department_head': is_department_head,
        'event_notifications': event_notifications,
        'messages_notifications': messages_notifications,
        'notifications_unread_count': notifications_unread_count,
        'messages_unread_count': messages_unread_count,
        'faculty': faculty,
        'form': form,
        'is_supervisor': is_supervisor,
        'department_faculty_filter': department_faculty_filter,
        'previous_evaluation_status': previous_evaluation_status,
        'academic_years': academic_years,
        'semesters': semesters,
        'academic_year': academic_year,
        'semester': semester,

    }

    return render(request, 'pages/department_head_view_department.html', context)
@login_required(login_url='signin')
@allowed_users(allowed_roles=['head of OSAS', 'department head/program coordinator'])
def department_head_view_department_peer_to_peer(request):
    user=request.user
    evaluation_status = EvaluationStatus.objects.first()
    current_academic_year = evaluation_status.academic_year 
    current_semester = evaluation_status.semester
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()
    faculty = Faculty.objects.filter(email=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)
    
    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()

    department = Department.objects.get(name=faculty.department)

    first_academic_year = LikertEvaluation.objects.aggregate(Min('academic_year'))['academic_year__min']
    if first_academic_year:
        start_year, _ = map(int, first_academic_year.split('-'))
        current_start_year, _ = map(int, current_academic_year.split('-'))
        academic_years = [f"{year} - {year + 1}" for year in range(start_year, current_start_year + 1)]
    else:
        academic_years = [current_academic_year]  # Fallback to the current academic year if no data exists

    semesters = ["1st Semester", "2nd Semester"]

    # Get query parameters for filtering
    academic_year = request.GET.get('academic_year', current_academic_year)
    semester = request.GET.get('semester', current_semester)



        # Parse the academic year to determine the previous academic year
    start_year, end_year = map(int, academic_year.split('-'))

    if semester == "1st Semester":
        # Previous semester is the "2nd" semester of the previous academic year
        previous_academic_year = f"{start_year - 1} - {end_year - 1}"
        previous_semester = "2nd Semester"
    else:
        # Previous semester is the "1st" semester of the same academic year
        previous_academic_year = academic_year
        previous_semester = "1st Semester"

    # Query the EvaluationStatus model for the previous semester's data
    previous_evaluation_status = EvaluationStatus.objects.filter(
        Q(academic_year=previous_academic_year) & Q(semester=previous_semester)
    ).first()

    if previous_evaluation_status:
        print(f"Previous Semester: {previous_evaluation_status.academic_year}, {previous_evaluation_status.semester}")
    else:
        print(start_year, end_year, previous_academic_year, previous_semester)
        print("No previous semester data found.")

    # Base queryset for peer-to-peer evaluations
    faculties = Faculty.objects.filter(department=request.user.faculty.department).annotate(
       average_rating=Round(
           Coalesce(
        Avg(
            'peertopeerevaluation__average_rating',
            filter=Q(
                peertopeerevaluation__academic_year=academic_year,
                peertopeerevaluation__semester=semester
            ),
        ),
        Value(0), output_field=FloatField()  # Default value for null average_rating
       ), 2
      ),
       previous_average_rating=Round(
           Coalesce(
        Avg(
            'peertopeerevaluation__average_rating',
            filter=Q(
                peertopeerevaluation__academic_year=previous_academic_year,
                peertopeerevaluation__semester=previous_semester,
            ),
        ),
        Value(0), output_field=FloatField()  # Default value for null average_rating
       ), 2
      ),        
      previous_rank=Window(
            expression=Rank(),
            order_by=F('previous_average_rating').desc(nulls_last=True)
        ),
      num_of_evaluators=Count(
            'peertopeerevaluation',
            filter=Q(
                peertopeerevaluation__academic_year=academic_year,
                peertopeerevaluation__semester=semester,
            ),
        ),
    ).order_by('-average_rating')  

    
    is_supervisor = faculty.is_supervisor
    department_faculty_filter = FacultyFilter(request.GET, queryset=faculties)
    faculties = department_faculty_filter.qs
    



    # Handle sorting
    ordering = request.GET.get('sort')
    if ordering in ['last_name', '-last_name', 'average_rating', '-average_rating', 'previous_average_rating', '-previous_average_rating']:
        faculties = faculties.order_by(ordering)
 


    form = TeacherForm()
    if request.method == 'POST':
        form = TeacherForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Faculty added successfully')
            return redirect('department_head_view_department')



    paginator = Paginator(faculties, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'department': department,
        'faculties': faculties,
        'page_obj': page_obj, 
        'is_department_head': is_department_head,
        'event_notifications': event_notifications,
        'messages_notifications': messages_notifications,
        'notifications_unread_count': notifications_unread_count,
        'messages_unread_count': messages_unread_count,
        'faculty': faculty,
        'form': form,
        'is_supervisor': is_supervisor,
        'department_faculty_filter': department_faculty_filter,
        'previous_evaluation_status': previous_evaluation_status,
        'academic_years': academic_years,
        'semesters': semesters,
        'academic_year': academic_year,
        'semester': semester
    }

    return render(request, 'pages/department_head_view_department_peer_to_peer.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['head of OSAS', 'department head/program coordinator'])
def mark_as_supervisor(request, pk):
    if request.method == 'POST':
        faculty = get_object_or_404(Faculty, pk=pk)
        faculty.is_supervisor = True  # Ensure this field exists in the Faculty model
        faculty.save()
        messages.success(request, f"{faculty.first_name} {faculty.last_name} has been successfully marked as a supervisor.")
    return redirect('department_head_view_department')  # Replace with the appropriate redirect URL

@login_required(login_url='signin')
@allowed_users(allowed_roles=['head of OSAS', 'department head/program coordinator'])
def unmark_as_supervisor(request, pk):
    if request.method == 'POST':
        faculty = get_object_or_404(Faculty, pk=pk)
        faculty.is_supervisor = False  # Ensure this field exists in the Faculty model
        faculty.save()
        messages.success(request, f"{faculty.first_name} {faculty.last_name} has been successfully unmarked as a supervisor.")
    return redirect('department_head_view_department')  # Replace with the appropriate redirect URL

@login_required(login_url='signin')
@allowed_users(allowed_roles=['head of OSAS', 'department head/program coordinator'])
def download_report(request, department_id):
    department = get_object_or_404(Department, id=department_id)

    # Path to the department-specific PDF
    pdf_path = os.path.join(settings.MEDIA_ROOT, f'faculty_evaluations_summary_{department_id}.pdf')
    if os.path.exists(pdf_path):
        return FileResponse(open(pdf_path, 'rb'), as_attachment=True, filename=f'{department.name}_Summary_Report.pdf')
    else:
        return HttpResponse("Report not found.")
    
@login_required(login_url='signin')
@allowed_users(allowed_roles=['head of OSAS', 'department head/program coordinator'])
def send_summary_reports(request):
    evaluation_status = EvaluationStatus.objects.first()
    current_academic_year = evaluation_status.academic_year
    current_semester = evaluation_status.semester
    if request.method == 'POST':
        faculty_ids = request.POST.getlist('selected_faculties')  # Get selected faculty IDs
        selected_faculties = Faculty.objects.filter(id__in=faculty_ids)

        evaluation_status = EvaluationStatus.objects.first()
        current_academic_year = evaluation_status.academic_year
        current_semester = evaluation_status.semester

        # Generate and send individual reports for each selected faculty
        for faculty in selected_faculties:
            evaluations = LikertEvaluation.objects.filter(
                section_subject_faculty__faculty=faculty,
                academic_year=current_academic_year,
                semester=current_semester
            )
            num_evaluators = evaluations.count()

            summary_data = []
            comments_data = []

            if num_evaluators > 0:
                category_sums = {
                    'Subject Matter Content': 0,
                    'Organization': 0,
                    'Teacher-Student Rapport': 0,
                    'Teaching Methods': 0,
                    'Presentation': 0,
                    'Classroom Management': 0,
                    'Sensitivity and Support to Students': 0,
                    'Overall': 0
                }

                for evaluation in evaluations:
                    category_averages = evaluation.calculate_category_averages()
                    for category, average in category_averages.items():
                        if average is not None:
                            category_sums[category] += average
                    category_sums['Overall'] += evaluation.average_rating

                category_averages = {category: round(total / num_evaluators, 2) for category, total in category_sums.items()}
                avg_rating = category_averages['Overall']
                rating_category = (
                    "Poor" if avg_rating <= 1.49 else
                    "Unsatisfactory" if avg_rating <= 2.49 else
                    "Satisfactory" if avg_rating <= 3.49 else
                    "Very Satisfactory" if avg_rating <= 4.49 else
                    "Outstanding"
                )

                summary_data.append({
                    'faculty': f"{faculty.first_name} {faculty.last_name}",
                    'num_evaluators': num_evaluators,
                    'subject_matter_content_avg': category_averages['Subject Matter Content'],
                    'organization_avg': category_averages['Organization'],
                    'teacher_student_rapport_avg': category_averages['Teacher-Student Rapport'],
                    'teaching_methods_avg': category_averages['Teaching Methods'],
                    'presentation_avg': category_averages['Presentation'],
                    'classroom_management_avg': category_averages['Classroom Management'],
                    'sensitivity_support_students_avg': category_averages['Sensitivity and Support to Students'],
                    'overall_avg': category_averages['Overall'],
                    'rating_category': rating_category
                })

                for evaluation in evaluations:
                    comments_data.append({
                        'faculty': f"{faculty.first_name} {faculty.last_name}",
                        'requires_less_task_for_credit': evaluation.requires_less_task_for_credit,
                        'strengths_of_the_faculty': evaluation.strengths_of_the_faculty,
                        'other_suggestions_for_improvement': evaluation.other_suggestions_for_improvement,
                        'comments': evaluation.comments
                    })

            # Render the summary report for this faculty
            image_path = os.path.join(settings.BASE_DIR, 'static/images/cvsulogo.png')
            html = render_to_string('pages/faculty_evaluations_summary_report.html', {
                'summary_data': summary_data,
                'image_path': image_path,
                'comments_data': comments_data,
                'current_academic_year': current_academic_year,
                'current_semester': current_semester
            })

            pdf_path = os.path.join(settings.MEDIA_ROOT, f'faculty_evaluations_summary_{faculty.id}.pdf')
            with open(pdf_path, 'wb') as pdf_file:
                pisa_status = pisa.CreatePDF(html, dest=pdf_file, link_callback=link_callback)

            if pisa_status.err:
                return HttpResponse(f'Error generating PDF for faculty {faculty.full_name}')

            # Send the email to the faculty
            email = EmailMessage(
                subject=f'Faculty Evaluations Summary Report - {faculty.full_name}',
                body='Please find attached your faculty evaluations summary report.',
                from_email='admin@example.com',
                to=[faculty.email],
            )
            email.attach_file(pdf_path)
            email.send()

            faculty.email_sent = True
            faculty.save()

            notify.send(request.user, 
                        recipient=faculty.user, 
                        verb='Summary Report Sent', 
                        description='Your faculty evaluation summary report has been sent to your email.',
                        level='info')

        messages.success(request, "The individual summary reports have been successfully sent to the selected faculty members.") 
        return redirect('department_head_view_department')  # Redirect to a success page
    
@login_required(login_url='signin')
@allowed_users(allowed_roles=['head of OSAS', 'department head/program coordinator'])
def department_head_faculty_evaluations(request, pk):
    user = request.user
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()
    faculty = Faculty.objects.filter(email=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()
    evaluation_status = EvaluationStatus.objects.first()
    current_academic_year = evaluation_status.academic_year 
    current_semester = evaluation_status.semester
    current_status = evaluation_status.evaluation_status
    academic_year = request.GET.get('academic_year', current_academic_year)
    semester = request.GET.get('semester', current_semester)
    teacher = get_object_or_404(Faculty, pk=pk)
    teacher_evaluations = LikertEvaluation.objects.filter(section_subject_faculty__faculty=teacher,academic_year=academic_year, semester=semester)
   
         # Create mutable copy of GET parameters
    filter_params = request.GET.copy()

    # Set defaults only when values are missing or empty
    if not filter_params.get('academic_year'):
        filter_params['academic_year'] = current_academic_year
    if not filter_params.get('semester'):
        filter_params['semester'] = current_semester

    #filter and search
    faculty_evaluation_filter = EvaluationFilter(filter_params, queryset=LikertEvaluation.objects.filter(section_subject_faculty__faculty=teacher, admin_status='Approved'), faculty=teacher)
    
    if not request.GET or (request.GET.get('academic_year') == '' and request.GET.get('semester') == ''):
        teacher_evaluations = LikertEvaluation.objects.filter(section_subject_faculty__faculty=teacher, admin_status='Approved', academic_year=current_academic_year, semester=current_semester)
    else:    
        teacher_evaluations = faculty_evaluation_filter.qs
    
    # ordering functionality
   
    ordering = request.GET.get('ordering', "")

     
    if ordering:
        teacher_evaluations = teacher_evaluations.order_by(ordering) 

    #pagination
    page_number = request.GET.get('page', 1)
    evaluation_paginator = Paginator(teacher_evaluations, 10)

    try:
        page = evaluation_paginator.page(page_number)
    except EmptyPage:
        page = evaluation_paginator.page(evaluation_paginator.num_pages)

    context = {'teacher': teacher, 'teacher_evaluations':  teacher_evaluations, 'is_department_head': is_department_head, 'event_notifications': event_notifications,
        'messages_notifications': messages_notifications,
        'notifications_unread_count': notifications_unread_count,
        'messages_unread_count': messages_unread_count,
        'faculty': faculty, 'faculty_evaluation_filter': faculty_evaluation_filter, 'page_obj':page, 'is_paginated': True, 'paginator':evaluation_paginator, 'teacher_evaluations': page.object_list,'current_status': current_status, 'current_academic_year': current_academic_year, 'current_semester': current_semester, }

    return render(request, 'pages/department_head_faculty_evaluations.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['head of OSAS', 'department head/program coordinator'])
def department_head_peer_to_peer_evaluations(request, pk):
    user = request.user
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()
    faculty = Faculty.objects.filter(email=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()
    evaluation_status = EvaluationStatus.objects.first()
    current_academic_year = evaluation_status.academic_year 
    current_semester = evaluation_status.semester
    current_status = evaluation_status.evaluation_status
    academic_year = request.GET.get('academic_year', current_academic_year)
    semester = request.GET.get('semester', current_semester)
    teacher = get_object_or_404(Faculty, pk=pk)
    teacher_evaluations = PeertoPeerEvaluation.objects.filter(peer=teacher,academic_year=academic_year, semester=semester)
   
    # Create mutable copy of GET parameters
    filter_params = request.GET.copy()

    # Set defaults only when values are missing or empty
    if not filter_params.get('academic_year'):
        filter_params['academic_year'] = current_academic_year
    if not filter_params.get('semester'):
        filter_params['semester'] = current_semester

    #filter and search
    faculty_evaluation_filter = PeertoPeerEvaluationFilter(filter_params, queryset=PeertoPeerEvaluation.objects.filter(peer=teacher))
    
    if not request.GET or (request.GET.get('academic_year') == '' and request.GET.get('semester') == ''):
        teacher_evaluations = PeertoPeerEvaluation.objects.filter(peer=teacher, academic_year=current_academic_year, semester=current_semester)
    else:    
        teacher_evaluations = faculty_evaluation_filter.qs
    
    # ordering functionality
   
    ordering = request.GET.get('ordering', "")

     
    if ordering:
        teacher_evaluations = teacher_evaluations.order_by(ordering) 

    #pagination
    page_number = request.GET.get('page', 1)
    evaluation_paginator = Paginator(teacher_evaluations, 10)

    try:
        page = evaluation_paginator.page(page_number)
    except EmptyPage:
        page = evaluation_paginator.page(evaluation_paginator.num_pages)

    context = {'teacher': teacher, 'teacher_evaluations':  teacher_evaluations, 'is_department_head': is_department_head, 'event_notifications': event_notifications,
        'messages_notifications': messages_notifications,
        'notifications_unread_count': notifications_unread_count,
        'messages_unread_count': messages_unread_count,
        'faculty': faculty, 'faculty_evaluation_filter': faculty_evaluation_filter, 'page_obj':page, 'is_paginated': True, 'paginator':evaluation_paginator, 'teacher_evaluations': page.object_list,'current_status': current_status, 'current_academic_year': current_academic_year, 'current_semester': current_semester, 'academic_year':academic_year, 'semester':semester }

    return render(request, 'pages/department_head_peer_to_peer_evaluations.html', context)

@login_required(login_url='signin')
@allowed_users(allowed_roles=['head of OSAS', 'department head/program coordinator'])
def department_head_send_message(request, pk):
    user = request.user
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()
    faculty = Faculty.objects.filter(email=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()
    faculty = get_object_or_404(Faculty, pk=pk)
    faculty_user = faculty.user
    if request.method == 'POST':
        form = MessageForm(request.POST, request.FILES)
        if form.is_valid():
            message = form.save(commit=False)
            message.sender = user
            message.recipient = faculty_user
            message.save()

            # Send email with attachment if it exists
            email = EmailMessage(
                subject=f'New Message: {message.subject}',
                body=message.content,
                from_email='your-email@example.com',
                to=[faculty.email],
            )
            if message.attachment:
                email.attach(message.attachment.name, message.attachment.read())
            email.send()

            notify.send(
                sender=user,
                recipient=faculty_user,
                verb=message.subject,
                description=message.content
            )
            return redirect('department_head_view_department')
    else:
        form = MessageForm()
    return render(request, 'pages/department_head_send_message.html', {'form': form, 'faculty': faculty, 'is_department_head': is_department_head, 'event_notifications': event_notifications,
        'messages_notifications': messages_notifications,
        'notifications_unread_count': notifications_unread_count,
        'messages_unread_count': messages_unread_count,
        'faculty': faculty})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['head of OSAS', 'department head/program coordinator'])
def department_head_pending_evaluations(request):
    user = request.user
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()
    faculty = Faculty.objects.filter(email=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()
    evaluations = LikertEvaluation.objects.filter(admin_status='Approved to Department head/program coordinator',  section_subject_faculty__faculty__department=faculty.department).order_by('-updated')
    paginator = Paginator(evaluations, 10) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request,'pages/department_head_pending_evaluations.html',{'evaluations': evaluations, 'page_obj': page_obj, 'faculty': faculty, 'is_department_head': is_department_head, 'event_notifications': event_notifications,
        'messages_notifications': messages_notifications,
        'notifications_unread_count': notifications_unread_count,
        'messages_unread_count': messages_unread_count,
        'faculty': faculty})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['head of OSAS','department head/program coordinator'])
def department_head_approve_evaluation(request, pk):
    evaluation = get_object_or_404(LikertEvaluation, pk=pk, admin_status='Approved to Department head/program coordinator')
    
    evaluation.admin_status = 'Approved'
    evaluation.save()
                
    return redirect('department_head_pending_evaluations')

@login_required(login_url='signin')
@allowed_users(allowed_roles=['head of OSAS','department head/program coordinator'])
def department_head_reject_evaluation(request, pk):
    evaluation = get_object_or_404(LikertEvaluation, pk=pk, admin_status='Pending')
    
    evaluation.admin_status = 'Rejected'
    evaluation.save()
                
    return redirect('department_head_pending_evaluations')

@login_required(login_url='signin')
@allowed_users(allowed_roles=['head of OSAS','department head/program coordinator']) 
def department_head_approve_all_pending_evaluations(request):
     if request.method == 'POST': LikertEvaluation.objects.filter(admin_status='Approved to Department head/program coordinator').update(admin_status='Approved') 
     messages.success(request, 'All pending evaluations have been approved.') 
     return redirect('department_head_pending_evaluations')

@login_required(login_url='signin')
@allowed_users(allowed_roles=['head of OSAS', 'department head/program coordinator'])
def department_head_faculty_evaluations_csv(request, faculty_pk=None):
    if faculty_pk is not None:
        teacher = get_object_or_404(Faculty, pk=faculty_pk)
    else:
        # Option B: If passed as a query parameter
        faculty_pk = request.GET.get('faculty_pk')
        teacher = get_object_or_404(Faculty, pk=faculty_pk)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=faculty_evaluations.csv'

    # Create a csv writer
    writer = csv.writer(response)
    evaluation_status = EvaluationStatus.objects.first()
    current_academic_year = evaluation_status.academic_year 
    current_semester = evaluation_status.semester
    academic_year = request.GET.get('academic_year', current_academic_year)
    semester = request.GET.get('semester', current_semester)
    # Apply filters from the EvaluationFilter based on the request data
    evaluation_filter = EvaluationFilter(request.GET, queryset=LikertEvaluation.objects.filter(section_subject_faculty__faculty=teacher, academic_year=academic_year, semester=semester))

    # Get the filtered queryset
    filtered_evaluations = evaluation_filter.qs

    # Add column headings to csv file

    writer.writerow(['Subject', 'Faculty', 'Average', 'Rating', 'Overall Impression', 'Polarity', 'Academic Year', 'Semester'])

    # Loop thru and output
    for i in filtered_evaluations:
        writer.writerow([i.section_subject_faculty.subjects, i.section_subject_faculty.faculty, i.average_rating, i.get_rating_category(), i.comments, i.predicted_sentiment, i.academic_year, i.semester ])

    return response

@login_required(login_url='signin')
@allowed_users(allowed_roles=['head of OSAS', 'department head/program coordinator'])
def department_head_peer_to_peer_faculty_evaluations_csv(request, faculty_pk=None):
    if faculty_pk is not None:
        teacher = get_object_or_404(Faculty, pk=faculty_pk)
    else:
        # Option B: If passed as a query parameter
        faculty_pk = request.GET.get('faculty_pk')
        teacher = get_object_or_404(Faculty, pk=faculty_pk)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=peer_to_peer_evaluations.csv'

    # Create a csv writer
    writer = csv.writer(response)
    evaluation_status = EvaluationStatus.objects.first()
    current_academic_year = evaluation_status.academic_year 
    current_semester = evaluation_status.semester
    academic_year = request.GET.get('academic_year', current_academic_year)
    semester = request.GET.get('semester', current_semester)
    # Apply filters from the EvaluationFilter based on the request data
    evaluation_filter = EvaluationFilter(request.GET, queryset=PeertoPeerEvaluation.objects.filter(peer=teacher, academic_year=academic_year, semester=semester))

    # Get the filtered queryset
    filtered_evaluations = evaluation_filter.qs

    # Add column headings to csv file

    writer.writerow(['Faculty', 'Average', 'Rating', 'Overall Impression', 'Polarity', 'Academic Year', 'Semester'])

    # Loop thru and output
    for i in filtered_evaluations:
        writer.writerow([i.peer, i.average_rating, i.get_rating_category(), i.comments, i.predicted_sentiment, i.academic_year, i.semester ])

    return response

@login_required(login_url='signin')
@allowed_users(allowed_roles=['head of OSAS','department head/program coordinator'])
def department_head_view_evaluation_form(request, pk):
    faculty_evaluation_form = get_object_or_404(LikertEvaluation, pk=pk)
    questions = FacultyEvaluationQuestions.objects.all().order_by('order')
    user = request.user
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()
    faculty = Faculty.objects.filter(email=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()
         # List of all rating fields to check
    rating_fields = [
        'command_and_knowledge_of_the_subject',
        'depth_of_mastery',
        'practice_in_respective_discipline',
        'up_to_date_knowledge',
        'integrates_subject_to_practical_circumstances',
        'organizes_the_subject_matter',
        'provides_orientation_on_course_content',
        'efforts_of_class_preparation',
        'summarizes_main_points',
        'monitors_online_class',
        'holds_interest_of_students',
        'provides_relevant_feedback',
        'encourages_participation',
        'shows_enthusiasm',
        'shows_sense_of_humor',
        'teaching_methods',
        'flexible_learning_strategies',
        'student_engagement',
        'clear_examples',
        'focused_on_objectives',
        'starts_with_motivating_activities',
        'speaks_in_clear_and_audible_manner',
        'uses_appropriate_medium_of_instruction',
        'establishes_online_classroom_environment',
        'observes_proper_classroom_etiquette',
        'uses_time_wisely',
        'gives_ample_time_for_students_to_prepare',
        'updates_the_students_of_their_progress',
        'demonstrates_leadership_and_professionalism',
        'understands_possible_distractions',
        'sensitivity_to_student_culture',
        'responds_appropriately',
        'assists_students_on_concerns',
        'guides_the_students_in_accomplishing_tasks',
        'extends_consideration_to_students'
    ]
    # Initialize counters
    outstanding_count = 0
    very_satisfactory_count = 0
    satisfactory_count = 0
    unsatisfactory_count = 0
    poor_count = 0

    # Calculate counts in memory
    for field in rating_fields:
        value = getattr(faculty_evaluation_form, field)
        if value == 5:
            outstanding_count += 1
        elif value == 4:
            very_satisfactory_count += 1
        elif value == 3:
            satisfactory_count += 1
        elif value == 2:
            unsatisfactory_count += 1
        elif value == 1:
            poor_count += 1

    return render(request, 'pages/department_head_view_evaluation_form.html', {'faculty_evaluation_form': faculty_evaluation_form, 'faculty': faculty, 'questions': questions, 'outstanding_count': outstanding_count, 'very_satisfactory_count': very_satisfactory_count, 'satisfactory_count': satisfactory_count, 'unsatisfactory_count': unsatisfactory_count, 'poor_count': poor_count, 'faculty': faculty, 'is_department_head': is_department_head, 'event_notifications': event_notifications,
        'messages_notifications': messages_notifications,
        'notifications_unread_count': notifications_unread_count,
        'messages_unread_count': messages_unread_count,})

@login_required(login_url='signin')
@allowed_users(allowed_roles=['head of OSAS','department head/program coordinator'])
def department_head_peer_to_peer_view_evaluation_form(request, pk):
    faculty_evaluation_form = get_object_or_404(PeertoPeerEvaluation, pk=pk)
    questions = FacultyEvaluationQuestions.objects.all().order_by('order')
    user = request.user
    is_department_head = request.user.groups.filter(name='department head/program coordinator').exists()
    faculty = Faculty.objects.filter(email=request.user.username).first()   
    event_notifications = Notification.objects.filter(recipient=user, level='success')
    messages_notifications = Notification.objects.filter(recipient=user, level='info')
    unread_notifications = Notification.objects.filter(recipient=user, level='success', unread=True)
    unread_messages = Notification.objects.filter(recipient=user, level='info', unread=True)

    notifications_unread_count = unread_notifications.count()
    messages_unread_count = unread_messages.count()
         # List of all rating fields to check
    rating_fields = [
        'command_and_knowledge_of_the_subject',
        'depth_of_mastery',
        'practice_in_respective_discipline',
        'up_to_date_knowledge',
        'integrates_subject_to_practical_circumstances',
        'organizes_the_subject_matter',
        'provides_orientation_on_course_content',
        'efforts_of_class_preparation',
        'summarizes_main_points',
        'monitors_online_class',
        'holds_interest_of_students',
        'provides_relevant_feedback',
        'encourages_participation',
        'shows_enthusiasm',
        'shows_sense_of_humor',
        'teaching_methods',
        'flexible_learning_strategies',
        'student_engagement',
        'clear_examples',
        'focused_on_objectives',
        'starts_with_motivating_activities',
        'speaks_in_clear_and_audible_manner',
        'uses_appropriate_medium_of_instruction',
        'establishes_online_classroom_environment',
        'observes_proper_classroom_etiquette',
        'uses_time_wisely',
        'gives_ample_time_for_students_to_prepare',
        'updates_the_students_of_their_progress',
        'demonstrates_leadership_and_professionalism',
        'understands_possible_distractions',
        'sensitivity_to_student_culture',
        'responds_appropriately',
        'assists_students_on_concerns',
        'guides_the_students_in_accomplishing_tasks',
        'extends_consideration_to_students'
    ]
    # Initialize counters
    outstanding_count = 0
    very_satisfactory_count = 0
    satisfactory_count = 0
    unsatisfactory_count = 0
    poor_count = 0

    # Calculate counts in memory
    for field in rating_fields:
        value = getattr(faculty_evaluation_form, field)
        if value == 5:
            outstanding_count += 1
        elif value == 4:
            very_satisfactory_count += 1
        elif value == 3:
            satisfactory_count += 1
        elif value == 2:
            unsatisfactory_count += 1
        elif value == 1:
            poor_count += 1

    return render(request, 'pages/department_head_peer_to_peer_view_evaluation_form.html', {'faculty_evaluation_form': faculty_evaluation_form, 'faculty': faculty, 'questions': questions, 'outstanding_count': outstanding_count, 'very_satisfactory_count': very_satisfactory_count, 'satisfactory_count': satisfactory_count, 'unsatisfactory_count': unsatisfactory_count, 'poor_count': poor_count, 'faculty': faculty, 'is_department_head': is_department_head, 'event_notifications': event_notifications,
        'messages_notifications': messages_notifications,
        'notifications_unread_count': notifications_unread_count,
        'messages_unread_count': messages_unread_count,})


def facultylogout(request):
    logout(request)
    return redirect('facultylogin')